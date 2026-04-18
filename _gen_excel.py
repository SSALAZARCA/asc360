"""Genera el archivo Excel template para importar Centros de Servicio / Distribuidoras.
Incluye los nuevos campos: capacidad_bahias, numero_tecnicos, ciudad, departamento, tipo_servicio.
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Centros de Servicio"

headers_display = [
    "Nombre del Centro / Distribuidora *",
    "NIT",
    "Teléfono",
    "Tipo *",
    "Ciudad",
    "Departamento",
    "Capacidad Bahías",
    "Núm. Técnicos",
    "Tipo de Servicio *",
    "Subdominio (opcional)",
]

header_fill = PatternFill(start_color="1A3C5E", end_color="1A3C5E", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin = Side(style="thin")
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws.row_dimensions[1].height = 38
for col_idx, h in enumerate(headers_display, start=1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border

examples = [
    ("Motos Sur Bogotá", "900.123.456-7", "601-7234567", "Centro de Servicio", "Bogotá", "Cundinamarca", 4, 6, "Todos", ""),
    ("MotoPlus Medellín", "800.987.654-3", "604-5671234", "Centro de Servicio", "Medellín", "Antioquia", 3, 4, "Todos", ""),
    ("Express Moto Cali", "830.221.999-1", "602-5563890", "Centro de Servicio", "Cali", "Valle", 2, 2, "Revisiones/Express", ""),
    ("Repuestos Colombia SAS", "901.555.111-0", "604-4441234", "Distribuidor / Repuestero", "Medellín", "Antioquia", None, None, "Todos", ""),
]

fill_even = PatternFill(start_color="EAF4FB", end_color="EAF4FB", fill_type="solid")
fill_odd  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
data_font = Font(size=10, name="Calibri")

for row_idx, ex in enumerate(examples, start=2):
    fill = fill_even if row_idx % 2 == 0 else fill_odd
    for col_idx, val in enumerate(ex, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.fill = fill
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")

# Leyenda
ws.cell(row=7, column=1, value="TIPOS DE CENTRO:").font = Font(bold=True, color="1A3C5E", size=9)
ws.cell(row=8, column=1, value="  •  Centro de Servicio").font = Font(italic=True, size=9, color="555555")
ws.cell(row=9, column=1, value="  •  Distribuidor / Repuestero").font = Font(italic=True, size=9, color="555555")
ws.cell(row=7, column=9, value="TIPO DE SERVICIO:").font = Font(bold=True, color="1A3C5E", size=9)
ws.cell(row=8, column=9, value="  •  Todos").font = Font(italic=True, size=9, color="555555")
ws.cell(row=9, column=9, value="  •  Revisiones/Express").font = Font(italic=True, size=9, color="555555")

col_widths = [42, 18, 16, 26, 16, 18, 16, 14, 20, 22]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"

out = r"c:\proyectos IA\UM Colombia\Aplicación red de servicio\centros_de_servicio.xlsx"
wb.save(out)
print("Excel creado:", out)
