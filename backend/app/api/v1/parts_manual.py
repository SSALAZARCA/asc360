import base64
import io
import json
import logging
import os
import re
import tempfile
import uuid as _uuid
from datetime import datetime
from pathlib import Path
from typing import Literal, Optional

import fitz  # PyMuPDF
import pdfplumber
from fastapi import APIRouter, Depends, File, Form, Header, HTTPException, UploadFile
from fastapi.responses import Response, StreamingResponse
from minio import Minio
from pydantic import BaseModel, Field
from sqlalchemy import delete as sa_delete, update as sa_update, text, exists as sa_exists, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from openai import AsyncOpenAI

from app.api.deps import get_current_user, get_optional_user
from app.core.security import verify_sonia_secret
from app.config import settings
from app.database import get_db, async_session_maker
from app.models.order import ServiceOrder
from app.models.vehicle import Vehicle
from app.models.imports import VehicleModel, SparePartItem
from app.models.logistics import PartCatalog
from app.models.parts_manual import (
    PartsManualItem, PartsManualSection, PartsReference, VehicleCatalogMap,
    PartsCodeReviewTask, PartSubstitute, PartCostHistory, PartOrderDecision,
    PartsManualSectionHistory,
)
from app.models.system_config import SystemConfig
from app.services.pricing_service import get_pricing_factors, compute_prices

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/parts", tags=["Parts Manual"])

_openai = AsyncOpenAI(api_key=settings.OPENAI_API_KEY)

PARTS_BUCKET = "parts-manuals"

# Solo usado por replace_section_diagram (subida directa, sin re-procesar).
_DIAGRAM_ALLOWED_MIME_TYPES = {"image/png", "image/jpeg", "image/webp"}


# ── Schemas ────────────────────────────────────────────────────────────────────

class PartsSearchRequest(BaseModel):
    order_id: str
    description: str

class PartLookupResult(BaseModel):
    section_id: str
    section_code: str
    section_name: str
    diagram_url: Optional[str]
    model_code: Optional[str] = None


class PartsByModelRequest(BaseModel):
    model_code: str
    description: str


class PartItemByCodeResult(BaseModel):
    factory_part_number: str
    description: str
    description_es: Optional[str] = None
    section_code: str
    section_name: str
    order_num: str

class PartItemResult(BaseModel):
    """Shared by `get_part_by_number` (single position, INNER JOIN + 404) and
    `get_all_items_for_section` (full section, LEFT JOIN, never drops a row).
    Reference-derived fields are Optional because of that LEFT JOIN: a numbered
    position with no catalog reference still occupies its slot so the list
    always matches the diagram legend (design ADR 3/6). Nulls stay nulls --
    "N/D" and "Sin precio" are the frontend's job, never the API's (ADR 8).

    `precio_publico` is visible to EVERY authenticated caller by design
    (no per-screen or per-role gating anywhere): both `/distribuidor/repuestos`
    and `/tg/parts` read this same shape."""
    id: str
    section_id: str
    section_code: str
    section_name: str
    order_num: str
    factory_part_number: str
    um_part_number: Optional[str] = None
    description: Optional[str] = None
    description_es: Optional[str] = None
    unit: Optional[str] = None
    precio_publico: Optional[float] = None
    precio_es_preliminar: bool = False
    # "disponible" | "ingresando" | "sin_stock" | None (not computed by this
    # endpoint -- `get_part_by_number` leaves it None, only
    # `get_all_items_for_section` resolves it today, ADR 8's "nulls stay
    # nulls" applies here too).
    inventory_status: Optional[str] = None

class PartReferenceResult(BaseModel):
    factory_part_number: str
    um_part_number: str
    description: str
    unit: Optional[str]

class LoadSectionResult(BaseModel):
    section_code: str
    section_name: str
    diagram_url: Optional[str]
    parts_loaded: int
    references_new: int


# ── Helpers internos ───────────────────────────────────────────────────────────

async def _classify_sections(description: str, sections: list[dict]) -> list[str]:
    sections_block = "\n".join(
        f"- {s['section_code']}: {s['section_name']}" for s in sections
    )
    system = (
        "Eres un experto en repuestos de motocicletas UM Colombia. "
        "El técnico describe en español la parte que necesita. "
        "Identifica las 2 o 3 secciones del catálogo de despiece que MÁS PROBABLEMENTE "
        "contienen esa parte. Devuelve ÚNICAMENTE un JSON: {\"codes\": [\"B2\", \"E11\"]}"
    )
    user = f"Descripción del técnico: {description}\n\nSecciones disponibles:\n{sections_block}"
    try:
        resp = await _openai.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "system", "content": system}, {"role": "user", "content": user}],
            response_format={"type": "json_object"},
            max_tokens=80,
            temperature=0,
        )
        data = json.loads(resp.choices[0].message.content)
        return data.get("codes", [])[:3]
    except Exception as e:
        logger.error(f"_classify_sections error: {e}")
        return []


def _parse_section_filename(filename: str) -> tuple[str, str]:
    """
    'RENEGADE 200 SPORT_B1_FRAME.pdf'        → ('B1',  'FRAME')
    'RENEGADE 200 SPORT_B12_REAR FENDER_REAR TURN SIGNAL.pdf' → ('B12', 'REAR FENDER / REAR TURN SIGNAL')
    'B01_BODY COMP FRAME.pdf'                → ('B01', 'BODY COMP FRAME')
    'B01 - AIR CLEANER ASSY.pdf'             → ('B01', 'AIR CLEANER ASSY')
    'B04 FOOTREST & PEDALS.pdf'              → ('B04', 'FOOTREST & PEDALS')
    """
    stem = Path(filename).stem
    parts = [p.strip() for p in stem.split("_") if p.strip()]
    if len(parts) >= 3:
        # Formato: MODELO_CODIGO_DESCRIPCION[_DESCRIPCION...]
        code = parts[1]
        name = " / ".join(parts[2:])
    elif len(parts) == 2:
        # Formato simple: CODIGO_DESCRIPCION
        code = parts[0]
        name = parts[1]
    else:
        # Formato "B01 - DESCRIPCION" o "B04 DESCRIPCION" (sin guiones bajos, guion opcional)
        m = re.match(r'^([A-Z]\d+)\s*-*\s*(.+)$', stem, re.IGNORECASE)
        if m:
            code = m.group(1).strip().upper()
            name = m.group(2).strip()
        else:
            code = stem
            name = stem
    return code, name


_HEADER_KEYWORDS = {
    "order_num":   ["page", "no.", "no ", "item", "pos", "orden"],
    "factory":     ["factory", "part no", "part num", "code", "bom"],
    "um":          ["um part", "um no"],
    "description": ["description", "descrip"],
    "unit":        ["unit"],
}


def _detect_col(header_row: list, keywords: list, start: int = 0) -> int:
    for i in range(start, len(header_row)):
        cell = header_row[i]
        if cell is None:
            continue
        cell_lower = str(cell).lower().strip()
        for kw in keywords:
            if kw in cell_lower:
                return i
    return -1


def _find_col_groups(header_row: list) -> list[dict]:
    """Return all column groups (order_num, factory, …) found in a header row."""
    groups = []
    row_lower = [str(c).lower().strip() if c else "" for c in header_row]
    i = 0
    while i < len(row_lower):
        cell = row_lower[i]
        if cell and (any(kw in cell for kw in ["no.", "no ", "n0.", "item", "pos", "orden"]) or cell == "n"):
            col_map = {field: _detect_col(header_row, kws, start=i) for field, kws in _HEADER_KEYWORDS.items()}
            col_map["order_num"] = i
            if col_map.get("factory", -1) > i:
                groups.append(col_map)
                i = max(v for v in col_map.values() if v > 0) + 1
                continue
        i += 1
    return groups


def _extract_tables_from_page(page) -> list[list[list]]:
    """Try default line-based extraction; fall back to text-alignment strategy."""
    tables = page.extract_tables() or []
    if tables:
        return tables
    for strategy in (
        {"vertical_strategy": "text", "horizontal_strategy": "lines"},
        {"vertical_strategy": "text", "horizontal_strategy": "text",
         "snap_tolerance": 5, "join_tolerance": 5, "text_tolerance": 5},
    ):
        try:
            t = page.extract_table(table_settings=strategy)
            if t and len(t) >= 2:
                return [t]
        except Exception:
            pass
    return []


def _parse_parts_from_text(text: str) -> list[dict]:
    """Fallback for PDFs without structured tables (plain-text layout)."""
    import re as _re
    parts = []
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    header_idx = -1
    for i, line in enumerate(lines):
        low = line.lower()
        if ("no" in low or "n0" in low or "item" in low) and ("factory" in low or "bom" in low or "part" in low):
            header_idx = i
            break
    if header_idx < 0:
        return parts
    pattern = _re.compile(r'^(\S+)\s+(\S+)\s+(.+)$')
    skip = {"no.", "no", "item", "pos", "page"}
    for line in lines[header_idx + 1:]:
        m = pattern.match(line)
        if not m:
            continue
        order_num, factory, description = m.group(1), m.group(2), m.group(3).strip()
        if order_num.lower() in skip:
            continue
        if len(factory) < 3 or not _re.search(r'[A-Z0-9]', factory, _re.I):
            continue
        parts.append({
            "order_num":           order_num,
            "factory_part_number": factory,
            "um_part_number":      "",
            "description":         description,
            "unit":                None,
        })
    return parts


def _parse_parts_table(pdf_path: str) -> list[dict]:
    parts = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            page_parts: list[dict] = []
            for table in _extract_tables_from_page(page):
                if not table or len(table) < 2:
                    continue
                header_idx, col_groups = -1, []
                for row_i, row in enumerate(table):
                    if not row:
                        continue
                    row_lower = [str(c).lower() if c else "" for c in row]
                    hits = sum(
                        1 for kw_list in _HEADER_KEYWORDS.values()
                        for kw in kw_list if any(kw in cell for cell in row_lower)
                    )
                    if hits >= 3:
                        header_idx = row_i
                        col_groups = _find_col_groups(row)
                        break
                if header_idx < 0 or not col_groups:
                    continue
                skip = {"page", "no.", "no", "item", "pos", ""}
                for row in table[header_idx + 1:]:
                    if not row:
                        continue
                    for col_map in col_groups:
                        def get(field, _cm=col_map, _row=row):
                            idx = _cm.get(field, -1)
                            if idx < 0 or idx >= len(_row):
                                return None
                            v = _row[idx]
                            return str(v).strip() if v is not None else None

                        order_num = get("order_num")
                        factory   = get("factory")
                        if not order_num or not factory:
                            continue
                        if order_num.lower() in skip:
                            continue
                        page_parts.append({
                            "order_num":           order_num,
                            "factory_part_number": factory,
                            "um_part_number":      get("um") or "",
                            "description":         get("description") or "",
                            "unit":                get("unit"),
                        })

            if page_parts:
                parts.extend(page_parts)
            else:
                text = page.extract_text() or ""
                parts.extend(_parse_parts_from_text(text))
    return parts


def _extract_section_illustration(pdf_path: str) -> bytes:
    """Extracts the exploded-view illustration from a section PDF's first
    page as flat PNG bytes.

    Some source PDFs split the illustration across MULTIPLE embedded images
    stacked on the page (e.g. the main assembly plus a separate bracket/
    cover placed directly below it) instead of a single image. Extracting
    only the first embedded image (`imgs[0]`) silently drops the rest --
    confirmed against a real source PDF where a lower bracket's own
    separately-embedded image never made it into the generated card.

    Instead of re-extracting and stitching embedded images back together,
    this computes the union of every embedded image's on-page placement
    rectangle and RENDERS that exact page region as one flat image --
    correctly composited regardless of how many images (or vector graphics)
    the source PDF actually uses. Falls back to rendering the whole page
    when the page has no embedded images at all (pure vector content).
    """
    doc = fitz.open(pdf_path)
    try:
        page = doc[0]
        imgs = page.get_images(full=True)
        rects = [r for img in imgs for r in page.get_image_rects(img[0])]
        if rects:
            clip = rects[0]
            for r in rects[1:]:
                clip |= r
            pix = page.get_pixmap(matrix=fitz.Matrix(4, 4), clip=clip)
        else:
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
        return pix.tobytes("png")
    finally:
        doc.close()


_NAT = re.compile(r"^([A-Za-z]*)(\d*)")


def _natural_order_key(order_num: str) -> tuple:
    """`order_num` is `String(20)` holding values like A1 / B3 / C12, so plain
    lexicographic ordering yields A1, A10, A2. The whole point of the
    Distribuidor screen is a legend that reads in the same order as the
    numbers stamped on the diagram (design ADR 7). Sections are tens of rows
    -- sorting in Python is free."""
    m = _NAT.match(order_num or "")
    alpha, digits = (m.group(1), m.group(2)) if m else ("", "")
    return (alpha.upper(), int(digits) if digits else 0, order_num or "")


def _resolve_public_price(ref, manual_price, factors) -> tuple[Optional[float], bool]:
    """ONE price path for both parts-detail endpoints (design ADR 2).
    Reproduces Maestro de Partes' hierarchy EXACTLY (`_build_item` below):
      1. a manually set `PartCatalog.public_price` wins outright ("Precio Final");
      2. otherwise compute from `avg_fob_cost`, falling back to `preliminary_fob`;
      3. otherwise `None` -- never 0, never a guess.
    Whoever quotes this number -- Distribuidor or Técnico -- must say the same
    number the company's own price master says, or the quote is not honourable."""
    if manual_price is not None:
        return float(manual_price), False
    if ref is None:
        return None, False
    avg    = float(ref.avg_fob_cost)    if ref.avg_fob_cost    is not None else None
    prelim = float(ref.preliminary_fob) if ref.preliminary_fob is not None else None
    fob    = avg if avg is not None else prelim
    return compute_prices(fob, factors)["precio_publico"], (avg is None and prelim is not None)


def _extract_prev_codes(prev_codes_json) -> list[str]:
    extracted = []
    for entry in (prev_codes_json or []):
        if isinstance(entry, dict) and "code" in entry:
            extracted.append(entry["code"])
        elif isinstance(entry, str) and entry:
            extracted.append(entry)
    return extracted


async def _resolve_inventory_status(db: AsyncSession, fpn_and_refs: list[tuple[str, object]]) -> dict[str, str]:
    """Tri-state inventory badge per `factory_part_number` for the Distribuidor
    parts-search screen: 'disponible' (certified physical stock, some
    `spare_part_items` row has `qty_physical > 0`), 'ingresando' (received
    but not yet physically inspected -- `qty_received > 0` and
    `qty_physical IS NULL`), or 'sin_stock' (neither -- includes both "never
    declared" and "ordered but nothing physical yet", the latter folded in
    because this screen only distinguishes 3 states, unlike Maestro de
    Partes' 4-state `coverage_status` in `_list_catalog_impl` which this
    reuses the exact same scoring query shape from, for consistency).
    Matches by `factory_part_number` AND any historical `prev_codes` alias.
    ONE query for the whole request, batched, never per-row (ADR 2/5)."""
    code_to_fpn: dict[str, str] = {}
    for fpn, ref in fpn_and_refs:
        code_to_fpn[fpn.upper().strip().replace(' ', '')] = fpn
        if ref is not None:
            for prev in _extract_prev_codes(ref.prev_codes):
                norm = prev.upper().strip().replace(' ', '')
                code_to_fpn.setdefault(norm, fpn)

    if not code_to_fpn:
        return {}

    from app.models.imports import SparePartLot, ShipmentOrder as _SO
    from sqlalchemy import case as sa_case, and_ as sa_and, or_ as sa_or

    fn = func.upper(func.trim(func.replace(SparePartItem.part_number, ' ', '')))
    cov_q = (
        select(
            fn.label('code'),
            func.max(
                sa_case(
                    (sa_and(SparePartItem.qty_physical.isnot(None), SparePartItem.qty_physical > 0), 2),
                    (sa_and(
                        SparePartItem.qty_received > 0,
                        SparePartItem.qty_physical.is_(None),
                        sa_or(_SO.bl_container.isnot(None), SparePartLot.packing_list_received == True),
                    ), 1),
                    else_=0,
                )
            ).label('score'),
        )
        .select_from(SparePartItem)
        .join(SparePartLot, SparePartLot.id == SparePartItem.lot_id)
        .outerjoin(_SO, _SO.id == SparePartLot.shipment_order_id)
        .where(fn.in_(list(code_to_fpn.keys())))
        .group_by(fn)
    )
    score_to_status = {2: "disponible", 1: "ingresando", 0: "sin_stock"}
    rank = {"disponible": 2, "ingresando": 1, "sin_stock": 0}
    status: dict[str, str] = {}
    for code_val, score in (await db.execute(cov_q)).all():
        fpn = code_to_fpn.get(code_val, code_val)
        new_status = score_to_status.get(score or 0, "sin_stock")
        if fpn not in status or rank[new_status] > rank[status[fpn]]:
            status[fpn] = new_status
    return status


def _minio_client() -> Minio:
    return Minio(
        settings.MINIO_ENDPOINT,
        access_key=settings.MINIO_ACCESS_KEY,
        secret_key=settings.MINIO_SECRET_KEY,
        secure=settings.MINIO_SECURE,
    )


def _ensure_parts_bucket(client: Minio) -> None:
    if not client.bucket_exists(PARTS_BUCKET):
        client.make_bucket(PARTS_BUCKET)
        policy = json.dumps({
            "Version": "2012-10-17",
            "Statement": [{
                "Effect": "Allow", "Principal": "*",
                "Action": "s3:GetObject",
                "Resource": f"arn:aws:s3:::{PARTS_BUCKET}/*",
            }]
        })
        client.set_bucket_policy(PARTS_BUCKET, policy)


def _diagram_public_url(object_name: str) -> str:
    public_base = settings.MINIO_PUBLIC_URL or f"http://{settings.MINIO_ENDPOINT}"
    return f"{public_base}/{PARTS_BUCKET}/{object_name}"


# ── Endpoint de administración — listado de modelos ───────────────────────────

@router.get("/admin/vehicle-models")
async def list_vehicle_models(
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Devuelve el catálogo de modelos UM con su catalog_model_code si ya tiene secciones cargadas."""
    if not current_user.is_superadmin:
        raise HTTPException(status_code=403, detail="Solo superadmin")

    result = await db.execute(
        select(VehicleModel.model_name, VehicleCatalogMap.catalog_model_code)
        .outerjoin(VehicleCatalogMap, VehicleModel.model_name == VehicleCatalogMap.vehicle_model_pattern)
        .order_by(VehicleModel.model_name)
    )
    rows = result.all()
    return [
        {"vehicle_model": r[0], "catalog_model_code": r[1]}
        for r in rows if r[0]
    ]


# ── Endpoints de búsqueda (bot) ────────────────────────────────────────────────

@router.post("/search", response_model=list[PartLookupResult])
async def search_parts(
    body: PartsSearchRequest,
    db: AsyncSession = Depends(get_db),
    x_sonia_secret: str = Header(default=""),
):
    if not verify_sonia_secret(x_sonia_secret, settings.SONIA_BOT_SECRET):
        raise HTTPException(status_code=403, detail="Forbidden")

    result = await db.execute(
        select(ServiceOrder, Vehicle)
        .join(Vehicle, ServiceOrder.vehicle_id == Vehicle.id)
        .where(ServiceOrder.id == body.order_id)
    )
    row = result.first()
    if not row:
        raise HTTPException(status_code=404, detail="Orden no encontrada")

    vehicle_model = row[1].model or ""

    map_result = await db.execute(
        select(VehicleCatalogMap).where(VehicleCatalogMap.vehicle_model_pattern == vehicle_model)
    )
    catalog_map = map_result.scalar_one_or_none()
    if not catalog_map:
        raise HTTPException(
            status_code=404,
            detail=f"Catálogo no disponible para el modelo '{vehicle_model}'."
        )

    sections_result = await db.execute(
        select(PartsManualSection).where(PartsManualSection.model_code == catalog_map.catalog_model_code)
    )
    all_sections = sections_result.scalars().all()
    if not all_sections:
        raise HTTPException(status_code=404, detail="Sin secciones cargadas para este modelo")

    sections_list = [{"section_code": s.section_code, "section_name": s.section_name} for s in all_sections]
    matched_codes = await _classify_sections(body.description, sections_list)
    matched = [s for s in all_sections if s.section_code in matched_codes] or list(all_sections[:2])

    return [
        PartLookupResult(
            section_id=str(s.id),
            section_code=s.section_code,
            section_name=s.section_name,
            diagram_url=s.diagram_url,
            model_code=catalog_map.catalog_model_code,
        )
        for s in matched
    ]


# ── Endpoints de catálogo para bot (superadmin) ───────────────────────────────

@router.get("/bot/catalog-models")
async def bot_catalog_models(
    db: AsyncSession = Depends(get_db),
    x_sonia_secret: str = Header(default=""),
    current_user=Depends(get_optional_user),
):
    if not verify_sonia_secret(x_sonia_secret, settings.SONIA_BOT_SECRET) and current_user is None:
        raise HTTPException(status_code=403, detail="Forbidden")

    result = await db.execute(
        select(VehicleModel.model_name, VehicleCatalogMap.catalog_model_code)
        .join(VehicleCatalogMap, VehicleModel.model_name == VehicleCatalogMap.vehicle_model_pattern)
        .where(VehicleCatalogMap.catalog_model_code.isnot(None))
        .order_by(VehicleModel.model_name)
    )
    return [
        {"vehicle_model": r[0], "catalog_model_code": r[1]}
        for r in result.all() if r[0]
    ]


@router.post("/search-by-model", response_model=list[PartLookupResult])
async def search_parts_by_model(
    body: PartsByModelRequest,
    db: AsyncSession = Depends(get_db),
    x_sonia_secret: str = Header(default=""),
    current_user=Depends(get_optional_user),
):
    if not verify_sonia_secret(x_sonia_secret, settings.SONIA_BOT_SECRET) and current_user is None:
        raise HTTPException(status_code=403, detail="Forbidden")

    sections_result = await db.execute(
        select(PartsManualSection).where(PartsManualSection.model_code == body.model_code)
    )
    all_sections = sections_result.scalars().all()
    if not all_sections:
        raise HTTPException(status_code=404, detail="Sin secciones cargadas para este modelo")

    sections_list = [{"section_code": s.section_code, "section_name": s.section_name} for s in all_sections]
    matched_codes = await _classify_sections(body.description, sections_list)
    matched = [s for s in all_sections if s.section_code in matched_codes] or list(all_sections[:3])

    return [
        PartLookupResult(
            section_id=str(s.id),
            section_code=s.section_code,
            section_name=s.section_name,
            diagram_url=s.diagram_url,
            model_code=body.model_code,
        )
        for s in matched
    ]


@router.get("/model/{model_code}/all-sections")
async def get_all_sections_for_model(
    model_code: str,
    db: AsyncSession = Depends(get_db),
    x_sonia_secret: str = Header(default=""),
    current_user=Depends(get_optional_user),
):
    if not verify_sonia_secret(x_sonia_secret, settings.SONIA_BOT_SECRET) and current_user is None:
        raise HTTPException(status_code=403, detail="Forbidden")

    result = await db.execute(
        select(PartsManualSection)
        .where(PartsManualSection.model_code == model_code)
        .order_by(PartsManualSection.section_code)
    )
    return [
        {
            "section_id": str(s.id),
            "section_code": s.section_code,
            "section_name": s.section_name,
            "diagram_url": s.diagram_url,
        }
        for s in result.scalars().all()
    ]


@router.get("/model/{model_code}/item/{order_num}", response_model=PartItemByCodeResult)
async def get_part_by_model_and_code(
    model_code: str,
    order_num: str,
    db: AsyncSession = Depends(get_db),
    x_sonia_secret: str = Header(default=""),
):
    if not verify_sonia_secret(x_sonia_secret, settings.SONIA_BOT_SECRET):
        raise HTTPException(status_code=403, detail="Forbidden")

    result = await db.execute(
        select(PartsManualItem, PartsManualSection, PartsReference)
        .join(PartsManualSection, PartsManualItem.section_id == PartsManualSection.id)
        .join(PartsReference, PartsManualItem.factory_part_number == PartsReference.factory_part_number)
        .where(
            PartsManualSection.model_code == model_code,
            PartsManualItem.order_num == order_num,
        )
    )
    row = result.first()
    if not row:
        raise HTTPException(status_code=404, detail="Parte no encontrada")

    item, section, ref = row
    return PartItemByCodeResult(
        factory_part_number=item.factory_part_number,
        description=ref.description,
        description_es=ref.description_es_manual,
        section_code=section.section_code,
        section_name=section.section_name,
        order_num=item.order_num,
    )


@router.get("/section/{section_id}/diagram-image")
async def get_diagram_image(
    section_id: str,
    db: AsyncSession = Depends(get_db),
    x_sonia_secret: str = Header(default=""),
    current_user=Depends(get_optional_user),
):
    """Proxy: descarga la imagen de diagrama desde MinIO. Acepta Sonia secret o JWT."""
    if not verify_sonia_secret(x_sonia_secret, settings.SONIA_BOT_SECRET) and current_user is None:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        section_uuid = _uuid.UUID(section_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="section_id inválido")

    section = await db.get(PartsManualSection, section_uuid)
    if not section or not section.diagram_url:
        raise HTTPException(status_code=404, detail="Diagrama no encontrado")

    public_base = f"{settings.MINIO_PUBLIC_URL}/{PARTS_BUCKET}/"
    if not section.diagram_url.startswith(public_base):
        raise HTTPException(status_code=404, detail="URL de diagrama no reconocida")

    object_name = section.diagram_url[len(public_base):]
    try:
        client = _minio_client()
        minio_response = client.get_object(PARTS_BUCKET, object_name)
        data = minio_response.read()
        minio_response.close()
    except Exception as e:
        logger.error(f"get_diagram_image MinIO error: {e}")
        raise HTTPException(status_code=502, detail="No se pudo obtener el diagrama")

    content_type = "image/png" if object_name.lower().endswith(".png") else "image/jpeg"
    return Response(content=data, media_type=content_type)


@router.get("/section/{section_id}/item/{order_num}", response_model=PartItemResult)
async def get_part_by_number(
    section_id: str,
    order_num: str,
    db: AsyncSession = Depends(get_db),
    x_sonia_secret: str = Header(default=""),
    current_user=Depends(get_optional_user),
):
    """Búsqueda por posición exacta (`/tg/parts` y el bot de Sonia). A
    diferencia de `get_all_items_for_section`, sigue usando INNER JOIN y
    404 -- una posición sin `PartsReference` es un genuino "no encontrado"
    para una búsqueda de un único recurso (design ADR 6).

    Distributor Parts Search (PR2, design ADR 1/2): resuelve `precio_publico`
    y `description_es` por el mismo `_resolve_public_price` que usa el
    endpoint plural, para que ambos handlers NUNCA puedan mostrar un precio
    distinto para la misma parte. No hay gating por rol -- el campo de
    precio es visible para cualquier llamador autenticado (o Sonia)."""
    if not verify_sonia_secret(x_sonia_secret, settings.SONIA_BOT_SECRET) and current_user is None:
        raise HTTPException(status_code=403, detail="Forbidden")

    result = await db.execute(
        select(PartsManualItem, PartsManualSection, PartsReference)
        .join(PartsManualSection, PartsManualItem.section_id == PartsManualSection.id)
        .join(PartsReference, PartsManualItem.factory_part_number == PartsReference.factory_part_number)
        .where(PartsManualItem.section_id == section_id, PartsManualItem.order_num == order_num)
    )
    row = result.first()
    if not row:
        raise HTTPException(status_code=404, detail="Parte no encontrada")

    item, section, ref = row

    factors = await get_pricing_factors(db)                                     # ONCE per request
    catalog_result = await db.execute(
        select(PartCatalog).where(PartCatalog.part_code.in_([item.factory_part_number]))
    )                                                                           # ONCE per request
    manual_price = next(
        (c.public_price for c in catalog_result.scalars() if c.part_code == item.factory_part_number),
        None,
    )
    precio_publico, precio_es_preliminar = _resolve_public_price(ref, manual_price, factors)

    return PartItemResult(
        id=str(item.id),
        section_id=str(section.id),
        section_code=section.section_code,
        section_name=section.section_name,
        order_num=item.order_num,
        factory_part_number=item.factory_part_number,
        um_part_number=ref.um_part_number,
        description=ref.description,
        description_es=ref.description_es_manual,
        unit=ref.unit,
        precio_publico=precio_publico,
        precio_es_preliminar=precio_es_preliminar,
    )


@router.get("/section/{section_id}/items", response_model=list[PartItemResult])
async def get_all_items_for_section(
    section_id: str,
    db: AsyncSession = Depends(get_db),
    x_sonia_secret: str = Header(default=""),
    current_user=Depends(get_optional_user),
):
    """Lista TODOS los repuestos numerados de una sección (para la vista de
    Distribuidor que muestra el diagrama junto con la lista completa de
    posiciones, en vez de buscar una sola posición por código exacto como
    `get_part_by_number`). Sección vacía o `section_id` inexistente → lista
    vacía, no 404 (es un listado, no una búsqueda de un único recurso).

    LEFT JOIN on `PartsReference` (design ADR 6): a numbered position never
    disappears from the list just because it has no catalog reference.
    Results are natural-sorted by `order_num` in Python (ADR 7), not via SQL
    `ORDER BY`. `precio_publico` is resolved once per row via the shared
    `_resolve_public_price` path (ADR 2), fed by exactly one
    `get_pricing_factors(db)` call and one batched `PartCatalog` lookup for
    the whole request (ADR 5) — never per row. `inventory_status` is resolved
    the same way, via one batched `_resolve_inventory_status` call.
    `description_es` falls back to the import-derived `SparePartItem.
    description_es` (via a batched `spi_description_es` lookup) whenever
    `PartsReference.description_es_manual` was never filled in -- same
    `COALESCE(description_es_manual, spi.description_es)` priority Maestro
    de Partes' `_list_catalog_impl` already uses, so a translation that only
    ever came from the packing list still shows up here."""
    if not verify_sonia_secret(x_sonia_secret, settings.SONIA_BOT_SECRET) and current_user is None:
        raise HTTPException(status_code=403, detail="Forbidden")

    result = await db.execute(
        select(PartsManualItem, PartsManualSection, PartsReference)
        .join(PartsManualSection, PartsManualItem.section_id == PartsManualSection.id)
        .outerjoin(PartsReference, PartsManualItem.factory_part_number == PartsReference.factory_part_number)
        .where(PartsManualItem.section_id == section_id)
    )
    rows = result.all()
    rows.sort(key=lambda r: _natural_order_key(r[0].order_num))

    factors = await get_pricing_factors(db)                                     # ONCE per request

    codes = [item.factory_part_number for item, _, _ in rows]
    manual_prices: dict[str, float] = {}
    spi_description_es: dict[str, str] = {}
    if codes:
        catalog_result = await db.execute(
            select(PartCatalog).where(PartCatalog.part_code.in_(codes))
        )                                                                       # ONCE per request
        manual_prices = {c.part_code: c.public_price for c in catalog_result.scalars()}

        # First non-empty import-derived translation per part_number -- same
        # `spi_latest` shape/ordering as `_list_catalog_impl` (design ADR:
        # description_es fallback), just filtered to this section's codes.
        spi_desc_result = await db.execute(
            select(SparePartItem.part_number, SparePartItem.description_es)
            .where(
                SparePartItem.part_number.in_(codes),
                SparePartItem.description_es.isnot(None),
                SparePartItem.description_es != "",
            )
            .distinct(SparePartItem.part_number)
            .order_by(SparePartItem.part_number, SparePartItem.created_at.asc())
        )                                                                       # ONCE per request
        spi_description_es = dict(spi_desc_result.all())

    inventory_status = await _resolve_inventory_status(
        db, [(item.factory_part_number, ref) for item, _, ref in rows]
    )                                                                           # ONCE per request

    items = []
    for item, section, ref in rows:
        precio_publico, precio_es_preliminar = _resolve_public_price(
            ref, manual_prices.get(item.factory_part_number), factors
        )
        description_es = (ref.description_es_manual if ref else None) or spi_description_es.get(item.factory_part_number)
        items.append(PartItemResult(
            id=str(item.id),
            section_id=str(section.id),
            section_code=section.section_code,
            section_name=section.section_name,
            order_num=item.order_num,
            factory_part_number=item.factory_part_number,
            um_part_number=ref.um_part_number if ref else None,
            description=ref.description if ref else None,
            description_es=description_es,
            unit=ref.unit if ref else None,
            precio_publico=precio_publico,
            precio_es_preliminar=precio_es_preliminar,
            inventory_status=inventory_status.get(item.factory_part_number, "sin_stock"),
        ))
    return items


@router.get("/factory/{factory_code}", response_model=PartReferenceResult)
async def get_part_by_factory_code(
    factory_code: str,
    db: AsyncSession = Depends(get_db),
    x_sonia_secret: str = Header(default=""),
    current_user=Depends(get_optional_user),
):
    if not verify_sonia_secret(x_sonia_secret, settings.SONIA_BOT_SECRET) and current_user is None:
        raise HTTPException(status_code=403, detail="Forbidden")

    result = await db.execute(
        select(PartsReference).where(PartsReference.factory_part_number == factory_code)
    )
    ref = result.scalar_one_or_none()
    if not ref:
        raise HTTPException(status_code=404, detail="Código de fábrica no encontrado")

    return PartReferenceResult(
        factory_part_number=ref.factory_part_number,
        um_part_number=ref.um_part_number,
        description=ref.description,
        unit=ref.unit,
    )


# ── Detección de cambios de código ────────────────────────────────────────────

async def _detect_code_changes(db: AsyncSession) -> int:
    """Detecta posibles cambios de código de fábrica usando similitud de descripción (pg_trgm).
    Crea tareas pendientes en parts_code_review_tasks. Inocuo si se ejecuta varias veces."""
    threshold_record = await db.get(SystemConfig, "parts_similarity_threshold")
    threshold = float(threshold_record.value) if threshold_record else 0.9

    result = await db.execute(text("""
        INSERT INTO parts_code_review_tasks
            (id, existing_code, candidate_code, existing_description, candidate_description,
             similarity_score, status, created_at)
        SELECT DISTINCT ON (spi.part_number, pr.factory_part_number)
            gen_random_uuid(),
            pr.factory_part_number,
            spi.part_number,
            pr.description,
            spi.description,
            similarity(spi.description, pr.description),
            'pending',
            now()
        FROM (
            SELECT DISTINCT ON (part_number) part_number, description, model_applicable
            FROM spare_part_items
            WHERE description IS NOT NULL AND description != ''
              AND model_applicable IS NOT NULL AND model_applicable != ''
            ORDER BY part_number, created_at DESC
        ) spi
        -- Restringir la comparación al mismo modelo de moto
        JOIN vehicle_catalog_map vcm ON vcm.vehicle_model_pattern = spi.model_applicable
        JOIN parts_manual_sections pms ON pms.model_code = vcm.catalog_model_code
        JOIN parts_manual_items pmi ON pmi.section_id = pms.id
        JOIN parts_references pr ON pr.factory_part_number = pmi.factory_part_number
        WHERE similarity(spi.description, pr.description) >= :threshold
          AND spi.part_number != pr.factory_part_number
          -- El candidato no existe ya como código activo en el catálogo con secciones asignadas
          AND NOT EXISTS (
              SELECT 1 FROM parts_references pr2
              JOIN parts_manual_items pmi2 ON pmi2.factory_part_number = pr2.factory_part_number
              WHERE pr2.factory_part_number = spi.part_number
          )
          -- El candidato no es un código previo ya conocido de esta parte (formato dict nuevo o string plano legado)
          AND NOT (
              pr.prev_codes @> to_jsonb(spi.part_number::text)
              OR pr.prev_codes @> jsonb_build_array(jsonb_build_object('code', spi.part_number))
          )
          -- No re-detectar el mismo par si ya está pendiente o rechazado
          AND NOT EXISTS (
              SELECT 1 FROM parts_code_review_tasks t
              WHERE t.candidate_code = spi.part_number
                AND t.existing_code = pr.factory_part_number
                AND t.status IN ('pending', 'rejected')
          )
          -- No agregar un segundo candidato mientras ya hay uno pendiente para este código
          AND NOT EXISTS (
              SELECT 1 FROM parts_code_review_tasks t2
              WHERE t2.existing_code = pr.factory_part_number
                AND t2.status = 'pending'
          )
        ORDER BY spi.part_number, pr.factory_part_number
    """), {"threshold": threshold})
    await db.commit()
    return result.rowcount


async def run_detection_bg() -> None:
    """Wrapper para ejecutar detección en segundo plano con sesión propia."""
    async with async_session_maker() as db:
        try:
            await _detect_code_changes(db)
        except Exception as e:
            logger.error(f"run_detection_bg error: {e}")


# ── Endpoint de consulta — tabla de repuestos cargados ────────────────────────

class PartSubstituteOut(BaseModel):
    substitute_part_code: str
    brand: str
    model: str
    position: int

class PartSubstituteIn(BaseModel):
    substitute_part_code: str
    brand: str
    model: str

class CatalogItemResult(BaseModel):
    factory_part_number: str
    description: str
    description_es: Optional[str]
    public_price: Optional[float]
    section_code: str
    section_name: str
    vehicle_model_name: Optional[str]
    pending_task_id: Optional[str] = None
    pending_candidate_code: Optional[str] = None
    pending_score: Optional[float] = None
    avg_fob_cost: Optional[float] = None
    preliminary_fob: Optional[float] = None
    fob_is_preliminary: bool = False          # True si el precio viene de PI, no de packing list
    costo_importado: Optional[float] = None
    precio_distribuidor: Optional[float] = None
    precio_publico_calculado: Optional[float] = None
    substitutes: list[PartSubstituteOut] = []
    rotation_class: Optional[str] = None
    prev_codes: list[str] = []
    needs_price_review: bool = False
    coverage_status: Optional[str] = None

class CatalogItemUpdate(BaseModel):
    description: Optional[str] = None
    description_es_manual: Optional[str] = None
    public_price: Optional[float] = None
    substitutes: Optional[list[PartSubstituteIn]] = None
    rotation_class: Optional[Literal['alta', 'media', 'baja']] = None
    prev_codes: Optional[list[str]] = None
    needs_price_review: Optional[bool] = None

class ReplaceCodeRequest(BaseModel):
    new_code: str
    description: Optional[str] = None
    description_es_manual: Optional[str] = None
    public_price: Optional[float] = None

class CatalogListResult(BaseModel):
    total: int
    items: list[CatalogItemResult]


@router.get("/admin/catalog", response_model=CatalogListResult)
async def list_catalog(
    search: str = "",
    model_code: str = "",
    only_pending: bool = False,
    only_price_review: bool = False,
    rotation_class: str = "",
    coverage_status: str = "",
    sort_col: str = "section_code",
    sort_dir: str = "asc",
    page: int = 1,
    page_size: int = 50,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Devuelve todos los repuestos cargados con su sección, modelo y datos del catálogo interno."""
    if not current_user.is_superadmin:
        raise HTTPException(status_code=403, detail="Solo superadmin")

    try:
        return await _list_catalog_impl(
            search, model_code, only_pending, only_price_review,
            rotation_class, coverage_status, sort_col, sort_dir, page, page_size, db,
        )
    except HTTPException:
        raise
    except Exception:
        logger.exception(
            "list_catalog failed: search=%r model=%r page=%d sort=%s/%s",
            search, model_code, page, sort_col, sort_dir,
        )
        raise HTTPException(status_code=500, detail="Error interno al consultar el catálogo")


async def _list_catalog_impl(
    search: str,
    model_code: str,
    only_pending: bool,
    only_price_review: bool,
    rotation_class: str,
    coverage_status: str,
    sort_col: str,
    sort_dir: str,
    page: int,
    page_size: int,
    db: AsyncSession,
) -> "CatalogListResult":
    pricing_factors = await get_pricing_factors(db)

    from sqlalchemy import func, or_

    # Subquery: primera descripción ES por part_number — la primera que llegó queda fija
    spi_latest = (
        select(
            SparePartItem.part_number.label("part_number"),
            SparePartItem.description_es.label("description_es"),
        )
        .where(SparePartItem.description_es.isnot(None))
        .where(SparePartItem.description_es != "")
        .distinct(SparePartItem.part_number)
        .order_by(SparePartItem.part_number, SparePartItem.created_at.asc())
        .subquery("spi_latest")
    )

    # Subquery: tarea pendiente más reciente por existing_code
    pending_sq = (
        select(
            PartsCodeReviewTask.existing_code.label("existing_code"),
            PartsCodeReviewTask.id.label("task_id"),
            PartsCodeReviewTask.candidate_code.label("candidate_code"),
            PartsCodeReviewTask.similarity_score.label("score"),
        )
        .where(PartsCodeReviewTask.status == "pending")
        .distinct(PartsCodeReviewTask.existing_code)
        .order_by(PartsCodeReviewTask.existing_code, PartsCodeReviewTask.similarity_score.desc(), PartsCodeReviewTask.created_at.desc())
        .subquery("pending_sq")
    )

    def _base_joins(q):
        q = (q
            .join(PartsManualItem, PartsManualItem.factory_part_number == PartsReference.factory_part_number)
            .join(PartsManualSection, PartsManualSection.id == PartsManualItem.section_id)
            .outerjoin(VehicleCatalogMap, VehicleCatalogMap.catalog_model_code == PartsManualSection.model_code)
            .outerjoin(PartCatalog, PartCatalog.part_code == PartsReference.factory_part_number)
            .outerjoin(spi_latest, spi_latest.c.part_number == PartsReference.factory_part_number)
            .outerjoin(pending_sq, pending_sq.c.existing_code == PartsReference.factory_part_number)
        )
        if model_code:
            q = q.where(PartsManualSection.model_code == model_code)
        if search:
            term = f"%{search.strip()}%"
            q = q.where(or_(
                PartsReference.factory_part_number.ilike(term),
                PartsReference.description.ilike(term),
                PartsReference.description_es_manual.ilike(term),
                spi_latest.c.description_es.ilike(term),
                text("""EXISTS (
                    SELECT 1 FROM jsonb_array_elements(parts_references.prev_codes) AS _pc(elem)
                    WHERE CASE WHEN jsonb_typeof(_pc.elem) = 'string' THEN _pc.elem #>> '{}'
                               ELSE _pc.elem->>'code'
                          END ILIKE :search_term
                )""").bindparams(search_term=term),
            ))
        if only_pending:
            q = q.where(pending_sq.c.task_id.isnot(None))
        if only_price_review:
            q = q.where(PartsReference.needs_price_review == True)
        if rotation_class == "none":
            q = q.where(PartsReference.rotation_class.is_(None))
        elif rotation_class in ("alta", "media", "baja"):
            q = q.where(PartsReference.rotation_class == rotation_class)
        if coverage_status == "aqui":
            # Use spare_part_availability VIEW so dispatched units are excluded.
            # Also matches stock received under any prev_code of this reference.
            q = q.where(
                or_(
                    sa_exists(
                        select(text("1")).select_from(text("spare_part_availability spa")).where(
                            text(
                                "UPPER(TRIM(REPLACE(spa.part_number, ' ', ''))) = "
                                "UPPER(TRIM(REPLACE(parts_references.factory_part_number, ' ', ''))) "
                                "AND spa.qty_available > 0"
                            )
                        )
                    ),
                    text("""EXISTS (
                        SELECT 1 FROM spare_part_availability _spa_pc
                        JOIN jsonb_array_elements(parts_references.prev_codes) AS _pc(elem) ON true
                        WHERE UPPER(TRIM(REPLACE(
                            CASE WHEN jsonb_typeof(_pc.elem) = 'string' THEN _pc.elem #>> '{}'
                                 ELSE _pc.elem->>'code'
                            END, ' ', ''))) = UPPER(TRIM(REPLACE(_spa_pc.part_number, ' ', '')))
                        AND _spa_pc.qty_available > 0
                    )"""),
                )
            )
        elif coverage_status == "en_camino":
            from app.models.imports import SparePartLot, ShipmentOrder
            fpn_ec = func.upper(func.trim(func.replace(SparePartItem.part_number, ' ', '')))
            ref_ec = func.upper(func.trim(PartsReference.factory_part_number))
            q = q.where(
                or_(
                    sa_exists(
                        select(SparePartItem.id)
                        .join(SparePartLot, SparePartLot.id == SparePartItem.lot_id)
                        .join(ShipmentOrder, ShipmentOrder.id == SparePartLot.shipment_order_id)
                        .where(
                            fpn_ec == ref_ec,
                            SparePartItem.qty_received > 0,
                            SparePartItem.qty_physical.is_(None),
                            or_(
                                ShipmentOrder.bl_container.isnot(None),
                                SparePartLot.packing_list_received == True,
                            ),
                        )
                    ),
                    text("""EXISTS (
                        SELECT 1 FROM spare_part_items _spi_ec
                        JOIN spare_part_lots _spl_ec ON _spl_ec.id = _spi_ec.lot_id
                        JOIN shipment_orders _so_ec  ON _so_ec.id  = _spl_ec.shipment_order_id
                        JOIN jsonb_array_elements(parts_references.prev_codes) AS _pc(elem) ON true
                        WHERE UPPER(TRIM(REPLACE(
                            CASE WHEN jsonb_typeof(_pc.elem) = 'string' THEN _pc.elem #>> '{}'
                                 ELSE _pc.elem->>'code'
                            END, ' ', ''))) = UPPER(TRIM(REPLACE(_spi_ec.part_number, ' ', '')))
                        AND _spi_ec.qty_received > 0
                        AND _spi_ec.qty_physical IS NULL
                        AND (_so_ec.bl_container IS NOT NULL OR _spl_ec.packing_list_received = true)
                    )"""),
                )
            )
        elif coverage_status == "pedido":
            from app.models.imports import SparePartLot, ShipmentOrder
            _fpn_ped = func.upper(func.trim(func.replace(SparePartItem.part_number, ' ', '')))
            _ref_ped = func.upper(func.trim(PartsReference.factory_part_number))
            # EXISTS en pedido: código actual O prev_code
            q = q.where(or_(
                sa_exists(
                    select(SparePartItem.id)
                    .join(SparePartLot, SparePartLot.id == SparePartItem.lot_id)
                    .where(
                        _fpn_ped == _ref_ped,
                        or_(
                            SparePartLot.packing_list_received == False,
                            SparePartItem.status.in_(['BACKORDER', 'BACKORDER_PARCIAL']),
                        ),
                        SparePartItem.status != 'CANCELLED',
                    )
                ),
                text("""EXISTS (
                    SELECT 1 FROM spare_part_items _spi_p
                    JOIN spare_part_lots _spl_p ON _spl_p.id = _spi_p.lot_id
                    JOIN jsonb_array_elements(parts_references.prev_codes) AS _pc(elem) ON true
                    WHERE UPPER(TRIM(REPLACE(
                        CASE WHEN jsonb_typeof(_pc.elem) = 'string' THEN _pc.elem #>> '{}'
                             ELSE _pc.elem->>'code'
                        END, ' ', ''))) = UPPER(TRIM(REPLACE(_spi_p.part_number, ' ', '')))
                    AND (_spl_p.packing_list_received = false OR _spi_p.status IN ('BACKORDER','BACKORDER_PARCIAL'))
                    AND _spi_p.status != 'CANCELLED'
                )"""),
            ))
            # NOT en_camino: ni código actual ni prev_code
            q = q.where(~or_(
                sa_exists(
                    select(SparePartItem.id)
                    .join(SparePartLot, SparePartLot.id == SparePartItem.lot_id)
                    .join(ShipmentOrder, ShipmentOrder.id == SparePartLot.shipment_order_id)
                    .where(
                        _fpn_ped == _ref_ped,
                        SparePartItem.qty_received > 0,
                        SparePartItem.qty_physical.is_(None),
                        or_(
                            ShipmentOrder.bl_container.isnot(None),
                            SparePartLot.packing_list_received == True,
                        ),
                    )
                ),
                text("""EXISTS (
                    SELECT 1 FROM spare_part_items _spi_ec2
                    JOIN spare_part_lots _spl_ec2 ON _spl_ec2.id = _spi_ec2.lot_id
                    JOIN shipment_orders _so_ec2  ON _so_ec2.id  = _spl_ec2.shipment_order_id
                    JOIN jsonb_array_elements(parts_references.prev_codes) AS _pc(elem) ON true
                    WHERE UPPER(TRIM(REPLACE(
                        CASE WHEN jsonb_typeof(_pc.elem) = 'string' THEN _pc.elem #>> '{}'
                             ELSE _pc.elem->>'code'
                        END, ' ', ''))) = UPPER(TRIM(REPLACE(_spi_ec2.part_number, ' ', '')))
                    AND _spi_ec2.qty_received > 0
                    AND _spi_ec2.qty_physical IS NULL
                    AND (_so_ec2.bl_container IS NOT NULL OR _spl_ec2.packing_list_received = true)
                )"""),
            ))
            # NOT aqui: ni código actual ni prev_code
            q = q.where(~or_(
                sa_exists(
                    select(SparePartItem.id).where(
                        _fpn_ped == _ref_ped,
                        SparePartItem.qty_physical.isnot(None),
                        SparePartItem.qty_physical > 0,
                    )
                ),
                text("""EXISTS (
                    SELECT 1 FROM spare_part_availability _spa_p2
                    JOIN jsonb_array_elements(parts_references.prev_codes) AS _pc(elem) ON true
                    WHERE UPPER(TRIM(REPLACE(
                        CASE WHEN jsonb_typeof(_pc.elem) = 'string' THEN _pc.elem #>> '{}'
                             ELSE _pc.elem->>'code'
                        END, ' ', ''))) = UPPER(TRIM(REPLACE(_spa_p2.part_number, ' ', '')))
                    AND _spa_p2.qty_available > 0
                )"""),
            ))
        elif coverage_status == "no_pedidas":
            from app.models.imports import SparePartLot, ShipmentOrder
            fpn_expr = func.upper(func.trim(func.replace(SparePartItem.part_number, ' ', '')))
            ref_expr = func.upper(func.trim(PartsReference.factory_part_number))
            # NOT aqui: ni código actual ni prev_code
            q = q.where(~or_(
                sa_exists(
                    select(SparePartItem.id).where(fpn_expr == ref_expr, SparePartItem.qty_physical.isnot(None), SparePartItem.qty_physical > 0)
                ),
                text("""EXISTS (
                    SELECT 1 FROM spare_part_availability _spa_np
                    JOIN jsonb_array_elements(parts_references.prev_codes) AS _pc(elem) ON true
                    WHERE UPPER(TRIM(REPLACE(
                        CASE WHEN jsonb_typeof(_pc.elem) = 'string' THEN _pc.elem #>> '{}'
                             ELSE _pc.elem->>'code'
                        END, ' ', ''))) = UPPER(TRIM(REPLACE(_spa_np.part_number, ' ', '')))
                    AND _spa_np.qty_available > 0
                )"""),
            ))
            # NOT en_camino (BL o packing list): ni código actual ni prev_code
            q = q.where(~or_(
                sa_exists(
                    select(SparePartItem.id)
                    .join(SparePartLot, SparePartLot.id == SparePartItem.lot_id)
                    .join(ShipmentOrder, ShipmentOrder.id == SparePartLot.shipment_order_id)
                    .where(
                        fpn_expr == ref_expr,
                        SparePartItem.qty_received > 0,
                        SparePartItem.qty_physical.is_(None),
                        or_(ShipmentOrder.bl_container.isnot(None), SparePartLot.packing_list_received == True),
                    )
                ),
                text("""EXISTS (
                    SELECT 1 FROM spare_part_items _spi_np
                    JOIN spare_part_lots _spl_np ON _spl_np.id = _spi_np.lot_id
                    JOIN shipment_orders _so_np  ON _so_np.id  = _spl_np.shipment_order_id
                    JOIN jsonb_array_elements(parts_references.prev_codes) AS _pc(elem) ON true
                    WHERE UPPER(TRIM(REPLACE(
                        CASE WHEN jsonb_typeof(_pc.elem) = 'string' THEN _pc.elem #>> '{}'
                             ELSE _pc.elem->>'code'
                        END, ' ', ''))) = UPPER(TRIM(REPLACE(_spi_np.part_number, ' ', '')))
                    AND _spi_np.qty_received > 0
                    AND _spi_np.qty_physical IS NULL
                    AND (_so_np.bl_container IS NOT NULL OR _spl_np.packing_list_received = true)
                )"""),
            ))
            # NOT en_camino (catalog price branch)
            q = q.where(PartCatalog.public_price.is_(None))
            # NOT pedido (sin packing list): ni código actual ni prev_code
            q = q.where(~or_(
                sa_exists(
                    select(SparePartItem.id)
                    .join(SparePartLot, SparePartLot.id == SparePartItem.lot_id)
                    .where(fpn_expr == ref_expr, SparePartLot.packing_list_received == False, SparePartItem.status != 'CANCELLED')
                ),
                text("""EXISTS (
                    SELECT 1 FROM spare_part_items _spi_np2
                    JOIN spare_part_lots _spl_np2 ON _spl_np2.id = _spi_np2.lot_id
                    JOIN jsonb_array_elements(parts_references.prev_codes) AS _pc(elem) ON true
                    WHERE UPPER(TRIM(REPLACE(
                        CASE WHEN jsonb_typeof(_pc.elem) = 'string' THEN _pc.elem #>> '{}'
                             ELSE _pc.elem->>'code'
                        END, ' ', ''))) = UPPER(TRIM(REPLACE(_spi_np2.part_number, ' ', '')))
                    AND _spl_np2.packing_list_received = false
                    AND _spi_np2.status != 'CANCELLED'
                )"""),
            ))
            # NOT backorder: ni código actual ni prev_code
            q = q.where(~or_(
                sa_exists(
                    select(SparePartItem.id)
                    .join(SparePartLot, SparePartLot.id == SparePartItem.lot_id)
                    .where(fpn_expr == ref_expr, SparePartItem.status.in_(['BACKORDER', 'BACKORDER_PARCIAL']))
                ),
                text("""EXISTS (
                    SELECT 1 FROM spare_part_items _spi_np3
                    JOIN spare_part_lots _spl_np3 ON _spl_np3.id = _spi_np3.lot_id
                    JOIN jsonb_array_elements(parts_references.prev_codes) AS _pc(elem) ON true
                    WHERE UPPER(TRIM(REPLACE(
                        CASE WHEN jsonb_typeof(_pc.elem) = 'string' THEN _pc.elem #>> '{}'
                             ELSE _pc.elem->>'code'
                        END, ' ', ''))) = UPPER(TRIM(REPLACE(_spi_np3.part_number, ' ', '')))
                    AND _spi_np3.status IN ('BACKORDER', 'BACKORDER_PARCIAL')
                )"""),
            ))
        return q

    # Total — cuenta pares únicos (parte, modelo)
    count_q = select(func.count()).select_from(
        _base_joins(
            select(PartsReference.factory_part_number, PartsManualSection.model_code)
        ).distinct().subquery()
    )
    total = (await db.execute(count_q)).scalar_one()

    # Filas — DISTINCT ON (factory_part_number, model_code): una fila por par parte+modelo
    from sqlalchemy import nullslast, cast
    from sqlalchemy import Numeric as SANumeric
    inner_sq = _base_joins(
        select(
            PartsReference.factory_part_number.label("fpn"),
            PartsReference.description.label("description"),
            func.coalesce(PartsReference.description_es_manual, spi_latest.c.description_es).label("description_es"),
            PartCatalog.public_price.label("public_price"),
            PartsManualSection.section_code.label("section_code"),
            PartsManualSection.section_name.label("section_name"),
            PartsManualSection.model_code.label("model_code"),
            VehicleCatalogMap.vehicle_model_pattern.label("vehicle_model_pattern"),
            pending_sq.c.task_id.label("task_id"),
            pending_sq.c.candidate_code.label("candidate_code"),
            pending_sq.c.score.label("score"),
            PartsReference.avg_fob_cost.label("avg_fob_cost"),           # r[11]
            PartsReference.rotation_class.label("rotation_class"),         # r[12]
            PartsReference.needs_price_review.label("needs_price_review"), # r[13]
            PartsReference.preliminary_fob.label("preliminary_fob"),       # r[14]
        )
        .distinct(PartsReference.factory_part_number, PartsManualSection.model_code)
    ).order_by(
        PartsReference.factory_part_number,
        PartsManualSection.model_code,
        PartsManualSection.section_code,
    ).subquery("inner_catalog")

    # Outer query — ORDER BY libre sobre el subquery
    # k_publico: multiplicador para precio_publico desde avg_fob_cost (USD → COP)
    # permite COALESCE(public_price, avg_fob_cost * k) para ordenar Precio Final
    # con fallback a precio calculado cuando no hay precio manual
    _pf = pricing_factors
    k_publico = (
        _pf["import_factor"]
        * (1 + _pf["provider_margin"])
        * (1 + _pf["distributor_margin"])
        * (1 + _pf["iva_rate"])
        * _pf["trm"]
    )
    # fob_effective: avg_fob_cost si existe, sino preliminary_fob (mismo que _build_item)
    fob_effective = func.coalesce(inner_sq.c.avg_fob_cost, inner_sq.c.preliminary_fob)
    _SORT_MAP = {
        "factory_part_number": inner_sq.c.fpn,
        "description":         inner_sq.c.description,
        "description_es":      inner_sq.c.description_es,
        "public_price":        func.coalesce(
            cast(inner_sq.c.public_price, SANumeric(15, 2)),
            cast(fob_effective * k_publico, SANumeric(15, 2)),
        ),
        "section_code":        inner_sq.c.section_code,
        "vehicle_model_name":  inner_sq.c.vehicle_model_pattern,
        "avg_fob_cost":        cast(fob_effective, SANumeric(12, 4)),
        "rotation_class":      inner_sq.c.rotation_class,
        "needs_price_review":  inner_sq.c.needs_price_review,
    }
    sort_expr = _SORT_MAP.get(sort_col, inner_sq.c.section_code)
    order_expr = nullslast(sort_expr.asc() if sort_dir == "asc" else sort_expr.desc())

    rows_q = (
        select(inner_sq)
        .order_by(order_expr, inner_sq.c.fpn)
        .offset((page - 1) * page_size)
        .limit(page_size)
    )
    rows = (await db.execute(rows_q)).all()

    # Sustitutos y prev_codes para las referencias de esta página
    fpns = [r[0] for r in rows]
    subs_by_fpn: dict[str, list[PartSubstitute]] = {}
    prev_codes_by_fpn: dict[str, list[str]] = {}
    if fpns:
        subs_q = (
            select(PartSubstitute)
            .where(PartSubstitute.factory_part_number.in_(fpns))
            .order_by(PartSubstitute.factory_part_number, PartSubstitute.position)
        )
        for s in (await db.execute(subs_q)).scalars().all():
            subs_by_fpn.setdefault(s.factory_part_number, []).append(s)

        refs_q = select(PartsReference.factory_part_number, PartsReference.prev_codes).where(
            PartsReference.factory_part_number.in_(fpns)
        )
        for fpn_r, codes in (await db.execute(refs_q)).all():
            extracted = []
            for entry in (codes or []):
                if isinstance(entry, dict) and "code" in entry:
                    extracted.append(entry["code"])
                elif isinstance(entry, str) and entry:
                    extracted.append(entry)
            prev_codes_by_fpn[fpn_r] = extracted

    coverage_by_fpn: dict[str, str] = {}
    if fpns:
        from app.models.imports import SparePartLot, ShipmentOrder as _SO
        from sqlalchemy import case as sa_case, and_ as sa_and
        _fn = func.upper(func.trim(func.replace(SparePartItem.part_number, ' ', '')))

        # Build norm_code → canonical FPN map, including prev_codes
        _code_to_fpn: dict[str, str] = {}
        for _f in fpns:
            _code_to_fpn[_f.upper().strip().replace(' ', '')] = _f
        for _fpn_r, _prev_list in prev_codes_by_fpn.items():
            for _prev in _prev_list:
                _norm = _prev.upper().strip().replace(' ', '')
                if _norm not in _code_to_fpn:
                    _code_to_fpn[_norm] = _fpn_r

        cov_q = (
            select(
                _fn.label('fpn'),
                func.max(
                    sa_case(
                        (sa_and(SparePartItem.qty_physical.isnot(None), SparePartItem.qty_physical > 0), 3),
                        (sa_and(
                            SparePartItem.qty_received > 0,
                            SparePartItem.qty_physical.is_(None),
                            or_(_SO.bl_container.isnot(None), SparePartLot.packing_list_received == True),
                        ), 2),
                        (sa_and(SparePartLot.packing_list_received == False, SparePartItem.status != 'CANCELLED'), 1),
                        (SparePartItem.status.in_(['BACKORDER', 'BACKORDER_PARCIAL']), 1),
                        else_=0,
                    )
                ).label('score'),
            )
            .select_from(SparePartItem)
            .join(SparePartLot, SparePartLot.id == SparePartItem.lot_id)
            .outerjoin(_SO, _SO.id == SparePartLot.shipment_order_id)
            .where(_fn.in_(list(_code_to_fpn.keys())))
            .group_by(_fn)
        )
        _SCORE = {3: 'aqui', 2: 'en_camino', 1: 'pedido', 0: 'sin_pedido'}
        _RANK  = {'aqui': 3, 'en_camino': 2, 'pedido': 1, 'sin_pedido': 0}
        for _fpn_val, _score in (await db.execute(cov_q)).all():
            _canonical = _code_to_fpn.get(_fpn_val, _fpn_val)
            _new_status = _SCORE.get(_score or 0, 'sin_pedido')
            _cur = coverage_by_fpn.get(_canonical)
            if _cur is None or _RANK[_new_status] > _RANK.get(_cur, 0):
                coverage_by_fpn[_canonical] = _new_status

    def _build_item(r) -> CatalogItemResult:
        avg_fob      = float(r[11]) if r[11] is not None else None
        prelim_fob   = float(r[14]) if r[14] is not None else None
        # Jerarquía: avg_fob_cost (packing list) > preliminary_fob (PI)
        fob_for_calc = avg_fob if avg_fob is not None else prelim_fob
        is_prelim    = avg_fob is None and prelim_fob is not None
        prices       = compute_prices(fob_for_calc, pricing_factors)
        fpn = r[0]
        public_price = float(r[3]) if r[3] is not None else None
        precio_distribuidor = prices["precio_distribuidor"]
        if public_price is not None:
            # Precio Final manual reemplaza al público calculado desde el FOB.
            # El Precio Distribuidor se deriva de ese valor manteniendo el
            # margen del distribuidor configurado (el IVA se cancela entre
            # ambas etapas: precio_distribuidor = precio_final / (1 + margen)).
            precio_distribuidor = round(
                public_price / (1 + pricing_factors["distributor_margin"]), 0
            )
        return CatalogItemResult(
            factory_part_number=fpn,
            description=r[1],
            description_es=r[2],
            public_price=public_price,
            section_code=r[4],
            section_name=r[5],
            vehicle_model_name=r[7],
            pending_task_id=str(r[8]) if r[8] else None,
            pending_candidate_code=r[9],
            pending_score=float(r[10]) if r[10] is not None else None,
            avg_fob_cost=avg_fob,
            preliminary_fob=prelim_fob,
            fob_is_preliminary=is_prelim,
            costo_importado=prices["costo_importado"],
            precio_distribuidor=precio_distribuidor,
            precio_publico_calculado=prices["precio_publico"],
            substitutes=[
                PartSubstituteOut(
                    substitute_part_code=s.substitute_part_code,
                    brand=s.brand,
                    model=s.model,
                    position=s.position,
                )
                for s in subs_by_fpn.get(fpn, [])
            ],
            rotation_class=r[12],
            prev_codes=prev_codes_by_fpn.get(fpn, []),
            needs_price_review=bool(r[13]) if r[13] is not None else False,
            coverage_status=coverage_by_fpn.get(fpn.upper().strip().replace(' ', ''), 'sin_pedido'),
        )

    return CatalogListResult(
        total=total,
        items=[_build_item(r) for r in rows],
    )


# ── Exportar maestro de partes a Excel ───────────────────────────────────────

@router.get("/admin/catalog/export")
async def export_catalog_excel(
    search: str = "",
    model_code: str = "",
    rotation_class: str = "",
    coverage_status: str = "",
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Descarga el maestro de partes completo como Excel con todas las columnas."""
    if not current_user.is_superadmin:
        raise HTTPException(status_code=403, detail="Solo superadmin")

    result = await _list_catalog_impl(
        search=search,
        model_code=model_code,
        only_pending=False,
        only_price_review=False,
        rotation_class=rotation_class,
        coverage_status=coverage_status,
        sort_col="section_code",
        sort_dir="asc",
        page=1,
        page_size=50_000,
        db=db,
    )

    fpns = [item.factory_part_number for item in result.items]
    extra: dict[str, dict] = {}
    if fpns:
        for fpn, unit, qty, updated in (await db.execute(
            select(
                PartsReference.factory_part_number,
                PartsReference.unit,
                PartsReference.total_fob_qty,
                PartsReference.last_cost_updated,
            ).where(PartsReference.factory_part_number.in_(fpns))
        )).all():
            extra[fpn] = {"unit": unit, "total_fob_qty": qty, "last_cost_updated": updated}

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    HEADERS = [
        ("Código Fábrica",             22),
        ("Descripción EN",             36),
        ("Descripción ES",             36),
        ("Unidad",                      8),
        ("Códigos Anteriores",         22),
        ("Sección Código",             12),
        ("Sección Nombre",             30),
        ("Modelo Moto",                20),
        ("Rotación",                   10),
        ("Cobertura",                  13),
        ("Req. Revisión Precio",       14),
        ("FOB Promedio USD",           14),
        ("FOB Preliminar USD",         14),
        ("FOB es Preliminar",          12),
        ("Fecha Último Costo",         18),
        ("Cant. Total FOB",            12),
        ("Costo Importado COP",        16),
        ("Precio Distribuidor COP",    16),
        ("Precio Público Calc. COP",   17),
        ("Precio Público Manual COP",  17),
        ("Sust. 1 Código",             20),
        ("Sust. 1 Marca",              14),
        ("Sust. 1 Modelo",             20),
        ("Sust. 2 Código",             20),
        ("Sust. 2 Marca",              14),
        ("Sust. 2 Modelo",             20),
        ("Sust. 3 Código",             20),
        ("Sust. 3 Marca",              14),
        ("Sust. 3 Modelo",             20),
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Maestro de Partes"

    hdr_font = Font(bold=True, color="FFFFFF", size=9)
    hdr_fill = PatternFill("solid", fgColor="1e1e2e")
    center   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left     = Alignment(horizontal="left",   vertical="center")

    for ci, (h, w) in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    COVERAGE_LABELS = {
        "aqui": "Aquí",
        "en_camino": "En camino",
        "pedido": "Pedido",
        "sin_pedido": "Sin pedido",
    }
    data_font = Font(size=9)

    for i, item in enumerate(result.items):
        ex = extra.get(item.factory_part_number, {})
        subs = item.substitutes

        def _sub(n: int, attr: str) -> str:
            return getattr(subs[n], attr, "") if n < len(subs) else ""

        values = [
            item.factory_part_number,
            item.description,
            item.description_es or "",
            ex.get("unit") or "",
            ", ".join(item.prev_codes),
            item.section_code,
            item.section_name,
            item.vehicle_model_name or "",
            (item.rotation_class or "").upper() if item.rotation_class else "",
            COVERAGE_LABELS.get(item.coverage_status or "", item.coverage_status or ""),
            "Sí" if item.needs_price_review else "No",
            item.avg_fob_cost,
            item.preliminary_fob,
            "Sí" if item.fob_is_preliminary else "No",
            ex.get("last_cost_updated"),
            ex.get("total_fob_qty"),
            item.costo_importado,
            item.precio_distribuidor,
            item.precio_publico_calculado,
            item.public_price,
            _sub(0, "substitute_part_code"),
            _sub(0, "brand"),
            _sub(0, "model"),
            _sub(1, "substitute_part_code"),
            _sub(1, "brand"),
            _sub(1, "model"),
            _sub(2, "substitute_part_code"),
            _sub(2, "brand"),
            _sub(2, "model"),
        ]

        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=i + 2, column=ci, value=val)
            cell.font = data_font
            cell.alignment = left

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = "maestro_partes"
    if model_code:
        filename += f"_{model_code}"
    if search:
        filename += "_filtrado"
    filename += ".xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Limpiar catálogo completo de un modelo ────────────────────────────────────

@router.delete("/admin/catalog/{model_code}", status_code=204)
async def delete_catalog(
    model_code: str,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Elimina todas las secciones (y sus ítems vía CASCADE) de un model_code. Solo superadmin."""
    if not current_user.is_superadmin:
        raise HTTPException(status_code=403, detail="Solo superadmin")

    await db.execute(
        sa_delete(PartsManualSection).where(PartsManualSection.model_code == model_code)
    )
    await db.commit()


# ── Eliminación de un repuesto individual del catálogo ───────────────────────

@router.delete("/admin/catalog/part/{factory_part_number:path}", status_code=204)
async def delete_catalog_part(
    factory_part_number: str,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Elimina un repuesto del catálogo. Bloqueado si tiene costos, historial o pedidos asociados."""
    if not current_user.is_superadmin:
        raise HTTPException(status_code=403, detail="Solo superadmin")

    ref = await db.get(PartsReference, factory_part_number)
    if not ref:
        raise HTTPException(status_code=404, detail="Referencia no encontrada")

    if ref.avg_fob_cost is not None:
        raise HTTPException(status_code=409, detail="No se puede eliminar: tiene costo promedio calculado.")

    history_count = (await db.execute(
        select(func.count()).select_from(PartCostHistory)
        .where(PartCostHistory.factory_part_number == factory_part_number)
    )).scalar_one()
    if history_count > 0:
        raise HTTPException(status_code=409, detail="No se puede eliminar: tiene historial de costos.")

    catalog = await db.get(PartCatalog, factory_part_number)
    if catalog:
        from app.models.logistics import PartsOrderItem, PurchaseOrderItem
        order_refs = (await db.execute(
            select(func.count()).select_from(PartsOrderItem)
            .where(PartsOrderItem.part_code == factory_part_number)
        )).scalar_one()
        purchase_refs = (await db.execute(
            select(func.count()).select_from(PurchaseOrderItem)
            .where(PurchaseOrderItem.part_code == factory_part_number)
        )).scalar_one()
        if order_refs > 0 or purchase_refs > 0:
            raise HTTPException(status_code=409, detail="No se puede eliminar: tiene órdenes de compra asociadas.")
        await db.delete(catalog)

    await db.execute(
        sa_delete(PartsManualItem).where(PartsManualItem.factory_part_number == factory_part_number)
    )
    await db.execute(
        sa_delete(PartsCodeReviewTask).where(
            (PartsCodeReviewTask.existing_code == factory_part_number) |
            (PartsCodeReviewTask.candidate_code == factory_part_number)
        )
    )
    await db.delete(ref)
    await db.commit()


# ── Edición inline de un repuesto del catálogo ───────────────────────────────

@router.patch("/admin/catalog/{factory_part_number:path}", status_code=200)
async def update_catalog_item(
    factory_part_number: str,
    payload: CatalogItemUpdate,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Actualiza descripción, descripción ES manual y/o precio público. Solo superadmin."""
    if not current_user.is_superadmin:
        raise HTTPException(status_code=403, detail="Solo superadmin")

    ref = await db.get(PartsReference, factory_part_number)
    if not ref:
        raise HTTPException(status_code=404, detail="Referencia no encontrada")

    if payload.description is not None:
        ref.description = payload.description
    if payload.description_es_manual is not None:
        ref.description_es_manual = payload.description_es_manual

    if payload.public_price is not None:
        catalog = await db.get(PartCatalog, factory_part_number)
        if catalog:
            catalog.public_price = payload.public_price
            if payload.description is not None:
                catalog.description = payload.description
        else:
            db.add(PartCatalog(
                part_code=factory_part_number,
                description=payload.description or ref.description,
                public_price=payload.public_price,
            ))

    if payload.rotation_class is not None:
        ref.rotation_class = payload.rotation_class

    if payload.needs_price_review is not None:
        ref.needs_price_review = payload.needs_price_review

    if payload.substitutes is not None:
        await db.execute(
            sa_delete(PartSubstitute).where(PartSubstitute.factory_part_number == factory_part_number)
        )
        for idx, sub in enumerate(payload.substitutes[:3], start=1):
            code = sub.substitute_part_code.strip()
            brand = sub.brand.strip()
            model = sub.model.strip()
            if code and brand and model:
                db.add(PartSubstitute(
                    factory_part_number=factory_part_number,
                    substitute_part_code=code,
                    brand=brand,
                    model=model,
                    position=idx,
                ))

    if payload.prev_codes is not None:
        clean = [c.strip() for c in payload.prev_codes if c.strip() and c.strip() != factory_part_number]
        ref.prev_codes = [{"code": c} for c in clean[:5]]
        from app.services.pricing_service import recalculate_part_cost
        await recalculate_part_cost(db, factory_part_number)

    await db.commit()
    return {"ok": True}


@router.post("/admin/catalog-replace/{factory_part_number:path}", status_code=200)
async def replace_catalog_code(
    factory_part_number: str,
    payload: ReplaceCodeRequest,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Reemplaza manualmente el código de fábrica de un repuesto. Solo superadmin."""
    if not current_user.is_superadmin:
        raise HTTPException(status_code=403, detail="Solo superadmin")

    new_code = payload.new_code.strip()
    if not new_code:
        raise HTTPException(status_code=422, detail="El nuevo código no puede estar vacío")
    if new_code == factory_part_number:
        raise HTTPException(status_code=422, detail="El nuevo código es igual al actual")

    existing_ref = await db.get(PartsReference, factory_part_number)
    if not existing_ref:
        raise HTTPException(status_code=404, detail="Referencia no encontrada")

    # Si el nuevo código ya existe en el catálogo, no permitir — evitar colisión de PK
    conflict = await db.get(PartsReference, new_code)
    if conflict is not None:
        raise HTTPException(
            status_code=409,
            detail=f"El código '{new_code}' ya existe en el catálogo. Usá el flujo de Verificar si querés unificar dos entradas."
        )

    # Construir prev_codes: el código que sale entra al historial, máx 5
    new_prev = ([{"code": factory_part_number}] + list(existing_ref.prev_codes or []))[:5]

    new_ref = PartsReference(
        factory_part_number=new_code,
        um_part_number=existing_ref.um_part_number,
        description=payload.description.strip() if payload.description else existing_ref.description,
        description_es_manual=payload.description_es_manual,
        unit=existing_ref.unit,
        prev_codes=new_prev,
        # Carry over classification/pricing so replacing a factory code never
        # silently wipes it -- previously the old row (with all of this) was
        # deleted below while the new row started blank, losing rotation and
        # price with no error or trace (found investigating a user report of
        # rotation/price disappearing from re-uploaded catalog parts).
        rotation_class=existing_ref.rotation_class,
        avg_fob_cost=existing_ref.avg_fob_cost,
        preliminary_fob=existing_ref.preliminary_fob,
        needs_price_review=existing_ref.needs_price_review,
        total_fob_qty=existing_ref.total_fob_qty,
        last_cost_updated=existing_ref.last_cost_updated,
    )
    db.add(new_ref)
    await db.flush()

    # Redirigir todos los ítems del catálogo al nuevo código
    await db.execute(
        sa_update(PartsManualItem)
        .where(PartsManualItem.factory_part_number == factory_part_number)
        .values(factory_part_number=new_code)
    )
    await db.flush()

    # Migrar entrada de PartCatalog si existe
    old_catalog = await db.get(PartCatalog, factory_part_number)
    if old_catalog:
        new_price = payload.public_price if payload.public_price is not None else float(old_catalog.public_price)
        new_desc  = (payload.description.strip() if payload.description else None) or old_catalog.description
        db.add(PartCatalog(part_code=new_code, description=new_desc, public_price=new_price))
        await db.delete(old_catalog)
        await db.flush()
    elif payload.public_price is not None:
        desc = (payload.description.strip() if payload.description else None) or new_ref.description
        db.add(PartCatalog(part_code=new_code, description=desc, public_price=payload.public_price))

    # Eliminar la referencia vieja
    await db.delete(existing_ref)

    # Cerrar tareas pendientes de revisión que involucren el código viejo
    await db.execute(
        sa_update(PartsCodeReviewTask)
        .where(
            PartsCodeReviewTask.existing_code == factory_part_number,
            PartsCodeReviewTask.status == "pending",
        )
        .values(status="rejected", resolved_at=datetime.utcnow())
    )

    # Recalcular costo promedio para el nuevo código — puede haber SparePartItems con precio
    from app.services.pricing_service import recalculate_part_cost
    await recalculate_part_cost(db, new_code)

    await db.commit()
    return {"ok": True, "new_code": new_code}


# ── Detección manual y revisión de cambios de código ─────────────────────────

@router.post("/admin/backfill-costs")
async def backfill_costs(
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Recalcula avg_fob_cost para todas las referencias sin costo que tienen precios en pedidos."""
    if not current_user.is_superadmin:
        raise HTTPException(status_code=403, detail="Solo superadmin")

    from app.services.pricing_service import recalculate_part_cost

    result = await db.execute(
        select(PartsReference.factory_part_number).where(PartsReference.avg_fob_cost.is_(None))
    )
    codes = [row[0] for row in result.all()]

    updated = 0
    for code in codes:
        before = (await db.get(PartsReference, code))
        await recalculate_part_cost(db, code)
        after = (await db.get(PartsReference, code))
        if after and after.avg_fob_cost is not None:
            updated += 1

    await db.commit()
    return {"checked": len(codes), "updated": updated}


# TEMPORARY debug endpoint -- diagnosing why "Recalcular costos" left many
# freshly-loaded Xtreet 401 references without a cost (2026-08-11). Read-only,
# superadmin-only. Remove after use.
@router.get("/admin/debug-cost-match/{model_code}")
async def debug_cost_match(
    model_code: str,
    section_code: str = "",
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    if not current_user.is_superadmin:
        raise HTTPException(status_code=403, detail="Solo superadmin")

    stmt = (
        select(PartsReference.factory_part_number, PartsReference.avg_fob_cost)
        .join(PartsManualItem, PartsManualItem.factory_part_number == PartsReference.factory_part_number)
        .join(PartsManualSection, PartsManualSection.id == PartsManualItem.section_id)
        .where(PartsManualSection.model_code == model_code)
    )
    if section_code:
        stmt = stmt.where(PartsManualSection.section_code == section_code)
    refs = (await db.execute(stmt)).all()

    results = []
    for fpn, avg_fob_cost in refs:
        if avg_fob_cost is not None:
            continue
        exact = (await db.execute(
            select(func.count()).select_from(SparePartItem)
            .where(SparePartItem.part_number == fpn, SparePartItem.unit_price.isnot(None))
        )).scalar_one()
        prefix = fpn.split("-")[0]
        near = (await db.execute(
            select(SparePartItem.part_number, SparePartItem.unit_price)
            .where(SparePartItem.part_number.like(f"{prefix}%"))
            .limit(5)
        )).all()
        results.append({
            "factory_part_number": fpn,
            "exact_priced_matches": exact,
            "near_matches": [{"part_number": pn, "unit_price": float(up) if up is not None else None} for pn, up in near],
        })
    return {"missing_cost_count": len(results), "details": results}


# TEMPORARY debug endpoint -- per-code deep dive, same investigation as above
# but for a single reported code (2026-08-11). Read-only, superadmin-only.
# Remove after use.
@router.get("/admin/debug-cost-match-single/{factory_part_number}")
async def debug_cost_match_single(
    factory_part_number: str,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    if not current_user.is_superadmin:
        raise HTTPException(status_code=403, detail="Solo superadmin")

    ref = await db.get(PartsReference, factory_part_number)
    exact_items = (await db.execute(
        select(
            SparePartItem.id, SparePartItem.part_number, SparePartItem.unit_price,
            SparePartItem.qty_ordered, SparePartItem.qty_physical, SparePartItem.qty_received,
            SparePartItem.fob_pi,
        ).where(SparePartItem.part_number == factory_part_number)
    )).all()

    norm = factory_part_number.strip().upper().replace(" ", "")
    fuzzy_items = (await db.execute(
        select(
            SparePartItem.id, SparePartItem.part_number, SparePartItem.unit_price,
            SparePartItem.qty_ordered, SparePartItem.qty_physical,
        ).where(func.upper(func.trim(SparePartItem.part_number)) == norm)
    )).all()

    return {
        "queried_code": factory_part_number,
        "queried_code_repr": repr(factory_part_number),
        "reference_exists": ref is not None,
        "reference_factory_part_number_repr": repr(ref.factory_part_number) if ref else None,
        "reference_avg_fob_cost": float(ref.avg_fob_cost) if ref and ref.avg_fob_cost is not None else None,
        "reference_preliminary_fob": float(ref.preliminary_fob) if ref and ref.preliminary_fob is not None else None,
        "reference_prev_codes": ref.prev_codes if ref else None,
        "exact_match_items": [
            {
                "id": str(i), "part_number": repr(pn), "unit_price": float(up) if up is not None else None,
                "qty_ordered": qo, "qty_physical": qp, "qty_received": qr,
                "fob_pi": float(fp) if fp is not None else None,
            }
            for i, pn, up, qo, qp, qr, fp in exact_items
        ],
        "fuzzy_match_items_if_different": [
            {"id": str(i), "part_number": repr(pn), "unit_price": float(up) if up is not None else None,
             "qty_ordered": qo, "qty_physical": qp}
            for i, pn, up, qo, qp in fuzzy_items
        ] if len(fuzzy_items) != len(exact_items) else "same as exact_match_items",
    }


@router.post("/admin/detect-code-changes")
async def detect_code_changes(
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Lanza la detección de posibles cambios de código por similitud de descripción. Solo superadmin."""
    if not current_user.is_superadmin:
        raise HTTPException(status_code=403, detail="Solo superadmin")
    created = await _detect_code_changes(db)
    return {"tasks_created": created}


@router.get("/admin/diagnose-detection")
async def diagnose_detection(
    part_code: str = "",
    description: str = "",
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Diagnóstica por qué un spare_part_item no genera tarea de revisión de código.
    Pasá part_code (el código del pedido) o description (texto a buscar).
    """
    if not current_user.is_superadmin:
        raise HTTPException(status_code=403, detail="Solo superadmin")

    results = {}

    # Paso 1: ¿Existe el item en spare_part_items?
    if part_code:
        r1 = await db.execute(text("""
            SELECT part_number, description, description_es, model_applicable, unit_price, qty_ordered
            FROM spare_part_items
            WHERE part_number = :code
            ORDER BY created_at DESC LIMIT 5
        """), {"code": part_code.strip().upper().replace(" ", "")})
    else:
        r1 = await db.execute(text("""
            SELECT part_number, description, description_es, model_applicable, unit_price, qty_ordered
            FROM spare_part_items
            WHERE description ILIKE :desc OR description_es ILIKE :desc
            ORDER BY created_at DESC LIMIT 5
        """), {"desc": f"%{description}%"})
    results["1_spare_part_items"] = [dict(r._mapping) for r in r1.all()]

    if not results["1_spare_part_items"]:
        results["conclusion"] = "No existe ningún spare_part_item con ese código/descripción."
        return results

    sample = results["1_spare_part_items"][0]
    pn = sample["part_number"]
    desc = sample["description"]
    model_ap = sample["model_applicable"]

    # Paso 2: ¿Tiene description en inglés?
    results["2_description_en_null"] = desc is None or desc == ""
    results["2_description_es_null"] = sample["description_es"] is None or sample["description_es"] == ""

    # Paso 3: ¿Tiene model_applicable?
    results["3_model_applicable"] = model_ap
    results["3_model_applicable_null"] = model_ap is None or model_ap == ""

    # Paso 4: ¿Existe en vehicle_catalog_map?
    if model_ap:
        r4 = await db.execute(text("""
            SELECT vehicle_model_pattern, catalog_model_code
            FROM vehicle_catalog_map
            WHERE vehicle_model_pattern = :model
        """), {"model": model_ap})
        results["4_vehicle_catalog_map_match"] = [dict(r._mapping) for r in r4.all()]

        # También buscar coincidencias parciales para diagnóstico
        r4b = await db.execute(text("""
            SELECT vehicle_model_pattern, catalog_model_code
            FROM vehicle_catalog_map
            WHERE vehicle_model_pattern ILIKE :model
        """), {"model": f"%{model_ap}%"})
        results["4_vehicle_catalog_map_ilike"] = [dict(r._mapping) for r in r4b.all()]
    else:
        results["4_vehicle_catalog_map_match"] = []
        results["4_vehicle_catalog_map_ilike"] = []

    # Paso 5: ¿El código ya existe en parts_references?
    r5 = await db.execute(text("""
        SELECT factory_part_number, description FROM parts_references
        WHERE factory_part_number = :code
    """), {"code": pn})
    results["5_candidate_already_in_catalog"] = [dict(r._mapping) for r in r5.all()]

    # Paso 6: Similitud real contra el catálogo (sin filtros de modelo)
    if desc:
        r6 = await db.execute(text("""
            SELECT pr.factory_part_number, pr.description,
                   similarity(:desc, pr.description) AS score
            FROM parts_references pr
            WHERE similarity(:desc, pr.description) > 0.4
            ORDER BY score DESC LIMIT 10
        """), {"desc": desc})
        results["6_similarity_scores_in_catalog"] = [dict(r._mapping) for r in r6.all()]
    else:
        results["6_similarity_scores_in_catalog"] = "No hay description en inglés para comparar"

    # Paso 7: ¿Hay tareas ya existentes para este código?
    r7 = await db.execute(text("""
        SELECT id, existing_code, candidate_code, status, similarity_score
        FROM parts_code_review_tasks
        WHERE candidate_code = :code OR existing_code = :code
        ORDER BY created_at DESC LIMIT 5
    """), {"code": pn})
    results["7_existing_tasks"] = [dict(r._mapping) for r in r7.all()]

    # Conclusión automática
    issues = []
    if results["2_description_en_null"]:
        issues.append("description (inglés) es NULL — la query lo filtra antes de comparar")
    if results["3_model_applicable_null"]:
        issues.append("model_applicable es NULL — el JOIN con vehicle_catalog_map falla")
    elif not results["4_vehicle_catalog_map_match"]:
        issues.append(f"model_applicable '{model_ap}' no coincide exactamente con ningún vehicle_model_pattern")
    if results["5_candidate_already_in_catalog"]:
        issues.append("El código del pedido YA existe en parts_references — la detección lo ignora a propósito")
    if results["6_similarity_scores_in_catalog"] and isinstance(results["6_similarity_scores_in_catalog"], list):
        top = results["6_similarity_scores_in_catalog"]
        if top and top[0]["score"] < 0.9:
            issues.append(f"Mejor similitud encontrada: {top[0]['score']:.2f} (umbral es 0.9) — descripción difiere")

    results["conclusion"] = issues if issues else ["No se encontró el problema automáticamente — revisar manualmente los pasos"]
    return results


@router.post("/admin/review-tasks/{task_id}/approve", status_code=200)
async def approve_review_task(
    task_id: str,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Aprueba la sustitución de código: el candidato pasa a ser el código activo. Solo superadmin."""
    if not current_user.is_superadmin:
        raise HTTPException(status_code=403, detail="Solo superadmin")

    task = await db.get(PartsCodeReviewTask, _uuid.UUID(task_id))
    if not task or task.status != "pending":
        raise HTTPException(status_code=404, detail="Tarea no encontrada o ya resuelta")

    existing_ref = await db.get(PartsReference, task.existing_code)
    if not existing_ref:
        raise HTTPException(status_code=404, detail="Código existente no encontrado en el catálogo")

    # El candidato podría ya existir (cargado por otro camino)
    candidate_ref = await db.get(PartsReference, task.candidate_code)

    if candidate_ref is None:
        # Construir nuevo prev_codes: [código que sale] + prev anteriores, máx 5
        new_prev = ([{"code": task.existing_code}] + list(existing_ref.prev_codes or []))[:5]
        candidate_ref = PartsReference(
            factory_part_number=task.candidate_code,
            um_part_number=existing_ref.um_part_number,
            description=existing_ref.description,
            description_es_manual=existing_ref.description_es_manual,
            unit=existing_ref.unit,
            prev_codes=new_prev,
            rotation_class=existing_ref.rotation_class,
        )
        db.add(candidate_ref)
        await db.flush()
    else:
        # El candidato ya existe — heredar campos del existente si no los tiene
        prev = candidate_ref.prev_codes or []
        existing_in_prev = any(
            (e["code"] == task.existing_code if isinstance(e, dict) else e == task.existing_code)
            for e in prev
        )
        if not existing_in_prev:
            candidate_ref.prev_codes = ([{"code": task.existing_code}] + list(prev))[:5]
        if candidate_ref.rotation_class is None and existing_ref.rotation_class is not None:
            candidate_ref.rotation_class = existing_ref.rotation_class
        if candidate_ref.description_es_manual is None and existing_ref.description_es_manual is not None:
            candidate_ref.description_es_manual = existing_ref.description_es_manual

    # Redirigir todos los items del catálogo al nuevo código
    await db.execute(
        sa_update(PartsManualItem)
        .where(PartsManualItem.factory_part_number == task.existing_code)
        .values(factory_part_number=task.candidate_code)
    )
    await db.flush()

    # Eliminar la referencia vieja (ya no tiene items apuntando a ella)
    await db.delete(existing_ref)

    # Resolver tarea
    task.status = "approved"
    task.resolved_at = datetime.utcnow()
    task.resolved_by = current_user.user_id

    # Rechazar automáticamente otras tareas pendientes para el mismo código existente
    await db.execute(
        sa_update(PartsCodeReviewTask)
        .where(
            PartsCodeReviewTask.existing_code == task.existing_code,
            PartsCodeReviewTask.id != task.id,
            PartsCodeReviewTask.status == "pending",
        )
        .values(status="rejected", resolved_at=datetime.utcnow())
    )

    from app.services.pricing_service import recalculate_part_cost
    await recalculate_part_cost(db, task.candidate_code)

    await db.commit()
    return {"ok": True, "new_code": task.candidate_code}


@router.post("/admin/review-tasks/{task_id}/reject", status_code=200)
async def reject_review_task(
    task_id: str,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Descarta la sugerencia de cambio de código. Solo superadmin."""
    if not current_user.is_superadmin:
        raise HTTPException(status_code=403, detail="Solo superadmin")

    task = await db.get(PartsCodeReviewTask, _uuid.UUID(task_id))
    if not task or task.status != "pending":
        raise HTTPException(status_code=404, detail="Tarea no encontrada o ya resuelta")

    task.status = "rejected"
    task.resolved_at = datetime.utcnow()
    task.resolved_by = current_user.user_id
    await db.commit()
    return {"ok": True}


# ── Endpoint de administración (frontend) ──────────────────────────────────────

@router.post("/admin/load-section", response_model=LoadSectionResult)
async def load_section(
    pdf_file: UploadFile = File(...),
    model_code: str = Form(...),
    vehicle_model: str = Form(...),
    force: bool = Form(False),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Carga una sección del catálogo de partes desde un PDF. Solo superadmin."""
    if not current_user.is_superadmin:
        raise HTTPException(status_code=403, detail="Solo superadmin puede cargar catálogos")

    vehicle_model_exists = (await db.execute(
        select(VehicleModel.model_name).where(VehicleModel.model_name == vehicle_model)
    )).scalar_one_or_none()
    if not vehicle_model_exists:
        raise HTTPException(
            status_code=422,
            detail=f"El modelo '{vehicle_model}' no existe en el catálogo de modelos de vehículos (vehicle_models). "
                   "Verificá el nombre exacto o creá el modelo primero.",
        )

    filename = pdf_file.filename or "unknown.pdf"
    section_code, section_name = _parse_section_filename(filename)

    pdf_bytes = await pdf_file.read()

    fd, tmp_path = tempfile.mkstemp(suffix=".pdf")
    try:
        with os.fdopen(fd, "wb") as f:
            f.write(pdf_bytes)

        # FIX 1+2: Parsear tabla de repuestos PRIMERO — antes de tocar la DB
        parts: list[dict] = []
        parse_error: str | None = None
        try:
            parts = _parse_parts_table(tmp_path)
        except Exception as e:
            parse_error = str(e)
            logger.warning(f"load_section parse error ({filename}): {e}")

        if not parts:
            raise HTTPException(
                status_code=422,
                detail=(
                    f"No se detectaron partes en '{filename}'"
                    + (f": {parse_error}" if parse_error else
                       ". Verificá que el PDF contiene una tabla de partes válida.")
                ),
            )

        # FIX 4: Verificar sección existente ANTES de destruir nada
        existing_section = (await db.execute(
            select(PartsManualSection).where(
                PartsManualSection.model_code == model_code,
                PartsManualSection.section_code == section_code,
            )
        )).scalar_one_or_none()

        if existing_section and not force:
            existing_count = (await db.execute(
                select(func.count()).select_from(PartsManualItem)
                .where(PartsManualItem.section_id == existing_section.id)
            )).scalar_one()
            if existing_count > 0:
                raise HTTPException(
                    status_code=409,
                    detail={
                        "code": "section_exists",
                        "existing_parts": existing_count,
                        "section_code": section_code,
                        "section_name": existing_section.section_name,
                        "message": (
                            f"La sección '{section_code}' ya tiene {existing_count} partes cargadas. "
                            "Enviá force=true para sobreescribir."
                        ),
                    },
                )

        # 1. Extraer ilustración del PDF
        illus_bytes: bytes | None = None
        try:
            illus_bytes = _extract_section_illustration(tmp_path)
        except Exception as e:
            logger.warning(f"load_section illustration extract error ({filename}): {e}")

        # 2. Obtener logo de la configuración del sistema
        logo_bytes: bytes | None = None
        try:
            logo_record = await db.get(SystemConfig, "logo_base64")
            if logo_record and logo_record.value:
                b64 = logo_record.value
                if "," in b64:
                    b64 = b64.split(",", 1)[1]
                logo_bytes = base64.b64decode(b64)
        except Exception as e:
            logger.warning(f"load_section logo fetch error: {e}")

        # 3. Generar card estilizada y subir a MinIO
        diagram_url = None
        png_bytes: bytes | None = None

        if illus_bytes:
            try:
                from app.services.diagram_styler import create_diagram_card
                png_bytes = create_diagram_card(
                    illus_bytes=illus_bytes,
                    section_code=section_code,
                    section_name=section_name,
                    logo_bytes=logo_bytes,
                )
            except Exception:
                logger.exception(f"load_section diagram_styler failed ({filename})")

        # Fallback: render página completa si el styler falló o no había ilustración
        if not png_bytes:
            logger.warning(f"load_section using full-page render fallback ({filename})")
            try:
                doc = fitz.open(tmp_path)
                pix = doc[0].get_pixmap(matrix=fitz.Matrix(2, 2))
                png_bytes = pix.tobytes("png")
                doc.close()
            except Exception as e:
                logger.error(f"load_section fallback render failed ({filename}): {e}")

        if png_bytes:
            try:
                client = _minio_client()
                _ensure_parts_bucket(client)
                object_name = f"{model_code}/{section_code}.png"
                client.put_object(
                    bucket_name=PARTS_BUCKET,
                    object_name=object_name,
                    data=io.BytesIO(png_bytes),
                    length=len(png_bytes),
                    content_type="image/png",
                )
                diagram_url = _diagram_public_url(object_name)
            except Exception as e:
                logger.error(f"load_section MinIO upload error ({filename}): {e}")

    finally:
        os.unlink(tmp_path)

    # FIX 3: Guardar snapshot histórico ANTES de eliminar la sección anterior
    if existing_section:
        items_snap = (await db.execute(
            select(PartsManualItem).where(PartsManualItem.section_id == existing_section.id)
        )).scalars().all()
        db.add(PartsManualSectionHistory(
            model_code=existing_section.model_code,
            section_code=existing_section.section_code,
            section_name=existing_section.section_name,
            diagram_url=existing_section.diagram_url,
            parts_count=len(items_snap),
            snapshot=[
                {"order_num": i.order_num, "factory_part_number": i.factory_part_number}
                for i in items_snap
            ],
            replaced_at=datetime.utcnow(),
            replaced_by=_uuid.UUID(current_user.user_id) if current_user.user_id else None,
        ))

    # Eliminar sección anterior (ahora seguro: parse fue exitoso y snapshot guardado)
    await db.execute(
        sa_delete(PartsManualSection).where(
            PartsManualSection.model_code == model_code,
            PartsManualSection.section_code == section_code,
        )
    )

    # Insertar sección nueva
    section = PartsManualSection(
        model_code=model_code,
        section_code=section_code,
        section_name=section_name,
        diagram_url=diagram_url,
    )
    db.add(section)
    await db.flush()

    # Upsert references + insertar items
    refs_new = 0
    seen_refs: set[str] = set()
    for p in parts:
        factory = p.get("factory_part_number", "").strip()
        if not factory:
            continue

        if factory not in seen_refs:
            existing_ref = await db.get(PartsReference, factory)
            if not existing_ref:
                db.add(PartsReference(
                    factory_part_number=factory,
                    um_part_number=p.get("um_part_number", ""),
                    description=p.get("description", ""),
                    unit=p.get("unit"),
                ))
                refs_new += 1
            seen_refs.add(factory)

        db.add(PartsManualItem(
            section_id=section.id,
            order_num=p["order_num"],
            factory_part_number=factory,
        ))

    # Upsert VehicleCatalogMap
    catalog_map = await db.get(VehicleCatalogMap, vehicle_model)
    if catalog_map:
        catalog_map.catalog_model_code = model_code
    else:
        db.add(VehicleCatalogMap(
            vehicle_model_pattern=vehicle_model,
            catalog_model_code=model_code,
        ))

    await db.commit()

    return LoadSectionResult(
        section_code=section_code,
        section_name=section_name,
        diagram_url=diagram_url,
        parts_loaded=len(parts),
        references_new=refs_new,
    )


@router.post("/admin/sections/{section_id}/diagram")
async def replace_section_diagram(
    section_id: str,
    image_file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Reemplaza SOLO la imagen del diagrama de una sección ya existente,
    subiendo el archivo tal cual -- sin pasar por create_diagram_card ni
    ningún otro procesamiento con PIL (sin resize, sin recompresión).

    Útil cuando el superadmin ya tiene una imagen corregida (p. ej. un PNG
    con fondo transparente) y quiere reemplazarla sin re-parsear el PDF
    completo de la sección. Reutiliza el mismo esquema determinístico de
    nombre de objeto que `load_section` (`{model_code}/{section_code}.png`)
    para sobreescribir exactamente el objeto que la sección ya referencia."""
    if not current_user.is_superadmin:
        raise HTTPException(status_code=403, detail="Solo superadmin")

    if image_file.content_type not in _DIAGRAM_ALLOWED_MIME_TYPES:
        raise HTTPException(
            status_code=422,
            detail=(
                f"Tipo de archivo no permitido: {image_file.content_type}. "
                "Solo se aceptan imágenes (PNG, JPEG, WEBP)."
            ),
        )

    try:
        section_uuid = _uuid.UUID(section_id)
    except ValueError:
        raise HTTPException(status_code=404, detail="Sección no encontrada")

    section = await db.get(PartsManualSection, section_uuid)
    if not section:
        raise HTTPException(status_code=404, detail="Sección no encontrada")

    image_bytes = await image_file.read()

    object_name = f"{section.model_code}/{section.section_code}.png"
    client = _minio_client()
    _ensure_parts_bucket(client)
    client.put_object(
        bucket_name=PARTS_BUCKET,
        object_name=object_name,
        data=io.BytesIO(image_bytes),
        length=len(image_bytes),
        content_type=image_file.content_type,
    )

    section.diagram_url = _diagram_public_url(object_name)
    await db.commit()

    return {"diagram_url": section.diagram_url}


# ── Historial de secciones — consulta y restauración ────────────────────────

@router.get("/admin/sections/{model_code}/history")
async def get_section_history(
    model_code: str,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Devuelve el historial de versiones anteriores de secciones para un model_code."""
    if not current_user.is_superadmin:
        raise HTTPException(status_code=403, detail="Solo superadmin")

    result = await db.execute(
        select(PartsManualSectionHistory)
        .where(PartsManualSectionHistory.model_code == model_code)
        .order_by(PartsManualSectionHistory.replaced_at.desc())
        .limit(100)
    )
    return [
        {
            "id": str(h.id),
            "model_code": h.model_code,
            "section_code": h.section_code,
            "section_name": h.section_name,
            "parts_count": h.parts_count,
            "replaced_at": h.replaced_at.isoformat(),
        }
        for h in result.scalars().all()
    ]


@router.post("/admin/sections/{model_code}/history/{history_id}/restore", status_code=200)
async def restore_section_from_history(
    model_code: str,
    history_id: str,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Restaura una sección desde el historial. Guarda snapshot del estado actual antes de restaurar."""
    if not current_user.is_superadmin:
        raise HTTPException(status_code=403, detail="Solo superadmin")

    try:
        h_uuid = _uuid.UUID(history_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="history_id inválido")

    entry = await db.get(PartsManualSectionHistory, h_uuid)
    if not entry or entry.model_code != model_code:
        raise HTTPException(status_code=404, detail="Entrada de historial no encontrada")

    # Snapshot del estado actual antes de restaurar
    current_section = (await db.execute(
        select(PartsManualSection).where(
            PartsManualSection.model_code == model_code,
            PartsManualSection.section_code == entry.section_code,
        )
    )).scalar_one_or_none()

    if current_section:
        current_items = (await db.execute(
            select(PartsManualItem).where(PartsManualItem.section_id == current_section.id)
        )).scalars().all()
        db.add(PartsManualSectionHistory(
            model_code=current_section.model_code,
            section_code=current_section.section_code,
            section_name=current_section.section_name,
            diagram_url=current_section.diagram_url,
            parts_count=len(current_items),
            snapshot=[
                {"order_num": i.order_num, "factory_part_number": i.factory_part_number}
                for i in current_items
            ],
            replaced_at=datetime.utcnow(),
            replaced_by=_uuid.UUID(current_user.user_id) if current_user.user_id else None,
        ))
        await db.execute(
            sa_delete(PartsManualSection).where(
                PartsManualSection.model_code == model_code,
                PartsManualSection.section_code == entry.section_code,
            )
        )

    # Recrear sección desde snapshot
    section = PartsManualSection(
        model_code=entry.model_code,
        section_code=entry.section_code,
        section_name=entry.section_name,
        diagram_url=entry.diagram_url,
    )
    db.add(section)
    await db.flush()

    restored = 0
    for item_data in (entry.snapshot or []):
        factory = str(item_data.get("factory_part_number", "")).strip()
        if not factory:
            continue
        ref = await db.get(PartsReference, factory)
        if ref:
            db.add(PartsManualItem(
                section_id=section.id,
                order_num=str(item_data.get("order_num", "")),
                factory_part_number=factory,
            ))
            restored += 1

    await db.commit()
    return {"restored_parts": restored, "section_code": entry.section_code, "section_name": entry.section_name}


# ── Rotation class — bulk import, coverage dashboard, unordered export ─────────

import openpyxl


_VALID_ROTATION = {"alta", "media", "baja"}
_ROTATION_MAP = {
    "alta": "alta", "a": "alta",
    "media": "media", "b": "media",
    "baja": "baja", "c": "baja",
}


def _normalize_part_code(s: str) -> str:
    return str(s).strip().upper().replace(" ", "")


def _find_header_row(sheet, expected_cols: set[str]) -> int:
    """Return the 1-based row index of the header row (case-insensitive match)."""
    for row_idx in range(1, min(sheet.max_row + 1, 20)):
        row_vals = {
            str(sheet.cell(row_idx, c).value or "").strip().lower()
            for c in range(1, sheet.max_column + 1)
        }
        if expected_cols & row_vals:
            return row_idx
    raise ValueError("No header row found with required columns")


def _build_col_map(sheet, header_row: int) -> dict[str, int]:
    """Return {normalized_col_name: col_index (1-based)} for the header row."""
    col_map: dict[str, int] = {}
    for c in range(1, sheet.max_column + 1):
        val = sheet.cell(header_row, c).value
        if val is not None:
            col_map[str(val).strip().lower()] = c
    return col_map


def _cell(sheet, row_idx: int, col_map: dict[str, int], *col_names: str):
    """Return the value of the first matching column name found in col_map."""
    for name in col_names:
        if name in col_map:
            return sheet.cell(row_idx, col_map[name]).value
    return None


@router.post("/admin/rotation-import", status_code=200)
async def import_rotation(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Bulk-assign rotation_class from an Excel file. Solo superadmin."""
    if not current_user.is_superadmin:
        raise HTTPException(status_code=403, detail="Solo superadmin")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    sheet = wb.active

    _CODE_ALIASES = {"part_code", "factory_part_number", "codigo", "código", "codigo_parte",
                     "referencia", "ref", "part number", "reference", "numero_parte", "numero parte"}
    _RC_ALIASES   = {"rotation_class", "rotacion", "rotación", "clase", "clase_rotacion",
                     "clase_rotación", "clase rotacion", "clase rotación", "tipo_rotacion",
                     "tipo_rotación", "tipo rotacion"}
    expected = _CODE_ALIASES | _RC_ALIASES
    try:
        header_row = _find_header_row(sheet, expected)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc))

    col_map = _build_col_map(sheet, header_row)

    # verify that at least one code column and one rotation column are present
    has_code = any(a in col_map for a in _CODE_ALIASES)
    has_rc   = any(a in col_map for a in _RC_ALIASES)
    if not has_code or not has_rc:
        missing = []
        if not has_code: missing.append("código de parte (ej: 'part_code' o 'codigo')")
        if not has_rc:   missing.append("clasificación (ej: 'rotation_class' o 'rotacion')")
        raise HTTPException(status_code=422, detail=f"Faltan columnas requeridas: {', '.join(missing)}")

    code_cols = list(_CODE_ALIASES)
    rc_cols   = list(_RC_ALIASES)

    updated, skipped, errors = 0, 0, []

    for row_idx in range(header_row + 1, sheet.max_row + 1):
        code_raw = _cell(sheet, row_idx, col_map, *code_cols)
        rc_raw   = _cell(sheet, row_idx, col_map, *rc_cols)

        if not code_raw and not rc_raw:
            # blank row — skip silently
            continue

        if not code_raw:
            skipped += 1
            errors.append({"row": row_idx, "code": None, "reason": "missing_part_code"})
            continue

        code_n = _normalize_part_code(code_raw)
        rc_n   = _ROTATION_MAP.get(str(rc_raw or "").strip().lower())

        if rc_n is None:
            skipped += 1
            errors.append({"row": row_idx, "code": code_n, "reason": "invalid_rotation_class"})
            continue

        res = await db.execute(
            sa_update(PartsReference)
            .where(PartsReference.factory_part_number == code_n)
            .where(PartsReference.rotation_class.is_(None))
            .values(rotation_class=rc_n)
        )
        if res.rowcount == 0:
            skipped += 1
            errors.append({"row": row_idx, "code": code_n, "reason": "part_not_found"})
        else:
            updated += 1

    await db.commit()
    return {"updated": updated, "skipped": skipped, "errors": errors}


# ── FOB preliminar desde PI ────────────────────────────────────────────────────

@router.post("/admin/fob-preliminary-import", status_code=200)
async def import_fob_preliminary(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Carga FOB de PI desde un Excel con columnas: referencia | pi_number | fob_price.
    1. Escribe fob_pi en cada spare_part_item correspondiente.
    2. Recalcula preliminary_fob en parts_references como promedio ponderado
       usando qty_ordered de spare_part_items.
    Guarda siempre como historia, independiente de si hay packing list.
    Solo superadmin.
    """
    if not current_user.is_superadmin:
        raise HTTPException(status_code=403, detail="Solo superadmin")

    content = await file.read()
    wb      = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    sheet   = wb.active

    _REF_ALIASES = {"referencia", "reference", "factory_part_number", "part_number",
                    "codigo", "código", "part_code", "ref", "numero_parte"}
    _PI_ALIASES  = {"pi_number", "pi", "numero_pi", "lot", "lot_identifier",
                    "lote", "pi number", "numero pi", "n_pi"}
    _FOB_ALIASES = {"fob", "fob_price", "precio_fob", "fob_pi",
                    "valor_fob", "fob price", "precio fob", "valor fob"}

    expected = _REF_ALIASES | _PI_ALIASES | _FOB_ALIASES
    try:
        header_row = _find_header_row(sheet, expected)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc))

    col_map = _build_col_map(sheet, header_row)
    missing = []
    if not any(a in col_map for a in _REF_ALIASES): missing.append("referencia")
    if not any(a in col_map for a in _PI_ALIASES):  missing.append("pi_number")
    if not any(a in col_map for a in _FOB_ALIASES): missing.append("fob_price")
    if missing:
        raise HTTPException(status_code=422, detail=f"Faltan columnas: {', '.join(missing)}")

    from collections import defaultdict
    from app.models.imports import SparePartLot as _SPL, SparePartItem as _SPI2

    parse_errors  = []
    items_updated = 0
    skipped_no_pi = 0
    skipped_no_item = 0

    # ── Paso 1: primera pasada para parsear filas ──────────────────────────────
    # Evita N+1 queries cargando todos los lotes de una vez con IN (...)
    parsed_rows: list[tuple[int, str, str | None, float]] = []  # (row_idx, fpn, lot_id, fob_val)
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        ref_raw = _cell(sheet, row_idx, col_map, *list(_REF_ALIASES))
        pi_raw  = _cell(sheet, row_idx, col_map, *list(_PI_ALIASES))
        fob_raw = _cell(sheet, row_idx, col_map, *list(_FOB_ALIASES))

        if not ref_raw and not pi_raw and not fob_raw:
            continue
        if not ref_raw or not fob_raw:
            parse_errors.append({"row": row_idx, "reason": "missing_ref_or_fob"})
            continue
        try:
            fob_val = float(str(fob_raw).replace(",", ".").strip())
        except (ValueError, TypeError):
            parse_errors.append({"row": row_idx, "ref": str(ref_raw), "reason": "invalid_fob_value"})
            continue

        fpn    = _normalize_part_code(ref_raw)
        lot_id = str(pi_raw).strip() if pi_raw else None
        parsed_rows.append((row_idx, fpn, lot_id, fob_val))

    # ── Paso 2: batch-load de lotes y sus items ────────────────────────────────
    lot_ids = {lot_id for _, _, lot_id, _ in parsed_rows if lot_id}
    lots_by_id: dict[str, _SPL] = {}
    if lot_ids:
        lot_rows = (await db.execute(
            select(_SPL).where(_SPL.lot_identifier.in_(lot_ids))
        )).scalars().all()
        lots_by_id = {l.lot_identifier: l for l in lot_rows}

    # Batch-load de items para todos los lotes encontrados
    lot_uuids = [l.id for l in lots_by_id.values()]
    items_by_lot: dict = defaultdict(list)  # lot_id (uuid) → [SparePartItem]
    if lot_uuids:
        all_items = (await db.execute(
            select(_SPI2).where(_SPI2.lot_id.in_(lot_uuids))
        )).scalars().all()
        for spi in all_items:
            items_by_lot[spi.lot_id].append(spi)

    # ── Paso 3: procesar filas con datos ya en memoria ─────────────────────────
    fob_accum: dict[str, list[tuple[float, int]]] = defaultdict(list)

    for _, fpn, lot_id, fob_val in parsed_rows:
        lot_obj = lots_by_id.get(lot_id) if lot_id else None

        if not lot_obj:
            skipped_no_pi += 1
            continue

        norm_fpn = func.upper(func.trim(fpn)).compile(compile_kwargs={"literal_binds": True}).string \
            if False else fpn.upper().strip().replace(' ', '')
        items_found = [
            spi for spi in items_by_lot.get(lot_obj.id, [])
            if spi.part_number.upper().strip().replace(' ', '') == norm_fpn
        ]

        if not items_found:
            skipped_no_item += 1
            continue

        qty_total = 0
        for item in items_found:
            item.fob_pi = fob_val
            qty_total  += (item.qty_ordered or 1)
            items_updated += 1

        fob_accum[fpn].append((fob_val, qty_total))

    # 4. Recalcular preliminary_fob en parts_references (promedio ponderado)
    refs_updated, refs_not_found = 0, 0
    for fpn, entries in fob_accum.items():
        ref = await db.get(PartsReference, fpn)
        if not ref:
            refs_not_found += 1
            continue
        total_cost = sum(f * q for f, q in entries)
        total_qty  = sum(q for _, q in entries)
        ref.preliminary_fob = round(total_cost / total_qty, 4)
        refs_updated += 1

    await db.commit()
    return {
        "items_updated":    items_updated,
        "refs_updated":     refs_updated,
        "skipped_no_pi":    skipped_no_pi,
        "skipped_no_item":  skipped_no_item,
        "refs_not_found":   refs_not_found,
        "parse_errors":     parse_errors,
    }


# ── Coverage dashboard ─────────────────────────────────────────────────────────

from pydantic import BaseModel as _BM


class CoverageBucket(_BM):
    rotation_class: str
    total: int
    aqui: int
    en_camino: int
    pedido: int
    no_pedidas: int
    pct_aqui: float
    pct_en_camino: float
    pct_pedido: float
    pct_no_pedidas: float


class CoverageResponse(_BM):
    sin_clasificar: int
    buckets: list[CoverageBucket]


@router.get("/admin/coverage", response_model=CoverageResponse)
async def get_coverage(
    model_code: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Resumen de cobertura de repuestos por rotation_class. Solo superadmin."""
    if not current_user.is_superadmin:
        raise HTTPException(status_code=403, detail="Solo superadmin")

    model_filter = """
        AND r.factory_part_number IN (
            SELECT i.factory_part_number
            FROM parts_manual_items i
            JOIN parts_manual_sections s ON s.id = i.section_id
            WHERE s.model_code = :model_code
        )
    """ if model_code else """
        AND EXISTS (
            SELECT 1 FROM parts_manual_items pmi
            WHERE pmi.factory_part_number = r.factory_part_number
        )
    """

    coverage_sql = text(f"""
        WITH
        part_all_codes AS (
            SELECT factory_part_number,
                   UPPER(TRIM(REPLACE(factory_part_number, ' ', ''))) AS norm_code
            FROM parts_references
            UNION ALL
            SELECT pr.factory_part_number,
                   UPPER(TRIM(REPLACE(
                       CASE WHEN jsonb_typeof(elem) = 'string' THEN elem #>> '{{}}'
                            ELSE elem->>'code'
                       END, ' ', ''))) AS norm_code
            FROM parts_references pr,
                 jsonb_array_elements(pr.prev_codes) AS elem
            WHERE jsonb_array_length(pr.prev_codes) > 0
              AND (
                  (jsonb_typeof(elem) = 'string'  AND (elem #>> '{{}}') != '')
                  OR (jsonb_typeof(elem) = 'object' AND elem->>'code' IS NOT NULL AND elem->>'code' != '')
              )
        ),
        aqui AS (
            SELECT UPPER(TRIM(REPLACE(part_number, ' ', ''))) AS pn
            FROM spare_part_availability
            WHERE qty_available > 0
            GROUP BY 1
        ),
        en_camino AS (
            SELECT UPPER(TRIM(REPLACE(spi.part_number, ' ', ''))) AS pn
            FROM spare_part_items spi
            JOIN spare_part_lots spl ON spl.id = spi.lot_id
            JOIN shipment_orders so  ON so.id  = spl.shipment_order_id
            WHERE spi.qty_received > 0
              AND spi.qty_physical IS NULL
              AND (so.bl_container IS NOT NULL OR spl.packing_list_received = true)
              AND UPPER(TRIM(REPLACE(spi.part_number, ' ', ''))) NOT IN (SELECT pn FROM aqui)
            UNION
            SELECT UPPER(TRIM(part_code)) AS pn
            FROM part_catalog
            WHERE public_price IS NOT NULL
              AND UPPER(TRIM(part_code)) NOT IN (SELECT pn FROM aqui)
        ),
        pedido AS (
            SELECT UPPER(TRIM(REPLACE(spi.part_number, ' ', ''))) AS pn
            FROM spare_part_items spi
            JOIN spare_part_lots spl ON spl.id = spi.lot_id
            WHERE (spl.packing_list_received = false OR spi.status IN ('BACKORDER', 'BACKORDER_PARCIAL'))
              AND spi.status != 'CANCELLED'
              AND UPPER(TRIM(REPLACE(spi.part_number, ' ', ''))) NOT IN (SELECT pn FROM aqui)
              AND UPPER(TRIM(REPLACE(spi.part_number, ' ', ''))) NOT IN (SELECT pn FROM en_camino)
            GROUP BY 1
        ),
        ref_coverage AS (
            SELECT r.factory_part_number, r.rotation_class,
                MAX(CASE
                    WHEN a.pn IS NOT NULL THEN 3
                    WHEN c.pn IS NOT NULL THEN 2
                    WHEN p.pn IS NOT NULL THEN 1
                    ELSE 0
                END) AS score
            FROM parts_references r
            JOIN part_all_codes pac ON pac.factory_part_number = r.factory_part_number
            LEFT JOIN aqui      a ON a.pn = pac.norm_code
            LEFT JOIN en_camino c ON c.pn = pac.norm_code
            LEFT JOIN pedido    p ON p.pn = pac.norm_code
            WHERE r.rotation_class IS NOT NULL
            {model_filter}
            GROUP BY r.factory_part_number, r.rotation_class
        ),
        coverage AS (
            SELECT rotation_class,
                COUNT(*)                               AS total,
                COUNT(CASE WHEN score = 3 THEN 1 END) AS aqui,
                COUNT(CASE WHEN score = 2 THEN 1 END) AS en_camino,
                COUNT(CASE WHEN score = 1 THEN 1 END) AS pedido,
                COUNT(CASE WHEN score = 0 THEN 1 END) AS no_pedidas
            FROM ref_coverage
            GROUP BY rotation_class
        )
        SELECT * FROM coverage
        ORDER BY CASE rotation_class WHEN 'alta' THEN 1 WHEN 'media' THEN 2 WHEN 'baja' THEN 3 ELSE 4 END
    """)

    sin_sql = text(f"""
        SELECT COUNT(*) FROM parts_references r
        WHERE rotation_class IS NULL
        {model_filter}
    """)

    params = {"model_code": model_code} if model_code else {}
    rows = (await db.execute(coverage_sql, params)).all()
    sin_clasificar = (await db.execute(sin_sql, params)).scalar_one()

    clases: list[CoverageBucket] = []
    for row in rows:
        total      = row.total
        aqui       = row.aqui
        en_camino  = row.en_camino
        pedido     = row.pedido
        no_pedidas = row.no_pedidas
        pct_aqui       = round((aqui      / total) * 100, 2) if total else 0.0
        pct_en_camino  = round((en_camino / total) * 100, 2) if total else 0.0
        pct_pedido     = round((pedido    / total) * 100, 2) if total else 0.0
        pct_no_pedidas = round((no_pedidas / total) * 100, 2) if total else 0.0
        clases.append(CoverageBucket(
            rotation_class=row.rotation_class,
            total=total,
            aqui=aqui,
            en_camino=en_camino,
            pedido=pedido,
            no_pedidas=no_pedidas,
            pct_aqui=pct_aqui,
            pct_en_camino=pct_en_camino,
            pct_pedido=pct_pedido,
            pct_no_pedidas=pct_no_pedidas,
        ))

    return CoverageResponse(sin_clasificar=int(sin_clasificar), buckets=clases)


# ── Análisis: baja/media rotación en estado pedido ────────────────────────────

class _LotQty(BaseModel):
    lot_identifier: str
    qty: int
    fob_unit: Optional[float] = None

class _LowRotItem(BaseModel):
    factory_part_number: str
    description: str
    description_es: Optional[str]
    rotation_class: str
    total_qty: int
    lots: list[_LotQty]
    models: list[str] = []
    fob_unit: Optional[float] = None
    total_fob: Optional[float] = None
    qty_stock: Optional[int] = None
    sugerido: Optional[int] = None
    flota_efectiva: Optional[float] = None
    demand_estimate: Optional[float] = None
    prev_codes: list[str] = []

class _LowRotResponse(BaseModel):
    items: list[_LowRotItem]
    total_references: int
    total_qty: int
    baja_count: int
    media_count: int
    alta_count: int


@router.get("/admin/analysis/low-rotation-ordered", response_model=_LowRotResponse)
async def low_rotation_ordered_analysis(
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Repuestos en pedido (aún cancelables/ajustables) por clase de rotación. Solo superadmin."""
    if not current_user.is_superadmin:
        raise HTTPException(status_code=403, detail="Solo superadmin")

    sql = text("""
        WITH
        aqui AS (
            SELECT UPPER(TRIM(REPLACE(part_number, ' ', ''))) AS pn
            FROM spare_part_availability
            WHERE qty_available > 0
            GROUP BY 1
        ),
        en_camino AS (
            SELECT UPPER(TRIM(REPLACE(spi.part_number, ' ', ''))) AS pn
            FROM spare_part_items spi
            JOIN spare_part_lots  spl ON spl.id = spi.lot_id
            JOIN shipment_orders  so  ON so.id  = spl.shipment_order_id
            WHERE spi.qty_received > 0
              AND spi.qty_physical IS NULL
              AND (so.bl_container IS NOT NULL OR spl.packing_list_received = true)
            UNION
            SELECT UPPER(TRIM(part_code)) AS pn
            FROM part_catalog
            WHERE public_price IS NOT NULL
        ),
        desc_es_src AS (
            SELECT DISTINCT ON (UPPER(TRIM(REPLACE(part_number, ' ', ''))))
                UPPER(TRIM(REPLACE(part_number, ' ', ''))) AS pn,
                description_es
            FROM spare_part_items
            WHERE description_es IS NOT NULL AND description_es != ''
            ORDER BY UPPER(TRIM(REPLACE(part_number, ' ', ''))), created_at ASC
        ),
        models_from_items AS (
            SELECT
                UPPER(TRIM(REPLACE(spi2.part_number, ' ', ''))) AS norm_fpn,
                array_agg(DISTINCT spi2.model_applicable ORDER BY spi2.model_applicable)
                    FILTER (WHERE spi2.model_applicable IS NOT NULL AND spi2.model_applicable != '') AS models
            FROM spare_part_items spi2
            GROUP BY 1
        ),
        models_per_part AS (
            -- Traduce el código de catálogo al nombre comercial UM vía vehicle_catalog_map,
            -- para poder combinarlo con models_from_items (que ya usa nombres comerciales).
            SELECT
                pmi.factory_part_number,
                array_agg(DISTINCT vcm.vehicle_model_pattern ORDER BY vcm.vehicle_model_pattern) AS models
            FROM parts_manual_items pmi
            JOIN parts_manual_sections pms ON pms.id = pmi.section_id
            JOIN vehicle_catalog_map vcm    ON vcm.catalog_model_code = pms.model_code
            GROUP BY pmi.factory_part_number
        ),
        part_all_codes AS (
            -- Maps every normalized code (current + prev_codes) to its canonical factory_part_number
            SELECT factory_part_number,
                   UPPER(TRIM(REPLACE(factory_part_number, ' ', ''))) AS norm_code
            FROM parts_references
            UNION ALL
            SELECT pr.factory_part_number,
                   UPPER(TRIM(REPLACE(
                       CASE WHEN jsonb_typeof(elem) = 'string' THEN elem #>> '{}'
                            ELSE elem->>'code'
                       END, ' ', ''))) AS norm_code
            FROM parts_references pr,
                 jsonb_array_elements(pr.prev_codes) AS elem
            WHERE jsonb_array_length(pr.prev_codes) > 0
              AND (
                  (jsonb_typeof(elem) = 'string'  AND (elem #>> '{}') != '')
                  OR (jsonb_typeof(elem) = 'object' AND elem->>'code' IS NOT NULL AND elem->>'code' != '')
              )
        ),
        stock_confirmado AS (
            -- Consolidates stock across current and prev_codes into one qty per canonical FPN
            SELECT pac.factory_part_number AS fpn, SUM(sc_raw.qs) AS qty_stock
            FROM (
                SELECT
                    UPPER(TRIM(REPLACE(part_number, ' ', ''))) AS pn,
                    SUM(qty_available) AS qs
                FROM spare_part_availability
                GROUP BY 1
                UNION ALL
                SELECT
                    UPPER(TRIM(REPLACE(spi_s.part_number, ' ', ''))) AS pn,
                    SUM(spi_s.qty_received) AS qs
                FROM spare_part_items spi_s
                JOIN spare_part_lots spl_s ON spl_s.id = spi_s.lot_id
                WHERE spl_s.packing_list_received = TRUE
                  AND spi_s.qty_physical IS NULL
                  AND spi_s.qty_received > 0
                GROUP BY 1
            ) sc_raw
            JOIN part_all_codes pac ON pac.norm_code = sc_raw.pn
            GROUP BY pac.factory_part_number
        ),
        sugerido_params AS (
            SELECT
                COALESCE(MAX(CASE WHEN key = 'sugerido.rate_alta'             THEN value::float END), 20.0) AS rate_alta,
                COALESCE(MAX(CASE WHEN key = 'sugerido.rate_media'            THEN value::float END), 10.0) AS rate_media,
                COALESCE(MAX(CASE WHEN key = 'sugerido.rate_baja'             THEN value::float END), 1.0)  AS rate_baja,
                COALESCE(MAX(CASE WHEN key = 'sugerido.pending_moto_factor'   THEN value::float END), 0.5)  AS pending_moto_factor,
                COALESCE(MAX(CASE WHEN key = 'sugerido.lead_time_months'     THEN value::float END), 3.0) AS lead_time_months,
                COALESCE(MAX(CASE WHEN key = 'sugerido.review_period_months' THEN value::float END), 3.0) AS review_period_months
            FROM system_config
            WHERE key IN ('sugerido.rate_alta', 'sugerido.rate_media', 'sugerido.rate_baja', 'sugerido.pending_moto_factor', 'sugerido.lead_time_months', 'sugerido.review_period_months')
        ),
        fleet_by_model AS (
            SELECT model, SUM(fleet_count) AS fleet_count
            FROM (
                SELECT COALESCE(mu.model, so.model) AS model, 1.0 AS fleet_count
                FROM shipment_moto_units mu
                JOIN shipment_orders so ON so.id = mu.shipment_order_id
                WHERE COALESCE(mu.model, so.model) IS NOT NULL
                UNION ALL
                SELECT so.model AS model, so.qty_numeric * sp.pending_moto_factor AS fleet_count
                FROM shipment_orders so CROSS JOIN sugerido_params sp
                WHERE NOT so.is_spare_part
                  AND so.computed_status IN ('en_preparacion', 'listo_fabrica', 'en_transito')
                  AND so.model IS NOT NULL
                UNION ALL
                SELECT so.model AS model, so.qty_numeric * 1.0 AS fleet_count
                FROM shipment_orders so
                WHERE NOT so.is_spare_part
                  AND so.computed_status = 'en_destino'
                  AND so.model IS NOT NULL
                  AND NOT EXISTS (
                      SELECT 1 FROM shipment_moto_units mu2
                      WHERE mu2.shipment_order_id = so.id
                  )
            ) fbm_raw
            GROUP BY model
        ),
        fleet_per_reference AS (
            -- Modelos que aplican a cada referencia = catálogo (manual PDF) UNIDO con
            -- model_applicable (modelo declarado al importar el pedido de repuestos).
            -- UNION (no ALL) evita sumar la misma flota dos veces cuando el modelo
            -- coincide en ambas fuentes para la misma referencia.
            SELECT fpn AS factory_part_number, SUM(fleet_count) AS flota_efectiva
            FROM (
                SELECT DISTINCT pmi.factory_part_number AS fpn,
                                fbm.model,
                                fbm.fleet_count
                FROM fleet_by_model fbm
                JOIN vehicle_catalog_map vcm ON LOWER(TRIM(vcm.vehicle_model_pattern)) = LOWER(TRIM(fbm.model))
                JOIN parts_manual_sections pms ON pms.model_code = vcm.catalog_model_code
                JOIN parts_manual_items pmi ON pmi.section_id = pms.id

                UNION

                SELECT DISTINCT pr_mfi.factory_part_number AS fpn,
                                fbm.model,
                                fbm.fleet_count
                FROM fleet_by_model fbm
                JOIN spare_part_items spi ON LOWER(TRIM(spi.model_applicable)) = LOWER(TRIM(fbm.model))
                JOIN parts_references pr_mfi
                    ON UPPER(TRIM(REPLACE(spi.part_number, ' ', ''))) = UPPER(TRIM(REPLACE(pr_mfi.factory_part_number, ' ', '')))
                WHERE spi.model_applicable IS NOT NULL AND spi.model_applicable != ''
            ) deduped
            GROUP BY fpn
        ),
        total_qty_per_fpn AS (
            SELECT UPPER(TRIM(REPLACE(spi.part_number, ' ', ''))) AS fpn,
                   SUM(spi.qty_ordered) AS total_qty
            FROM spare_part_items spi
            JOIN spare_part_lots spl ON spl.id = spi.lot_id
            WHERE spl.packing_list_received = FALSE
              AND spi.status != 'CANCELLED'
            GROUP BY 1
        )
        SELECT
            pr.factory_part_number,
            pr.description,
            COALESCE(pr.description_es_manual, d.description_es) AS description_es,
            pr.rotation_class,
            spl.lot_identifier,
            SUM(spi.qty_ordered)::int AS qty,
            (
                SELECT array_agg(DISTINCT m ORDER BY m)
                FROM unnest(COALESCE(mfi.models, ARRAY[]::text[]) || COALESCE(mp.models, ARRAY[]::text[])) AS m
            ) AS models,
            MAX(COALESCE(spi.unit_price, spi.fob_pi))::float AS fob_unit,
            MAX(COALESCE(sc.qty_stock, 0))::int AS qty_stock,
            pr.prev_codes,
            MAX(fpr.flota_efectiva)::float AS flota_efectiva,
            MAX(CASE
                WHEN fpr.flota_efectiva IS NOT NULL THEN
                    (fpr.flota_efectiva / 100.0)
                    * CASE pr.rotation_class
                        WHEN 'alta'  THEN sp.rate_alta
                        WHEN 'media' THEN sp.rate_media
                        WHEN 'baja'  THEN sp.rate_baja
                        ELSE 0
                    END
                    * (1.0 + sp.lead_time_months / NULLIF(sp.review_period_months, 0))
                ELSE NULL
            END)::float AS demand_estimate,
            MAX(CASE
                WHEN fpr.flota_efectiva IS NOT NULL THEN
                    GREATEST(0, CEIL(
                        (fpr.flota_efectiva / 100.0)
                        * CASE pr.rotation_class
                            WHEN 'alta'  THEN sp.rate_alta
                            WHEN 'media' THEN sp.rate_media
                            WHEN 'baja'  THEN sp.rate_baja
                            ELSE 0
                        END
                        * (1.0 + sp.lead_time_months / NULLIF(sp.review_period_months, 0))
                        - COALESCE(sc.qty_stock, 0)
                    ))
                ELSE NULL
            END)::int AS sugerido
        FROM parts_references pr
        JOIN spare_part_items spi
            ON UPPER(TRIM(REPLACE(spi.part_number, ' ', ''))) = UPPER(TRIM(pr.factory_part_number))
        JOIN spare_part_lots spl ON spl.id = spi.lot_id
        LEFT JOIN desc_es_src d         ON d.pn        = UPPER(TRIM(pr.factory_part_number))
        LEFT JOIN models_from_items mfi ON mfi.norm_fpn = UPPER(TRIM(pr.factory_part_number))
        LEFT JOIN models_per_part mp    ON mp.factory_part_number = pr.factory_part_number
        LEFT JOIN stock_confirmado sc   ON sc.fpn       = pr.factory_part_number
        LEFT JOIN fleet_per_reference fpr ON fpr.factory_part_number = pr.factory_part_number
        LEFT JOIN total_qty_per_fpn tqf   ON tqf.fpn = UPPER(TRIM(REPLACE(pr.factory_part_number, ' ', '')))
        CROSS JOIN sugerido_params sp
        WHERE pr.rotation_class IN ('baja', 'media', 'alta')
          AND spl.packing_list_received = FALSE
          AND spi.status != 'CANCELLED'
        GROUP BY
            pr.factory_part_number, pr.description,
            COALESCE(pr.description_es_manual, d.description_es),
            pr.rotation_class, spl.lot_identifier,
            mfi.models, mp.models,
            pr.prev_codes
        ORDER BY pr.rotation_class, pr.factory_part_number, spl.lot_identifier
    """)

    rows = (await db.execute(sql)).all()

    from collections import defaultdict  # noqa: F401
    parts_map: dict[str, dict] = {}
    for row in rows:
        fpn = row.factory_part_number
        if fpn not in parts_map:
            raw_prev = row.prev_codes or []
            prev_codes = [
                e["code"] if isinstance(e, dict) and "code" in e else e
                for e in raw_prev
                if (isinstance(e, dict) and e.get("code")) or (isinstance(e, str) and e)
            ]
            parts_map[fpn] = {
                'factory_part_number': fpn,
                'description': row.description or '',
                'description_es': row.description_es,
                'rotation_class': row.rotation_class,
                'lots': [],
                'total_qty': 0,
                'models': list(row.models) if row.models else [],
                'qty_stock': int(row.qty_stock) if row.qty_stock else 0,
                'sugerido': int(row.sugerido) if row.sugerido is not None else None,
                'flota_efectiva': float(row.flota_efectiva) if row.flota_efectiva is not None else None,
                'demand_estimate': float(row.demand_estimate) if row.demand_estimate is not None else None,
                'prev_codes': prev_codes,
            }
        qty = int(row.qty) if row.qty else 0
        fob_unit = float(row.fob_unit) if row.fob_unit is not None else None
        parts_map[fpn]['lots'].append({'lot_identifier': row.lot_identifier, 'qty': qty, 'fob_unit': fob_unit})
        parts_map[fpn]['total_qty'] += qty
        if fob_unit is not None:
            parts_map[fpn]['_fob_weighted'] = parts_map[fpn].get('_fob_weighted', 0) + fob_unit * qty
            parts_map[fpn]['_fob_qty'] = parts_map[fpn].get('_fob_qty', 0) + qty

    for part in parts_map.values():
        fob_qty = part.pop('_fob_qty', 0)
        fob_weighted = part.pop('_fob_weighted', 0)
        if fob_qty > 0:
            part['fob_unit'] = round(fob_weighted / fob_qty, 4)
            part['total_fob'] = round(fob_weighted, 2)

    items = sorted(
        parts_map.values(),
        key=lambda x: ({'baja': 0, 'media': 1, 'alta': 2}.get(x['rotation_class'], 3), -x['total_qty']),
    )

    baja_count  = sum(1 for i in items if i['rotation_class'] == 'baja')
    media_count = sum(1 for i in items if i['rotation_class'] == 'media')
    alta_count  = sum(1 for i in items if i['rotation_class'] == 'alta')
    total_qty   = sum(i['total_qty'] for i in items)

    return _LowRotResponse(
        items=[_LowRotItem(**i) for i in items],
        total_references=len(items),
        total_qty=total_qty,
        baja_count=baja_count,
        media_count=media_count,
        alta_count=alta_count,
    )


# ── Comparativa de referencias compartidas entre modelos ─────────────────────

class _ComparativaModelPrice(BaseModel):
    fob_cost: float
    is_confirmed: bool  # True = unit_price (packing list), False = fob_pi (PI)

class _ComparativaItem(BaseModel):
    factory_part_number: str
    description: str
    description_es: Optional[str]
    rotation_class: Optional[str]
    prices: dict[str, _ComparativaModelPrice]  # {model: price_info}

class _ComparativaResponse(BaseModel):
    items: list[_ComparativaItem]
    models: list[str]
    total: int


@router.get("/admin/analysis/comparativa", response_model=_ComparativaResponse)
async def comparativa_referencias(
    model_filter: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Referencias compartidas entre modelos con comparativa de precios FOB. Solo superadmin."""
    if not current_user.is_superadmin:
        raise HTTPException(status_code=403, detail="Solo superadmin")

    sql = text("""
        WITH price_per_model AS (
            SELECT
                UPPER(TRIM(REPLACE(spi.part_number, ' ', ''))) AS pn,
                so.model                                         AS model,
                MAX(COALESCE(spi.unit_price, spi.fob_pi))::float AS fob_cost,
                BOOL_OR(spi.unit_price IS NOT NULL)              AS is_confirmed
            FROM spare_part_items   spi
            JOIN spare_part_lots    spl ON spl.id = spi.lot_id
            JOIN shipment_orders    so  ON so.id  = spl.shipment_order_id
            WHERE COALESCE(spi.unit_price, spi.fob_pi) IS NOT NULL
              AND so.model IS NOT NULL
              AND so.is_spare_part = TRUE
            GROUP BY 1, 2
        ),
        multi_model AS (
            SELECT pn
            FROM   price_per_model
            GROUP  BY pn
            HAVING COUNT(DISTINCT model) >= 2
        )
        SELECT
            pr.factory_part_number,
            pr.description,
            pr.description_es_manual,
            pr.rotation_class,
            ppm.model,
            ppm.fob_cost,
            ppm.is_confirmed
        FROM   parts_references  pr
        JOIN   multi_model       mm  ON mm.pn  = UPPER(TRIM(REPLACE(pr.factory_part_number, ' ', '')))
        JOIN   price_per_model   ppm ON ppm.pn = UPPER(TRIM(REPLACE(pr.factory_part_number, ' ', '')))
        ORDER  BY pr.factory_part_number, ppm.model
    """)

    rows = (await db.execute(sql)).all()

    parts_map: dict[str, dict] = {}
    for row in rows:
        fpn = row.factory_part_number
        if fpn not in parts_map:
            parts_map[fpn] = {
                'factory_part_number': fpn,
                'description': row.description or '',
                'description_es': row.description_es_manual,
                'rotation_class': row.rotation_class,
                'prices': {},
            }
        parts_map[fpn]['prices'][row.model] = {
            'fob_cost': round(float(row.fob_cost), 4),
            'is_confirmed': bool(row.is_confirmed),
        }

    items = list(parts_map.values())

    if model_filter:
        items = [i for i in items if model_filter in i['prices']]

    items.sort(key=lambda x: x['factory_part_number'])

    all_models = sorted({
        m for i in items for m in i['prices']
    })

    return _ComparativaResponse(
        items=[_ComparativaItem(**i) for i in items],
        models=all_models,
        total=len(items),
    )


# ── Decisiones de análisis (persistencia en DB) ───────────────────────────────

_POD = PartOrderDecision


@router.get("/admin/analysis/decisions")
async def get_decisions(
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Devuelve todas las decisiones guardadas como {fpn::lot_id: {decision, new_quantity, change_note}}."""
    if not current_user.is_superadmin:
        raise HTTPException(status_code=403, detail="Solo superadmin")
    rows = (await db.execute(
        select(_POD).where(_POD.executed_at.is_(None))
    )).scalars().all()
    return {
        f"{r.factory_part_number}::{r.lot_identifier}": {
            "decision":          r.decision,
            "new_quantity":      r.new_quantity,
            "original_quantity": r.original_quantity,
            "change_note":       r.change_note,
        }
        for r in rows
    }


class _DecisionIn(BaseModel):
    factory_part_number: str
    lot_identifier: str
    decision: str  # 'cancelar' | 'cambiar' | '' (vacío = borrar)
    new_quantity: Optional[int] = Field(None, ge=1)
    original_quantity: Optional[int] = Field(None, ge=1)
    change_note: Optional[str] = None


@router.post("/admin/analysis/decisions", status_code=200)
async def save_decision(
    body: _DecisionIn,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Guarda o elimina una decisión para un par (referencia, PI)."""
    if not current_user.is_superadmin:
        raise HTTPException(status_code=403, detail="Solo superadmin")

    existing = await db.get(_POD, (body.factory_part_number, body.lot_identifier))

    if not body.decision:
        if existing:
            await db.delete(existing)
            await db.commit()
        return {"status": "deleted"}

    if existing:
        existing.decision          = body.decision
        existing.new_quantity      = body.new_quantity
        existing.change_note       = body.change_note
        existing.updated_at        = datetime.utcnow()
        existing.updated_by        = _uuid.UUID(current_user.user_id)
        existing.executed_at       = None  # allow re-execution after a previous run
        # Preserve original_quantity: only set on first save, never overwrite
        if existing.original_quantity is None and body.original_quantity is not None:
            existing.original_quantity = body.original_quantity
    else:
        db.add(_POD(
            factory_part_number=body.factory_part_number,
            lot_identifier=body.lot_identifier,
            decision=body.decision,
            new_quantity=body.new_quantity,
            original_quantity=body.original_quantity,
            change_note=body.change_note,
            updated_by=_uuid.UUID(current_user.user_id),
        ))
    await db.commit()
    return {"status": "saved"}


@router.post("/admin/analysis/decisions/execute", status_code=200)
async def execute_adjustments(
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Ejecuta todas las decisiones pendientes:
    - 'cancelar': marca ítems CANCELLED y qty_pending = 0
    - 'cambiar': actualiza qty_ordered/qty_pending si new_quantity está definido
    Solo actúa sobre lotes SIN packing list recibido.
    Recalcula preliminary_fob para las referencias afectadas.
    """
    if not current_user.is_superadmin:
        raise HTTPException(status_code=403, detail="Solo superadmin")

    from app.models.imports import SparePartLot, SparePartItem
    from app.models.parts_manual import PartsReference

    all_pending = (await db.execute(
        select(_POD).where(_POD.executed_at.is_(None))
    )).scalars().all()

    cancelar_decisions = [d for d in all_pending if d.decision == 'cancelar']
    cambiar_decisions  = [d for d in all_pending if d.decision == 'cambiar']
    revisado_decisions = [d for d in all_pending if d.decision == 'revisado']

    if not all_pending:
        return {"cancelled_items": 0, "changed_items": 0, "refs_updated": 0, "affected_references": [], "skipped_lots": []}

    now = datetime.utcnow()
    cancelled_items = 0
    changed_items   = 0
    skipped_lots: list[str] = []
    affected_fpns: set[str] = set()

    for decision in cancelar_decisions:
        fpn = decision.factory_part_number
        lot_id_str = decision.lot_identifier

        # Buscar lote por lot_identifier
        lot = (await db.execute(
            select(SparePartLot).where(SparePartLot.lot_identifier == lot_id_str)
        )).scalar_one_or_none()

        if not lot:
            skipped_lots.append(f"{lot_id_str} (no encontrado)")
            continue

        # Guardia crítica: solo cancelar si el packing list NO fue recibido
        if lot.packing_list_received:
            skipped_lots.append(f"{lot_id_str} (packing list ya recibido)")
            decision.executed_at = now  # consume so it never re-surfaces
            continue

        # Normalización de part_number — misma que usa el sistema en todos lados
        norm_fpn = fpn.upper().strip().replace(' ', '')

        # Buscar ítems del lote que matcheen la referencia y estén cancelables
        items_result = await db.execute(
            select(SparePartItem).where(
                SparePartItem.lot_id == lot.id,
                SparePartItem.status.notin_(['CANCELLED', 'RECEIVED', 'DECLARED', 'BACKORDER']),
                func.upper(func.trim(func.replace(SparePartItem.part_number, ' ', ''))) == norm_fpn,
            )
        )
        items = items_result.scalars().all()

        for item in items:
            item.status = 'CANCELLED'
            item.qty_ordered = 0
            item.qty_pending = 0
            item.updated_at = now
            cancelled_items += 1

        decision.executed_at = now
        if items:
            affected_fpns.add(fpn)

    # ── Procesar decisiones 'cambiar' ─────────────────────────────────────────
    for decision in cambiar_decisions:
        fpn       = decision.factory_part_number
        lot_id_str = decision.lot_identifier

        lot = (await db.execute(
            select(SparePartLot).where(SparePartLot.lot_identifier == lot_id_str)
        )).scalar_one_or_none()

        if not lot:
            skipped_lots.append(f"{lot_id_str} (no encontrado)")
            continue

        if lot.packing_list_received:
            skipped_lots.append(f"{lot_id_str} (packing list ya recibido)")
            decision.executed_at = now  # consume so it never re-surfaces
            continue

        qty_changed = (
            decision.new_quantity is not None
            and decision.new_quantity != decision.original_quantity
        )
        if qty_changed:
            norm_fpn = fpn.upper().strip().replace(' ', '')
            items_result = await db.execute(
                select(SparePartItem).where(
                    SparePartItem.lot_id == lot.id,
                    SparePartItem.status.notin_(['CANCELLED', 'RECEIVED', 'DECLARED', 'BACKORDER']),
                    func.upper(func.trim(func.replace(SparePartItem.part_number, ' ', ''))) == norm_fpn,
                )
            )
            for item in items_result.scalars().all():
                item.qty_ordered = decision.new_quantity
                item.qty_pending = decision.new_quantity
                item.updated_at  = now
                changed_items += 1
            affected_fpns.add(fpn)

        decision.executed_at = now

    # Marcar decisiones 'revisado' como ejecutadas (no modifican datos, solo se cierran)
    for decision in revisado_decisions:
        decision.executed_at = now

    # Recalcular preliminary_fob para todas las referencias afectadas
    refs_updated = 0
    for fpn in affected_fpns:
        ref = await db.get(PartsReference, fpn)
        if not ref:
            continue

        norm_fpn = fpn.upper().strip().replace(' ', '')

        # Todos los ítems activos (no cancelados) con fob_pi para esta referencia
        active_items_result = await db.execute(
            select(SparePartItem).where(
                func.upper(func.trim(func.replace(SparePartItem.part_number, ' ', ''))) == norm_fpn,
                SparePartItem.status != 'CANCELLED',
                SparePartItem.fob_pi.isnot(None),
                SparePartItem.qty_ordered > 0,
            )
        )
        active_items = active_items_result.scalars().all()

        if active_items:
            total_cost = sum(float(i.fob_pi) * (i.qty_ordered or 0) for i in active_items)
            total_qty = sum(i.qty_ordered or 0 for i in active_items)
            ref.preliminary_fob = round(total_cost / total_qty, 4) if total_qty > 0 else None
        else:
            ref.preliminary_fob = None

        refs_updated += 1

    await db.commit()

    return {
        "cancelled_items": cancelled_items,
        "changed_items":   changed_items,
        "refs_updated":    refs_updated,
        "affected_references": list(affected_fpns),
        "skipped_lots": skipped_lots,
    }


class _ExportRequest(BaseModel):
    marked: dict[str, str] = {}  # { "fpn::lot_id": "cancelar" | "cambiar" | "revisado" }


_DECISION_EXPORT_LABELS = {"cancelar": "CANCELAR", "cambiar": "CAMBIAR", "revisado": "APROBADO"}


def _decision_export_label(decision: str) -> str:
    """Etiqueta de la columna Decisión en el Excel de Ajuste de Pedidos.
    Las referencias sin decisión tomada quedan en blanco (no un guion),
    para distinguirlas visualmente de las que sí tienen una decisión."""
    return _DECISION_EXPORT_LABELS.get(decision, "")


@router.post("/admin/analysis/low-rotation-ordered/export")
async def export_low_rotation_ordered(
    body: _ExportRequest,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Exporta análisis de baja/media rotación a Excel con decisiones marcadas."""
    if not current_user.is_superadmin:
        raise HTTPException(status_code=403, detail="Solo superadmin")

    analysis = await low_rotation_ordered_analysis(db=db, current_user=current_user)

    db_decisions = {
        f"{r.factory_part_number}::{r.lot_identifier}": r
        for r in (await db.execute(select(_POD).where(_POD.executed_at.is_(None)))).scalars().all()
    }

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Para Revisar"

    # Estilos
    hdr_font  = Font(bold=True, color="FFFFFF", size=9)
    hdr_fill  = PatternFill("solid", fgColor="1e1e2e")
    center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left      = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin_side = Side(style="thin", color="444455")
    thin_brd  = Border(bottom=Side(style="thin", color="333344"))

    fill_cancelar = PatternFill("solid", fgColor="3b1219")
    fill_cambiar  = PatternFill("solid", fgColor="3b2e0a")
    fill_alt      = PatternFill("solid", fgColor="12121e")

    headers = ["Rotación", "Código", "Descripción", "Modelos", "PI Number", "Cantidad", "Total ref.", "Sugerido", "Decisión", "Nueva Cant.", "Observación"]
    col_widths = [10, 20, 35, 25, 18, 10, 10, 12, 12, 12, 40]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 22

    row = 2
    for item in analysis.items:
        models_str = ", ".join(item.models) if item.models else "—"
        for lot in item.lots:
            key = f"{item.factory_part_number}::{lot.lot_identifier}"
            decision = body.marked.get(key, "")
            decision_label = _decision_export_label(decision)
            db_dec = db_decisions.get(key)

            if decision == "cancelar":
                row_fill = fill_cancelar
                dec_font = Font(bold=True, color="F87171", size=9)
            elif decision == "cambiar":
                row_fill = fill_cambiar
                dec_font = Font(bold=True, color="FCD34D", size=9)
            else:
                row_fill = None
                dec_font = Font(color="888899", size=9)

            new_qty    = db_dec.new_quantity if (db_dec and decision == "cambiar") else None
            obs        = db_dec.change_note  if (db_dec and decision == "cambiar") else None

            values = [
                item.rotation_class.upper(),
                item.factory_part_number,
                item.description_es or item.description,
                models_str,
                lot.lot_identifier,
                lot.qty,
                item.total_qty,
                item.sugerido if item.sugerido is not None else "—",
                decision_label,
                new_qty or "",
                obs or "",
            ]
            aligns = [center, left, left, left, left, center, center, center, center, center, left]

            for ci, (val, aln) in enumerate(zip(values, aligns), 1):
                cell = ws.cell(row=row, column=ci, value=val)
                cell.alignment = aln
                cell.font = dec_font if ci == 9 else Font(size=9)
                cell.border = thin_brd
                if row_fill:
                    cell.fill = row_fill

            row += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{row - 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from fastapi.responses import StreamingResponse
    filename = f"analisis_repuestos_{datetime.utcnow().date()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Unordered parts export ─────────────────────────────────────────────────────

@router.get("/admin/coverage/unordered")
async def export_unordered(
    rotation_class: Optional[Literal['alta', 'media', 'baja', 'all']] = None,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Exporta repuestos clasificados sin SparePartItems a Excel. Solo superadmin."""
    if not current_user.is_superadmin:
        raise HTTPException(status_code=403, detail="Solo superadmin")

    params: dict = {}
    rc_clause = ""
    if rotation_class and rotation_class != 'all':
        rc_clause = "AND r.rotation_class = :rc"
        params["rc"] = rotation_class

    # Misma lógica que GET /admin/coverage (score=0 == "no pedidas"): considera
    # códigos previos/sustituidos (prev_codes), la vista real de disponibilidad
    # (spare_part_availability, ya descuenta lo despachado por remisión), y
    # excluye cancelados — para que este Excel coincida con la tarjeta.
    unordered_sql = text(f"""
        WITH
        part_all_codes AS (
            SELECT factory_part_number,
                   UPPER(TRIM(REPLACE(factory_part_number, ' ', ''))) AS norm_code
            FROM parts_references
            UNION ALL
            SELECT pr.factory_part_number,
                   UPPER(TRIM(REPLACE(
                       CASE WHEN jsonb_typeof(elem) = 'string' THEN elem #>> '{{}}'
                            ELSE elem->>'code'
                       END, ' ', ''))) AS norm_code
            FROM parts_references pr,
                 jsonb_array_elements(pr.prev_codes) AS elem
            WHERE jsonb_array_length(pr.prev_codes) > 0
              AND (
                  (jsonb_typeof(elem) = 'string'  AND (elem #>> '{{}}') != '')
                  OR (jsonb_typeof(elem) = 'object' AND elem->>'code' IS NOT NULL AND elem->>'code' != '')
              )
        ),
        aqui AS (
            SELECT UPPER(TRIM(REPLACE(part_number, ' ', ''))) AS pn
            FROM spare_part_availability
            WHERE qty_available > 0
            GROUP BY 1
        ),
        en_camino AS (
            SELECT UPPER(TRIM(REPLACE(spi.part_number, ' ', ''))) AS pn
            FROM spare_part_items spi
            JOIN spare_part_lots spl ON spl.id = spi.lot_id
            JOIN shipment_orders so  ON so.id  = spl.shipment_order_id
            WHERE spi.qty_received > 0
              AND spi.qty_physical IS NULL
              AND (so.bl_container IS NOT NULL OR spl.packing_list_received = true)
              AND UPPER(TRIM(REPLACE(spi.part_number, ' ', ''))) NOT IN (SELECT pn FROM aqui)
            UNION
            SELECT UPPER(TRIM(part_code)) AS pn
            FROM part_catalog
            WHERE public_price IS NOT NULL
              AND UPPER(TRIM(part_code)) NOT IN (SELECT pn FROM aqui)
        ),
        pedido AS (
            SELECT UPPER(TRIM(REPLACE(spi.part_number, ' ', ''))) AS pn
            FROM spare_part_items spi
            JOIN spare_part_lots spl ON spl.id = spi.lot_id
            WHERE (spl.packing_list_received = false OR spi.status IN ('BACKORDER', 'BACKORDER_PARCIAL'))
              AND spi.status != 'CANCELLED'
              AND UPPER(TRIM(REPLACE(spi.part_number, ' ', ''))) NOT IN (SELECT pn FROM aqui)
              AND UPPER(TRIM(REPLACE(spi.part_number, ' ', ''))) NOT IN (SELECT pn FROM en_camino)
            GROUP BY 1
        ),
        ref_coverage AS (
            SELECT r.factory_part_number, r.rotation_class,
                MAX(CASE
                    WHEN a.pn IS NOT NULL THEN 3
                    WHEN c.pn IS NOT NULL THEN 2
                    WHEN p.pn IS NOT NULL THEN 1
                    ELSE 0
                END) AS score
            FROM parts_references r
            JOIN part_all_codes pac ON pac.factory_part_number = r.factory_part_number
            LEFT JOIN aqui      a ON a.pn = pac.norm_code
            LEFT JOIN en_camino c ON c.pn = pac.norm_code
            LEFT JOIN pedido    p ON p.pn = pac.norm_code
            WHERE r.rotation_class IS NOT NULL
              AND EXISTS (
                  SELECT 1 FROM parts_manual_items pmi
                  WHERE pmi.factory_part_number = r.factory_part_number
              )
            GROUP BY r.factory_part_number, r.rotation_class
        )
        SELECT
            r.factory_part_number,
            r.um_part_number,
            r.description,
            r.description_es_manual,
            r.rotation_class,
            r.avg_fob_cost,
            r.preliminary_fob,
            CASE
                WHEN r.avg_fob_cost IS NOT NULL THEN 'Confirmado'
                WHEN r.preliminary_fob IS NOT NULL THEN 'Preliminar (PI)'
                ELSE NULL
            END AS costo_origen,
            (
                SELECT STRING_AGG(DISTINCT s.model_code, ', ' ORDER BY s.model_code)
                FROM parts_manual_items i
                JOIN parts_manual_sections s ON s.id = i.section_id
                WHERE i.factory_part_number = r.factory_part_number
            ) AS models
        FROM ref_coverage rc
        JOIN parts_references r ON r.factory_part_number = rc.factory_part_number
        WHERE rc.score = 0
          {rc_clause}
        ORDER BY r.rotation_class, r.factory_part_number
    """)

    rows = (await db.execute(unordered_sql, params)).all()

    import openpyxl as _xl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = _xl.Workbook()
    ws = wb.active
    ws.title = "No Pedidas"

    headers = [
        "Código fábrica",
        "Descripción",
        "Descripción ES",
        "Rotación",
        "Costo FOB USD",
        "Origen costo",
        "Modelos",
    ]
    col_widths = [22, 40, 40, 10, 14, 18, 30]

    hdr_font = Font(bold=True, color="FFFFFF", size=9)
    hdr_fill = PatternFill("solid", fgColor="1e1e2e")
    center   = Alignment(horizontal="center", vertical="center")
    left     = Alignment(horizontal="left",   vertical="center")

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(ci)].width = w

    fill_prel = PatternFill("solid", fgColor="3b2e0a")  # naranja oscuro para preliminares
    fill_alt  = PatternFill("solid", fgColor="12121e")

    for ri, row in enumerate(rows, 2):
        fob = row.avg_fob_cost if row.avg_fob_cost is not None else row.preliminary_fob
        is_prel = row.avg_fob_cost is None and row.preliminary_fob is not None
        row_fill = fill_prel if is_prel else (fill_alt if ri % 2 == 0 else None)

        values = [
            row.factory_part_number,
            row.description,
            row.description_es_manual,
            row.rotation_class.upper() if row.rotation_class else None,
            float(fob) if fob is not None else None,
            row.costo_origen,
            row.models,
        ]
        aligns = [left, left, left, center, center, center, left]

        for ci, (val, aln) in enumerate(zip(values, aligns), 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = aln
            cell.font = Font(size=9, color="FCD34D" if is_prel and ci == 6 else "FFFFFF")
            if row_fill:
                cell.fill = row_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{len(rows) + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    rc_label = rotation_class or "all"
    date_str = datetime.utcnow().strftime("%Y-%m-%d")
    filename = f"unordered_parts_{rc_label}_{date_str}.xlsx"

    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


