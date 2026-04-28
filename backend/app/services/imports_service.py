import io
import uuid
import logging
from datetime import datetime, timedelta
from typing import Optional

import openpyxl
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select

from app.models.imports import (
    ShipmentOrder, ShipmentMotoUnit, SparePartLot, SparePartItem,
    PackingList, PackingListItem, ReconciliationResult, Backorder, ImportAuditLog,
)
from app.api.deps import CurrentUser
from app.schemas.imports import ImportExcelResult

logger = logging.getLogger(__name__)

# Columnas esperadas en el Excel de Shipment Status
SHIPMENT_EXPECTED_COLS = {
    "cycle", "pi_number", "pi number", "invoice_number", "invoice number",
    "model", "qty", "etd", "eta", "etr", "etl"
}

# Columnas esperadas en packing list de motos
MOTO_PL_EXPECTED_COLS = {"vin no", "vin no.", "vin", "engine no", "engine no.", "item no", "item description", "year", "año", "año modelo", "model year"}

# Columnas esperadas en packing list de repuestos
SP_PL_EXPECTED_COLS = {"part #", "part#", "qty(pcs)", "qty (pcs)", "n.w", "g.w", "complete description"}

# Columnas esperadas en invoice de repuestos
SP_INV_EXPECTED_COLS = {"part #", "part#", "unit price", "amount", "qty(pcs)", "qty (pcs)"}


# ---------------------------------------------------------------------------
# Utilidades de parseo
# ---------------------------------------------------------------------------

def normalize_part_number(pn: str) -> str:
    if not pn:
        return ""
    return str(pn).strip().upper().replace(" ", "")


def parse_excel_date(value) -> tuple[Optional[datetime], Optional[str]]:
    """
    Convierte un valor de celda Excel a (datetime, raw_string).
    - float/int: serial de Excel (días desde 1899-12-30)
    - string "PENDING", "READY", "READY-partial": retorna (None, value)
    - string de fecha: intenta parsear
    - datetime: retorna directo
    """
    if value is None or value == "":
        return None, None

    if isinstance(value, datetime):
        return value, value.strftime("%Y-%m-%d")

    if isinstance(value, (int, float)):
        try:
            # Epoch base Excel: 1899-12-30 (corrección del bug del año 1900)
            base = datetime(1899, 12, 30)
            dt = base + timedelta(days=float(value))
            return dt, dt.strftime("%Y-%m-%d")
        except Exception:
            return None, str(value)

    raw = str(value).strip()
    upper = raw.upper()

    # Valores textuales conocidos
    if upper in ("PENDING", "TBD", "", "-", "N/A"):
        return None, raw
    if upper.startswith("READY"):
        return None, raw

    # Intentar parsear como fecha
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            dt = datetime.strptime(raw, fmt)
            return dt, raw
        except ValueError:
            continue

    return None, raw


def detect_is_spare_part(qty: str, pi_number: str) -> bool:
    if not qty and not pi_number:
        return False
    qty_str = str(qty).strip().upper() if qty else ""
    pi_str = str(pi_number).strip().upper() if pi_number else ""
    return qty_str == "1 LOT" or "-SP" in pi_str


def compute_parent_pi(pi_number: str) -> Optional[str]:
    """'E0000573-SP-1' → 'E0000573', 'E0000573' → None"""
    if not pi_number:
        return None
    upper = str(pi_number).upper()
    idx = upper.find("-SP")
    if idx == -1:
        return None
    return pi_number[:idx]


def compute_status(etr_raw, etl_raw, etd, eta) -> str:
    now = datetime.utcnow()
    if eta and isinstance(eta, datetime) and eta < now:
        return "en_destino"
    if etd and isinstance(etd, datetime) and etd < now:
        return "en_transito"
    if etr_raw and str(etr_raw).upper().startswith("READY"):
        return "listo_fabrica"
    return "en_preparacion"


def _find_header_row(sheet, expected_cols: set, max_rows: int = 35) -> Optional[int]:
    """
    Escanea las primeras max_rows filas buscando la que tenga al menos 3
    coincidencias con expected_cols (case-insensitive).
    Retorna el índice de fila (base-1) o None.
    """
    for row_idx in range(1, max_rows + 1):
        row_values = {
            str(sheet.cell(row=row_idx, column=c).value or "").strip().lower()
            for c in range(1, sheet.max_column + 1)
        }
        matches = sum(1 for col in expected_cols if col.lower() in row_values)
        if matches >= 3:
            return row_idx
    return None


def _build_col_map(sheet, header_row: int) -> dict[str, int]:
    """Mapea nombre_de_columna_normalizado → índice (base-1)."""
    col_map = {}
    for col_idx in range(1, sheet.max_column + 1):
        val = sheet.cell(row=header_row, column=col_idx).value
        if val:
            col_map[str(val).strip().lower()] = col_idx
    return col_map


def _cell(sheet, row: int, col_map: dict, *keys) -> any:
    """Obtiene el valor de la primera clave encontrada en col_map."""
    for key in keys:
        if key.lower() in col_map:
            return sheet.cell(row=row, column=col_map[key.lower()]).value
    return None


# ---------------------------------------------------------------------------
# Importación del Excel de Shipment Status
# ---------------------------------------------------------------------------

async def import_shipment_excel(
    db: AsyncSession,
    file_bytes: bytes,
    actor: CurrentUser,
) -> ImportExcelResult:
    result = ImportExcelResult()
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = wb.active

    header_row = _find_header_row(sheet, SHIPMENT_EXPECTED_COLS)
    if header_row is None:
        result.errors.append({"code": "HEADERS_NOT_FOUND", "reason": "No se encontraron cabeceras en las primeras 35 filas"})
        return result

    col_map = _build_col_map(sheet, header_row)

    for row_idx in range(header_row + 1, sheet.max_row + 1):
        try:
            pi_number = _cell(sheet, row_idx, col_map, "pi_number", "pi number", "pi no", "pi no.")
            model = _cell(sheet, row_idx, col_map, "model", "model name")

            if not pi_number or not model:
                continue  # Fila vacía o de totales

            pi_number = str(pi_number).strip()
            model = str(model).strip()

            # Extraer todos los campos
            cycle_raw = _cell(sheet, row_idx, col_map, "cycle")
            cycle = int(cycle_raw) if cycle_raw and str(cycle_raw).isdigit() else None

            invoice_number = _cell(sheet, row_idx, col_map, "invoice_number", "invoice number", "invoice no")
            invoice_number = str(invoice_number).strip() if invoice_number else None

            model_year_raw = _cell(sheet, row_idx, col_map, "model_year", "model year", "year")
            model_year = int(model_year_raw) if model_year_raw and str(model_year_raw).isdigit() else None

            qty_raw = _cell(sheet, row_idx, col_map, "qty", "quantity")
            qty = str(qty_raw).strip() if qty_raw is not None else None
            qty_numeric = None
            if qty and qty.replace(",", "").isdigit():
                qty_numeric = int(qty.replace(",", ""))

            total_raw = _cell(sheet, row_idx, col_map, "total")
            total_units = int(total_raw) if total_raw and str(total_raw).isdigit() else None

            containers_raw = _cell(sheet, row_idx, col_map, "containers", "container", "40hq")
            containers = int(containers_raw) if containers_raw and str(containers_raw).isdigit() else None

            etr, etr_raw = parse_excel_date(_cell(sheet, row_idx, col_map, "etr"))
            etl, etl_raw = parse_excel_date(_cell(sheet, row_idx, col_map, "etl"))
            etd, etd_raw = parse_excel_date(_cell(sheet, row_idx, col_map, "etd"))
            eta, eta_raw = parse_excel_date(_cell(sheet, row_idx, col_map, "eta"))

            departure_port = _cell(sheet, row_idx, col_map, "departure_port", "departure port", "port of loading")
            departure_port = str(departure_port).strip() if departure_port else None

            bl_container = _cell(sheet, row_idx, col_map, "bl_container", "bl container", "bl/container", "b/l")
            bl_container = str(bl_container).strip() if bl_container else None

            vessel = _cell(sheet, row_idx, col_map, "vessel", "vessel name", "ship")
            vessel = str(vessel).strip() if vessel else None

            digital_docs = _cell(sheet, row_idx, col_map, "digital_docs_status", "digital docs", "digital docs status")
            digital_docs = str(digital_docs).strip().upper() if digital_docs else "PENDING"

            original_docs = _cell(sheet, row_idx, col_map, "original_docs_status", "original docs", "original docs status")
            original_docs = str(original_docs).strip().upper() if original_docs else "PENDING"

            remarks = _cell(sheet, row_idx, col_map, "remarks", "notes", "observations")
            remarks = str(remarks).strip() if remarks else None

            is_spare = detect_is_spare_part(qty, pi_number)
            parent_pi = compute_parent_pi(pi_number)
            status = compute_status(etr_raw, etl_raw, etd, eta)

            # Upsert por (pi_number, model)
            stmt = select(ShipmentOrder).where(
                ShipmentOrder.pi_number == pi_number,
                ShipmentOrder.model == model,
            )
            existing = (await db.execute(stmt)).scalar_one_or_none()

            if existing:
                # Detectar cambios para el audit log
                changes = {}
                fields = {
                    "cycle": cycle, "invoice_number": invoice_number,
                    "model_year": model_year, "qty": qty, "qty_numeric": qty_numeric,
                    "total_units": total_units, "containers": containers,
                    "etr": etr, "etr_raw": etr_raw,
                    "etl": etl, "etl_raw": etl_raw,
                    "etd": etd, "etd_raw": etd_raw,
                    "eta": eta, "eta_raw": eta_raw,
                    "departure_port": departure_port, "bl_container": bl_container,
                    "vessel": vessel, "digital_docs_status": digital_docs,
                    "original_docs_status": original_docs, "remarks": remarks,
                    "computed_status": status,
                }
                for field, new_val in fields.items():
                    old_val = getattr(existing, field)
                    if str(old_val) != str(new_val):
                        changes[field] = {"old": str(old_val), "new": str(new_val)}
                        setattr(existing, field, new_val)

                if changes:
                    existing.updated_at = datetime.utcnow()
                    _log_audit(db, actor, "UPDATE", "ShipmentOrder", str(existing.id), changes, existing.id)
                    result.updated += 1
                else:
                    result.skipped += 1

                # Garantizar que exista el lote SP aunque la orden sea un update
                if is_spare:
                    existing_lot = (await db.execute(
                        select(SparePartLot).where(SparePartLot.lot_identifier == pi_number)
                    )).scalar_one_or_none()
                    if not existing_lot:
                        lot = SparePartLot(
                            shipment_order_id=existing.id,
                            lot_identifier=pi_number,
                            created_at=datetime.utcnow(),
                        )
                        db.add(lot)
                        await db.flush()

            else:
                new_order = ShipmentOrder(
                    cycle=cycle,
                    pi_number=pi_number,
                    invoice_number=invoice_number,
                    model=model,
                    model_year=model_year,
                    qty=qty,
                    qty_numeric=qty_numeric,
                    total_units=total_units,
                    containers=containers,
                    etr=etr, etr_raw=etr_raw,
                    etl=etl, etl_raw=etl_raw,
                    etd=etd, etd_raw=etd_raw,
                    eta=eta, eta_raw=eta_raw,
                    departure_port=departure_port,
                    bl_container=bl_container,
                    vessel=vessel,
                    digital_docs_status=digital_docs,
                    original_docs_status=original_docs,
                    remarks=remarks,
                    is_spare_part=is_spare,
                    parent_pi_number=parent_pi,
                    computed_status=status,
                    created_at=datetime.utcnow(),
                )
                db.add(new_order)
                await db.flush()  # Obtener ID antes del commit

                # Auto-crear SparePartLot para registros SP
                if is_spare:
                    existing_lot = (await db.execute(
                        select(SparePartLot).where(SparePartLot.lot_identifier == pi_number)
                    )).scalar_one_or_none()

                    if not existing_lot:
                        lot = SparePartLot(
                            shipment_order_id=new_order.id,
                            lot_identifier=pi_number,
                            created_at=datetime.utcnow(),
                        )
                        db.add(lot)

                _log_audit(db, actor, "INSERT", "ShipmentOrder", str(new_order.id),
                           {"pi_number": pi_number, "model": model}, new_order.id)
                result.inserted += 1

        except Exception as e:
            logger.error(f"Error procesando fila {row_idx}: {e}", exc_info=True)
            result.errors.append({"row": row_idx, "reason": str(e)})

    await db.flush()
    return result


# ---------------------------------------------------------------------------
# Importación del archivo de Packing List de motos (VINs por pedido)
# ---------------------------------------------------------------------------

async def import_shipping_doc_excel(
    db: AsyncSession,
    file_bytes: bytes,
    actor: CurrentUser,
) -> dict:
    """
    Procesa el archivo multi-hoja de Packing List de motos.
    Solo procesa hojas de tipo moto_packing_list → ShipmentMotoUnit.
    Cada hoja debe contener el PI number en sus filas de encabezado
    (ej. "INV-E0000574") para vincularse al pedido correspondiente.
    Las hojas de repuestos (sp_packing_list, sp_invoice) se ignoran.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    summary = {"moto_units_added": 0, "moto_units_updated": 0, "sheets_processed": 0, "sheets_skipped": 0, "errors": [], "duplicates_skipped": []}

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet_type = _detect_sheet_type(sheet)
        logger.info(f"Hoja '{sheet_name}' detectada como: {sheet_type}")

        if sheet_type == "moto_packing_list":
            try:
                added, upd, dups = await _process_moto_packing_list(db, sheet, actor)
                summary["moto_units_added"] += added
                summary["moto_units_updated"] += upd
                summary["sheets_processed"] += 1
                summary["duplicates_skipped"].extend(dups)
            except ValueError as e:
                summary["errors"].append({"sheet": sheet_name, "reason": str(e)})
        else:
            summary["sheets_skipped"] += 1

    return summary


def _detect_sheet_type(sheet) -> str:
    """Detecta el tipo de hoja por los nombres de columnas que contiene."""
    all_values = set()
    for row in sheet.iter_rows(min_row=1, max_row=35, values_only=True):
        for cell in row:
            if cell:
                all_values.add(str(cell).strip().lower())

    has_vin = any(v in all_values for v in ("vin no.", "vin no", "vin"))
    has_engine = any(v in all_values for v in ("engine no.", "engine no"))
    has_part = any(v in all_values for v in ("part #", "part#"))
    has_weight = any(v in all_values for v in ("n.w (kgs)", "n.w(kgs)", "n.w"))
    has_price = any(v in all_values for v in ("unit price", "amount"))
    has_meas = any(v in all_values for v in ("meas (cbm)", "cbm", "meas"))

    if has_vin and has_engine:
        return "moto_packing_list"
    if has_part and has_weight and has_meas and not has_price:
        return "sp_packing_list"
    if has_part and has_price and not has_weight:
        return "sp_invoice"
    if has_meas and not has_part and not has_vin:
        return "summary"
    return "unknown"


async def _process_moto_packing_list(db: AsyncSession, sheet, actor: CurrentUser) -> int:
    header_row = _find_header_row(sheet, MOTO_PL_EXPECTED_COLS)
    if not header_row:
        return 0

    col_map = _build_col_map(sheet, header_row)

    # Cargar todos los PI numbers de pedidos de motos registrados en la DB
    all_pi_numbers = [
        row[0] for row in (await db.execute(
            select(ShipmentOrder.pi_number).where(ShipmentOrder.is_spare_part == False)
        )).all()
    ]
    # Ordenar de mayor a menor longitud para evitar matches parciales
    all_pi_numbers.sort(key=len, reverse=True)

    # Buscar en las filas de encabezado cualquier PI number registrado
    source_pi = None
    for r in range(1, header_row):
        for c in range(1, sheet.max_column + 1):
            val = str(sheet.cell(row=r, column=c).value or "").strip().upper()
            if not val:
                continue
            for pi in all_pi_numbers:
                if pi.upper() in val:
                    source_pi = pi
                    break
        if source_pi:
            break

    # Buscar ShipmentOrder para este PI
    if not source_pi:
        raise ValueError(
            "No se encontró ningún PI number registrado en las filas de encabezado de la hoja. "
            "El archivo debe contener el PI number antes del encabezado de columnas."
        )

    stmt = select(ShipmentOrder).where(
        ShipmentOrder.pi_number == source_pi,
        ShipmentOrder.is_spare_part == False,
    ).limit(1)
    order = (await db.execute(stmt)).scalar_one_or_none()

    if not order:
        raise ValueError(
            f"No existe ningún pedido con PI number '{source_pi}' en el sistema. "
            f"Cargá primero el Shipment Status para registrar el pedido antes de subir el Packing List."
        )

    # Mapa VIN → unidad existente en la DB (solo de este pedido para upsert)
    existing_units_map: dict = {
        u.vin_number: u
        for u in (await db.execute(
            select(ShipmentMotoUnit).where(
                ShipmentMotoUnit.vin_number.isnot(None),
                ShipmentMotoUnit.shipment_order_id == order.id,
            )
        )).scalars().all()
    }
    # VINs de OTROS pedidos — no se tocan, solo se avisa si hay colisión
    other_order_vins = set(
        row[0] for row in (await db.execute(
            select(ShipmentMotoUnit.vin_number).where(
                ShipmentMotoUnit.vin_number.isnot(None),
                ShipmentMotoUnit.shipment_order_id != order.id,
            )
        )).all()
    )

    added = 0
    updated = 0
    duplicates = []
    seen_in_file = set()

    for row_idx in range(header_row + 1, sheet.max_row + 1):
        vin = _cell(sheet, row_idx, col_map, "vin no.", "vin no", "vin")
        if not vin:
            continue

        vin_clean = str(vin).strip()

        if vin_clean in seen_in_file:
            duplicates.append({"vin": vin_clean, "row": row_idx, "reason": "duplicado en el archivo"})
            continue

        if vin_clean in other_order_vins:
            duplicates.append({"vin": vin_clean, "row": row_idx, "reason": "ya existe en otro pedido"})
            continue

        seen_in_file.add(vin_clean)

        raw_year = _cell(sheet, row_idx, col_map, "year", "año", "año modelo", "model year")
        parsed_year = None
        if raw_year is not None:
            try:
                parsed_year = int(float(str(raw_year).strip()))
            except (ValueError, TypeError):
                parsed_year = None

        engine = str(_cell(sheet, row_idx, col_map, "engine no.", "engine no") or "").strip() or None
        color = str(_cell(sheet, row_idx, col_map, "color") or "").strip() or None
        item_no = _cell(sheet, row_idx, col_map, "item no", "item no.")

        if vin_clean in existing_units_map:
            # Upsert: actualizar campos del archivo, sin tocar datos de aduana/DIM
            unit = existing_units_map[vin_clean]
            unit.item_no = item_no
            unit.engine_number = engine
            unit.color = color
            if parsed_year is not None:
                unit.model_year = parsed_year
            updated += 1
        else:
            unit = ShipmentMotoUnit(
                shipment_order_id=order.id,
                item_no=item_no,
                vin_number=vin_clean,
                engine_number=engine,
                color=color,
                model_year=parsed_year,
                source_pi=source_pi,
                created_at=datetime.utcnow(),
            )
            db.add(unit)
            added += 1

    await db.flush()
    return added, updated, duplicates


async def _process_sp_packing_list(db: AsyncSession, sheet, actor: CurrentUser) -> int:
    """Crea o actualiza spare_part_items con datos físicos: qty, peso, volumen, cartons.
    Detecta el PI number en las filas de encabezado para vincular al lote correcto.
    Si el ítem no existe en ese lote, lo crea (upsert por part_number + lot_id).
    """
    from app.models.imports import SparePartItem
    import re as _re

    header_row = _find_header_row(sheet, SP_PL_EXPECTED_COLS)
    if not header_row:
        return 0

    col_map = _build_col_map(sheet, header_row)

    # Detectar PI number en las filas previas al encabezado
    source_pi = None
    for r in range(1, header_row):
        for c in range(1, sheet.max_column + 1):
            val = str(sheet.cell(row=r, column=c).value or "")
            match = _re.search(r'E\d{7}', val, _re.IGNORECASE)
            if match:
                source_pi = match.group(0).upper()
                break
        if source_pi:
            break

    # Buscar el SparePartLot vinculado a ese PI
    target_lot = None
    if source_pi:
        stmt = select(SparePartLot).where(SparePartLot.lot_identifier == source_pi)
        target_lot = (await db.execute(stmt)).scalar_one_or_none()

    count = 0
    current_carton_data = {}

    for row_idx in range(header_row + 1, sheet.max_row + 1):
        part_raw = _cell(sheet, row_idx, col_map, "part #", "part#")
        if not part_raw:
            continue

        part_number = normalize_part_number(str(part_raw))
        if not part_number:
            continue

        # La primera fila de cada cargón tiene N.W, G.W, CBM
        carton_no = _cell(sheet, row_idx, col_map, "carton no", "carton no.")
        if carton_no:
            current_carton_data = {
                "net_weight_kg": _cell(sheet, row_idx, col_map, "n.w (kgs)", "n.w(kgs)", "n.w"),
                "gross_weight_kg": _cell(sheet, row_idx, col_map, "g.w (kgs)", "g.w(kgs)", "g.w"),
                "cbm": _cell(sheet, row_idx, col_map, "cbm"),
            }

        qty_pcs_raw = _cell(sheet, row_idx, col_map, "qty(pcs)", "qty (pcs)", "qty")
        try:
            qty_pcs = int(float(str(qty_pcs_raw))) if qty_pcs_raw is not None else None
        except (ValueError, TypeError):
            qty_pcs = None

        qty_ctns_raw = _cell(sheet, row_idx, col_map, "qty(ctns)", "qty (ctns)", "qty ctns")
        try:
            qty_ctns = int(float(str(qty_ctns_raw))) if qty_ctns_raw is not None else None
        except (ValueError, TypeError):
            qty_ctns = None

        desc_en = _cell(sheet, row_idx, col_map, "complete description", "description")
        desc_es_raw = _cell(sheet, row_idx, col_map, "spanish description", "spanish desc", "descripción")
        try:
            desc_es = str(desc_es_raw).encode("latin-1").decode("utf-8") if desc_es_raw else None
        except Exception:
            desc_es = str(desc_es_raw) if desc_es_raw else None

        model_applicable = _cell(sheet, row_idx, col_map, "model")

        def _apply_fields(item: SparePartItem):
            if qty_pcs is not None:
                item.qty_ordered = qty_pcs
                item.qty_pending = item.qty_ordered - item.qty_received
            if qty_ctns is not None:
                item.qty_cartons = qty_ctns
            if desc_en:
                item.description = str(desc_en).strip()
            if desc_es:
                item.description_es = desc_es
            if model_applicable:
                item.model_applicable = str(model_applicable).strip()
            if current_carton_data.get("net_weight_kg") is not None:
                item.net_weight_kg = current_carton_data["net_weight_kg"]
            if current_carton_data.get("gross_weight_kg") is not None:
                item.gross_weight_kg = current_carton_data["gross_weight_kg"]
            if current_carton_data.get("cbm") is not None:
                item.cbm = current_carton_data["cbm"]
            item.updated_at = datetime.utcnow()

        if target_lot:
            # Upsert: buscar en el lote específico
            stmt = select(SparePartItem).where(
                SparePartItem.lot_id == target_lot.id,
                SparePartItem.part_number == part_number,
            )
            item = (await db.execute(stmt)).scalar_one_or_none()
            if item is None:
                item = SparePartItem(
                    lot_id=target_lot.id,
                    part_number=part_number,
                    qty_ordered=qty_pcs or 0,
                    qty_received=0,
                    qty_pending=qty_pcs or 0,
                    status="PENDING",
                    created_at=datetime.utcnow(),
                )
                db.add(item)
            _apply_fields(item)
            count += 1
        else:
            # Sin PI detectado: actualizar cualquier item existente con ese part_number
            stmt = select(SparePartItem).where(SparePartItem.part_number == part_number)
            items = (await db.execute(stmt)).scalars().all()
            for item in items:
                _apply_fields(item)
                count += 1

    await db.flush()
    return count


async def _process_sp_invoice(db: AsyncSession, sheet, actor: CurrentUser) -> int:
    """Actualiza spare_part_items con precios: unit_price, amount."""
    from app.models.imports import SparePartItem

    header_row = _find_header_row(sheet, SP_INV_EXPECTED_COLS)
    if not header_row:
        return 0

    col_map = _build_col_map(sheet, header_row)
    count = 0
    parts_with_price: set[str] = set()

    for row_idx in range(header_row + 1, sheet.max_row + 1):
        part_raw = _cell(sheet, row_idx, col_map, "part #", "part#")
        if not part_raw:
            continue

        part_number = normalize_part_number(str(part_raw))
        if not part_number:
            continue

        unit_price_raw = _cell(sheet, row_idx, col_map, "unit price")
        amount_raw = _cell(sheet, row_idx, col_map, "amount")

        try:
            unit_price = float(str(unit_price_raw).replace("$", "").replace(",", "").strip()) if unit_price_raw else None
        except ValueError:
            unit_price = None

        try:
            amount = float(str(amount_raw).replace("$", "").replace(",", "").strip()) if amount_raw else None
        except ValueError:
            amount = None

        stmt = select(SparePartItem).where(SparePartItem.part_number == part_number)
        items = (await db.execute(stmt)).scalars().all()

        price_updated = False
        for item in items:
            if unit_price is not None:
                item.unit_price = unit_price
                price_updated = True
            if amount is not None:
                item.amount = amount
            item.updated_at = datetime.utcnow()
            count += 1

        if price_updated:
            parts_with_price.add(part_number)

    await db.flush()

    if parts_with_price:
        from app.services.pricing_service import recalculate_part_cost
        for pn in parts_with_price:
            await recalculate_part_cost(db, pn)

    return count


# ---------------------------------------------------------------------------
# Reconciliación: subida de packing list para un lote SP
# ---------------------------------------------------------------------------
# Carga de detalle de orden de compra (PI Detail) → crea SparePartItems
# ---------------------------------------------------------------------------
# Crear pedido de repuestos desde Excel (flujo inicial)
# ---------------------------------------------------------------------------

SP_ORDER_COLS = {"codigo parte", "nombre", "cantidad", "moto aplica", "codigo", "part #", "referencia", "qty ordered", "qty"}

async def create_sp_order_from_excel(
    db: AsyncSession,
    reference: str,
    file_bytes: bytes,
    actor: CurrentUser,
) -> dict:
    """
    Crea un ShipmentOrder SP + SparePartLot + SparePartItems desde el Excel
    de solicitud de repuestos del usuario.
    Columnas esperadas: Codigo Parte | Nombre | Cantidad | Moto Aplica
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = None
    header_row = None
    for name in wb.sheetnames:
        s = wb[name]
        hr = _find_header_row(s, SP_ORDER_COLS)
        if hr:
            sheet = s
            header_row = hr
            break

    if sheet is None or header_row is None:
        return {"inserted": 0, "errors": [{"row": 0, "reason": "No se encontraron columnas requeridas (Codigo Parte / Cantidad)"}]}

    col_map = _build_col_map(sheet, header_row)

    ref = reference.strip().upper()

    # Verificar si ya existe un pedido con esta referencia
    existing_order = (await db.execute(
        select(ShipmentOrder).where(ShipmentOrder.pi_number == ref)
    )).scalar_one_or_none()

    if existing_order:
        order = existing_order
        lot = (await db.execute(
            select(SparePartLot).where(SparePartLot.shipment_order_id == order.id)
        )).scalar_one_or_none()
        if not lot:
            lot = SparePartLot(
                shipment_order_id=order.id,
                lot_identifier=ref,
                detail_loaded=False,
                packing_list_received=False,
                created_at=datetime.utcnow(),
            )
            db.add(lot)
            await db.flush()
    else:
        order = ShipmentOrder(
            pi_number=ref,
            model="REPUESTOS",
            is_spare_part=True,
            digital_docs_status="PENDING",
            original_docs_status="PENDING",
            computed_status="en_preparacion",
            created_at=datetime.utcnow(),
        )
        db.add(order)
        await db.flush()
        lot = SparePartLot(
            shipment_order_id=order.id,
            lot_identifier=ref,
            detail_loaded=False,
            packing_list_received=False,
            created_at=datetime.utcnow(),
        )
        db.add(lot)
        await db.flush()

    # Borrar ítems previos del lote (re-carga limpia)
    old_items = (await db.execute(
        select(SparePartItem).where(SparePartItem.lot_id == lot.id)
    )).scalars().all()
    for old in old_items:
        await db.delete(old)
    await db.flush()

    existing_items: dict[str, SparePartItem] = {}

    inserted = updated = skipped = 0
    errors = []

    # --- Primer pase: agregar filas por part_number (suma cantidades de duplicados) ---
    aggregated: dict[str, dict] = {}  # part_number → {qty, nombre, modelo_moto}
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        try:
            part_raw = _cell(sheet, row_idx, col_map,
                             "codigo parte", "codigo", "part #", "part#", "referencia", "codigo de parte")
            if not part_raw:
                continue
            part_number = normalize_part_number(str(part_raw))
            if not part_number:
                continue

            qty_raw = _cell(sheet, row_idx, col_map, "cantidad", "qty", "qty ordered", "cantidad solicitada")
            try:
                qty = int(float(str(qty_raw))) if qty_raw is not None else None
            except (ValueError, TypeError):
                qty = None
            if not qty or qty <= 0:
                skipped += 1
                continue

            nombre = _cell(sheet, row_idx, col_map, "nombre", "descripcion", "descripción", "nombre es", "description")
            modelo_moto = _cell(sheet, row_idx, col_map, "moto aplica", "moto", "modelo moto", "modelo", "aplica")

            if part_number in aggregated:
                aggregated[part_number]["qty"] += qty  # sumar duplicados
            else:
                aggregated[part_number] = {
                    "qty": qty,
                    "nombre": str(nombre).strip() if nombre else None,
                    "modelo_moto": str(modelo_moto).strip() if modelo_moto else None,
                }
        except Exception as e:
            errors.append({"row": row_idx, "reason": str(e)})

    # --- Segundo pase: crear o actualizar SparePartItems con cantidades agregadas ---
    for part_number, data in aggregated.items():
        qty = data["qty"]
        nombre = data["nombre"]
        modelo_moto = data["modelo_moto"]

        item = SparePartItem(
            lot_id=lot.id,
            part_number=part_number,
            description_es=nombre,
            model_applicable=modelo_moto,
            qty_ordered=qty,
            qty_received=0,
            qty_pending=qty,
            status="PENDING",
            created_at=datetime.utcnow(),
        )
        db.add(item)
        inserted += 1

    if inserted > 0:
        lot.detail_loaded = True
        lot.updated_at = datetime.utcnow() if hasattr(lot, 'updated_at') else None

    await db.flush()
    return {
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
        "order_id": str(order.id),
        "lot_id": str(lot.id),
        "reference": ref,
    }


# ---------------------------------------------------------------------------
# Crear pedido de motos desde Excel (flujo inicial)
# ---------------------------------------------------------------------------

MOTO_ORDER_COLS = {"referencia", "descripcion", "cantidad", "modelo", "model", "moto", "qty"}

async def create_moto_order_from_excel(
    db: AsyncSession,
    cycle: int,
    file_bytes: bytes,
    actor: CurrentUser,
) -> dict:
    """
    Crea uno o más ShipmentOrders de motos desde el Excel de solicitud.
    Columnas esperadas: Referencia Moto | Descripcion | Cantidad
    Cada fila = un ShipmentOrder.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = None
    header_row = None
    for name in wb.sheetnames:
        s = wb[name]
        hr = _find_header_row(s, MOTO_ORDER_COLS)
        if hr:
            sheet = s
            header_row = hr
            break

    if sheet is None or header_row is None:
        return {"inserted": 0, "errors": [{"row": 0, "reason": "No se encontraron columnas requeridas (Referencia / Cantidad)"}]}

    col_map = _build_col_map(sheet, header_row)
    inserted = skipped = 0
    errors = []

    for row_idx in range(header_row + 1, sheet.max_row + 1):
        try:
            ref_raw = _cell(sheet, row_idx, col_map, "referencia", "modelo", "model", "moto", "referencia moto")
            if not ref_raw:
                continue
            model = str(ref_raw).strip().upper()
            if not model:
                continue

            qty_raw = _cell(sheet, row_idx, col_map, "cantidad", "qty", "cantidad solicitada")
            try:
                qty = int(float(str(qty_raw))) if qty_raw is not None else None
            except (ValueError, TypeError):
                qty = None
            if not qty or qty <= 0:
                skipped += 1
                continue

            desc = _cell(sheet, row_idx, col_map, "descripcion", "descripción", "description", "nombre")

            # Usar referencia temporal hasta que el proveedor asigne PI
            temp_pi = f"REQ-{cycle}-{model}" if cycle else f"REQ-{model}"

            existing = (await db.execute(
                select(ShipmentOrder).where(
                    ShipmentOrder.pi_number == temp_pi,
                    ShipmentOrder.model == model,
                )
            )).scalar_one_or_none()

            if existing:
                existing.qty = str(qty)
                existing.qty_numeric = qty
                existing.updated_at = datetime.utcnow()
            else:
                order = ShipmentOrder(
                    cycle=cycle if cycle else None,
                    pi_number=temp_pi,
                    model=model,
                    qty=str(qty),
                    qty_numeric=qty,
                    is_spare_part=False,
                    digital_docs_status="PENDING",
                    original_docs_status="PENDING",
                    computed_status="en_preparacion",
                    remarks=str(desc).strip() if desc else None,
                    created_at=datetime.utcnow(),
                )
                db.add(order)
                inserted += 1

        except Exception as e:
            errors.append({"row": row_idx, "reason": str(e)})

    await db.flush()
    return {"inserted": inserted, "skipped": skipped, "errors": errors}


# ---------------------------------------------------------------------------

# Columnas mínimas para detectar hoja de orden de compra
ORDER_DETAIL_EXPECTED_COLS = {"part #", "part#", "qty ordered", "qty", "description"}


async def load_order_detail_excel(
    db: AsyncSession,
    lot: SparePartLot,
    file_bytes: bytes,
    actor: CurrentUser,
) -> dict:
    """
    Carga el detalle de la orden de compra para un lote de repuestos.
    Crea SparePartItems con qty_ordered. Si el ítem ya existe, actualiza qty_ordered.
    Retorna { inserted, updated, skipped, errors }.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    # Buscar la primera hoja que tenga las columnas esperadas
    target_sheet = None
    for name in wb.sheetnames:
        s = wb[name]
        header_row = _find_header_row(s, ORDER_DETAIL_EXPECTED_COLS)
        if header_row:
            target_sheet = s
            break

    if target_sheet is None:
        return {"inserted": 0, "updated": 0, "skipped": 0,
                "errors": [{"row": 0, "reason": "No se encontraron las columnas requeridas (Part #, Qty Ordered, Description)"}]}

    header_row = _find_header_row(target_sheet, ORDER_DETAIL_EXPECTED_COLS)
    col_map = _build_col_map(target_sheet, header_row)

    # Índice de ítems existentes en este lote
    existing_stmt = select(SparePartItem).where(SparePartItem.lot_id == lot.id)
    existing_items: dict[str, SparePartItem] = {
        i.part_number: i
        for i in (await db.execute(existing_stmt)).scalars().all()
    }

    inserted = updated = skipped = 0
    errors = []
    parts_with_price: set[str] = set()  # part_numbers que tuvieron unit_price en esta carga

    for row_idx in range(header_row + 1, target_sheet.max_row + 1):
        try:
            part_raw = _cell(target_sheet, row_idx, col_map, "part #", "part#")
            if not part_raw:
                continue

            part_number = normalize_part_number(str(part_raw))
            if not part_number:
                continue

            qty_raw = _cell(target_sheet, row_idx, col_map, "qty ordered", "qty_ordered", "qty(pcs)", "qty (pcs)", "qty")
            try:
                qty_ordered = int(float(str(qty_raw))) if qty_raw is not None else None
            except (ValueError, TypeError):
                qty_ordered = None

            if qty_ordered is None or qty_ordered <= 0:
                skipped += 1
                continue

            desc_en = _cell(target_sheet, row_idx, col_map, "description", "complete description", "descripción en")
            desc_es = _cell(target_sheet, row_idx, col_map, "descripcion", "descripción", "spanish description", "descripcion es")
            model_raw = _cell(target_sheet, row_idx, col_map, "model", "modelo")
            unit_price_raw = _cell(target_sheet, row_idx, col_map, "unit price", "precio unitario", "precio")
            try:
                unit_price = float(str(unit_price_raw).replace("$", "").replace(",", "").strip()) if unit_price_raw else None
            except (ValueError, TypeError):
                unit_price = None

            if part_number in existing_items:
                item = existing_items[part_number]
                old_qty = item.qty_ordered
                item.qty_ordered = qty_ordered
                item.qty_pending = max(0, qty_ordered - item.qty_received)
                if desc_en:
                    item.description = str(desc_en).strip()
                if desc_es:
                    item.description_es = str(desc_es).strip()
                if model_raw:
                    item.model_applicable = str(model_raw).strip()
                if unit_price is not None:
                    item.unit_price = unit_price
                    item.amount = round(unit_price * qty_ordered, 4)
                    parts_with_price.add(part_number)
                item.updated_at = datetime.utcnow()
                updated += 1
            else:
                amount = round(unit_price * qty_ordered, 4) if unit_price else None
                item = SparePartItem(
                    lot_id=lot.id,
                    part_number=part_number,
                    description=str(desc_en).strip() if desc_en else None,
                    description_es=str(desc_es).strip() if desc_es else None,
                    model_applicable=str(model_raw).strip() if model_raw else None,
                    qty_ordered=qty_ordered,
                    qty_received=0,
                    qty_pending=qty_ordered,
                    unit_price=unit_price,
                    amount=amount,
                    status="PENDING",
                    created_at=datetime.utcnow(),
                )
                db.add(item)
                existing_items[part_number] = item
                if unit_price is not None:
                    parts_with_price.add(part_number)
                inserted += 1

        except Exception as e:
            errors.append({"row": row_idx, "reason": str(e)})

    # Marcar lote como detalle cargado
    if inserted + updated > 0:
        lot.detail_loaded = True
        lot.updated_at = datetime.utcnow()

    await db.flush()

    # Recalcular costo promedio FOB para cada parte con precio en este lote
    if parts_with_price:
        from app.services.pricing_service import recalculate_part_cost
        for pn in parts_with_price:
            await recalculate_part_cost(db, pn, lot_identifier=lot.lot_identifier)

    return {"inserted": inserted, "updated": updated, "skipped": skipped, "errors": errors}


# ---------------------------------------------------------------------------

async def reconcile_lot_packing_list(
    db: AsyncSession,
    lot: SparePartLot,
    file_bytes: bytes,
    file_name: str,
    minio_object_name: str,
    actor: CurrentUser,
) -> dict:
    """
    Acepta tanto el formato Packing List como el formato Invoice del proveedor.
    Cruza contra SparePartItems del lote y genera ReconciliationResult + Backorders.
    Si el archivo es una invoice, además actualiza unit_price y amount en los ítems.
    """
    import re as _re
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    # Buscar hoja que coincida con PL o Invoice; priorizar la que tenga el PI del lote
    target_sheet = None
    for name in wb.sheetnames:
        s = wb[name]
        t = _detect_sheet_type(s)
        if t in ("sp_packing_list", "sp_invoice"):
            # Preferir la hoja que mencione el PI del lote en su nombre o contenido
            if lot.lot_identifier.replace("-SP", "") in name.upper() or lot.lot_identifier in name.upper():
                target_sheet = s
                break
            if target_sheet is None:
                target_sheet = s

    if target_sheet is None:
        target_sheet = wb.active

    # Intentar cabeceras de invoice primero (más columnas útiles), luego packing list
    header_row = _find_header_row(target_sheet, SP_INV_EXPECTED_COLS)
    is_invoice = header_row is not None
    if not is_invoice:
        header_row = _find_header_row(target_sheet, SP_PL_EXPECTED_COLS)
    if header_row is None:
        return {"error": "No se encontraron cabeceras válidas. El archivo debe tener columnas: Part #, Qty(PCS), y Unit Price o N.W"}

    col_map = _build_col_map(target_sheet, header_row)

    # Eliminar resultados anteriores para este lote (re-conciliación)
    old_results = (await db.execute(
        select(ReconciliationResult).where(ReconciliationResult.lot_id == lot.id)
    )).scalars().all()
    for r in old_results:
        await db.delete(r)

    old_pls = (await db.execute(
        select(PackingList).where(PackingList.lot_id == lot.id)
    )).scalars().all()
    for pl in old_pls:
        await db.delete(pl)

    await db.flush()

    # Crear registro PackingList
    pl_record = PackingList(
        lot_id=lot.id,
        uploaded_by=uuid.UUID(actor.user_id) if actor.user_id else None,
        file_name=file_name,
        minio_object_name=minio_object_name,
        uploaded_at=datetime.utcnow(),
        processed=False,
    )
    db.add(pl_record)
    await db.flush()

    # Parsear filas → {part_number: qty_in_pl}
    pl_items: dict[str, int] = {}
    pl_prices: dict[str, tuple] = {}   # part_number → (unit_price, amount, desc_en, desc_es, model)
    pl_item_records: list[PackingListItem] = []

    for row_idx in range(header_row + 1, target_sheet.max_row + 1):
        part_raw = _cell(target_sheet, row_idx, col_map, "part #", "part#")
        if not part_raw:
            continue
        part_number = normalize_part_number(str(part_raw))
        if not part_number:
            continue

        qty_raw = _cell(target_sheet, row_idx, col_map, "qty(pcs)", "qty (pcs)", "qty")
        try:
            qty = int(float(str(qty_raw))) if qty_raw is not None else 0
        except (ValueError, TypeError):
            qty = 0

        desc_raw = _cell(target_sheet, row_idx, col_map, "complete description", "description")
        desc = str(desc_raw).strip() if desc_raw else None

        pl_items[part_number] = pl_items.get(part_number, 0) + qty

        # Si es invoice, extraer datos de precio y descripción adicional
        if is_invoice:
            desc_es_raw = _cell(target_sheet, row_idx, col_map, "spanish description", "descripcion")
            desc_es = str(desc_es_raw).strip() if desc_es_raw else None

            model_raw = _cell(target_sheet, row_idx, col_map, "model", "modelo")
            model_val = str(model_raw).strip() if model_raw else None

            price_raw = _cell(target_sheet, row_idx, col_map, "unit price", "unit_price", "price")
            try:
                unit_price = float(price_raw) if price_raw is not None else None
            except (ValueError, TypeError):
                unit_price = None

            amount_raw = _cell(target_sheet, row_idx, col_map, "amount", "total")
            try:
                amount = float(amount_raw) if amount_raw is not None else None
            except (ValueError, TypeError):
                amount = None

            if part_number in pl_prices:
                # Sumar amount para duplicados; mantener unit_price y textos del primer registro
                prev = pl_prices[part_number]
                new_amount = (prev[1] or 0) + (amount or 0) if (prev[1] is not None or amount is not None) else None
                pl_prices[part_number] = (prev[0] or unit_price, new_amount, prev[2] or desc, prev[3] or desc_es, prev[4] or model_val)
            else:
                pl_prices[part_number] = (unit_price, amount, desc, desc_es, model_val)

        pl_item_records.append(PackingListItem(
            packing_list_id=pl_record.id,
            part_number=part_number,
            description=desc,
            qty=qty,
        ))

    for pli in pl_item_records:
        db.add(pli)
    await db.flush()

    # Obtener SparePartItems del lote
    lot_items_result = await db.execute(
        select(SparePartItem).where(SparePartItem.lot_id == lot.id)
    )
    lot_items: dict[str, SparePartItem] = {i.part_number: i for i in lot_items_result.scalars().all()}

    # Cruzar
    counts = {"complete": 0, "partial": 0, "missing": 0, "extra": 0}
    reconciled_parts: set[str] = set()

    for part_number, qty_in_pl in pl_items.items():
        reconciled_parts.add(part_number)
        sp_item = lot_items.get(part_number)

        if sp_item is None:
            # EXTRA: está en el packing list pero no en la orden
            rr = ReconciliationResult(
                lot_id=lot.id,
                packing_list_id=pl_record.id,
                spare_part_item_id=None,
                part_number=part_number,
                qty_ordered=None,
                qty_in_packing=qty_in_pl,
                result="EXTRA",
            )
            db.add(rr)
            counts["extra"] += 1
        else:
            qty_ordered = sp_item.qty_ordered
            if qty_in_pl >= qty_ordered:
                result_code = "COMPLETE"
                counts["complete"] += 1
            else:
                result_code = "PARTIAL"
                counts["partial"] += 1

            rr = ReconciliationResult(
                lot_id=lot.id,
                packing_list_id=pl_record.id,
                spare_part_item_id=sp_item.id,
                part_number=part_number,
                qty_ordered=qty_ordered,
                qty_in_packing=qty_in_pl,
                result=result_code,
            )
            db.add(rr)

    # MISSING: en la orden pero no en el packing list
    for part_number, sp_item in lot_items.items():
        if part_number not in reconciled_parts:
            rr = ReconciliationResult(
                lot_id=lot.id,
                packing_list_id=pl_record.id,
                spare_part_item_id=sp_item.id,
                part_number=part_number,
                qty_ordered=sp_item.qty_ordered,
                qty_in_packing=0,
                result="MISSING",
            )
            db.add(rr)
            counts["missing"] += 1

    # Si es invoice, actualizar precios y descripciones en los SparePartItems
    # y calcular total_declared_value del lote sumando los amounts del invoice
    invoice_parts_with_price: set[str] = set()
    if is_invoice and pl_prices:
        total_declared = 0.0
        for part_number, (unit_price, amount, desc_en, desc_es, model_val) in pl_prices.items():
            if amount is not None:
                total_declared += amount
            sp_item = lot_items.get(part_number)
            if sp_item is None:
                continue
            if unit_price is not None:
                sp_item.unit_price = unit_price
                invoice_parts_with_price.add(part_number)
            if amount is not None:
                sp_item.amount = amount
            if desc_en:
                sp_item.description = desc_en
            if desc_es:
                sp_item.description_es = desc_es
            if model_val:
                sp_item.model_applicable = model_val

        if total_declared > 0:
            lot.total_declared_value = total_declared

    pl_record.processed = True
    lot.packing_list_received = True
    await db.flush()

    if invoice_parts_with_price:
        from app.services.pricing_service import recalculate_part_cost
        for pn in invoice_parts_with_price:
            await recalculate_part_cost(db, pn, lot_identifier=lot.lot_identifier)

    return {**counts, "pl_id": str(pl_record.id), "total_parts_in_pl": len(pl_items), "is_invoice": is_invoice}


async def confirm_reconciliation(
    db: AsyncSession,
    lot: SparePartLot,
    actor: CurrentUser,
) -> dict:
    """
    Aplica los resultados de reconciliación a los SparePartItems:
    - COMPLETE → qty_received = qty_ordered, status = RECEIVED
    - PARTIAL  → qty_received = qty_in_packing, status = PARTIAL, crea Backorder por pendiente
    - MISSING  → status = BACKORDER, crea Backorder por total
    Marca lot.packing_list_received = True.
    """
    results = (await db.execute(
        select(ReconciliationResult).where(ReconciliationResult.lot_id == lot.id)
    )).scalars().all()

    if not results:
        return {"error": "No hay resultados de reconciliación para confirmar"}

    # Obtener lot_identifier (= origin_pi para backorders)
    origin_pi = lot.lot_identifier
    backorders_created = 0
    updated = 0

    for rr in results:
        if rr.spare_part_item_id is None:
            continue  # EXTRA: no hay item que actualizar
        item = await db.get(SparePartItem, rr.spare_part_item_id)
        if not item:
            continue

        if rr.result == "COMPLETE":
            item.qty_received = item.qty_ordered
            item.qty_pending = 0
            item.status = "RECEIVED"
            # Resolver backorders previos de esta parte en este lote
            await _resolve_backorders_for_item(db, item, origin_pi)

        elif rr.result == "PARTIAL":
            item.qty_received = rr.qty_in_packing or 0
            item.qty_pending = max(0, item.qty_ordered - item.qty_received)
            item.status = "PARTIAL"
            if item.qty_pending > 0:
                bo = await _upsert_backorder(db, item, origin_pi, item.qty_pending)
                if bo:
                    backorders_created += 1

        elif rr.result == "MISSING":
            if item.qty_received == 0:
                item.status = "BACKORDER"
            item.qty_pending = item.qty_ordered - item.qty_received
            qty_bo = item.qty_ordered - item.qty_received
            if qty_bo > 0:
                bo = await _upsert_backorder(db, item, origin_pi, qty_bo)
                if bo:
                    backorders_created += 1

        item.updated_at = datetime.utcnow()
        updated += 1

    # Cruzar EXTRAs contra backorders abiertos (FIFO por fecha de creación)
    backorders_resolved_by_extra = 0
    for rr in results:
        if rr.result != "EXTRA":
            continue
        surplus_qty = rr.qty_in_packing or 0
        if surplus_qty <= 0:
            continue

        open_bos = (await db.execute(
            select(Backorder)
            .where(Backorder.part_number == rr.part_number, Backorder.resolved == False)
            .order_by(Backorder.created_at.asc())
        )).scalars().all()

        now = datetime.utcnow()
        applied_any = False

        for bo in open_bos:
            if surplus_qty <= 0:
                break
            apply_qty = min(surplus_qty, bo.qty_pending)
            surplus_qty -= apply_qty
            bo.qty_pending -= apply_qty

            history = list(bo.history or [])
            history.append({
                "date": now.isoformat(),
                "event": "FILLED_BY_EXTRA",
                "pi": origin_pi,
                "qty_applied": apply_qty,
            })

            sp_item = await db.get(SparePartItem, bo.spare_part_item_id)
            if bo.qty_pending == 0:
                bo.resolved = True
                bo.resolved_at = now
                history.append({"date": now.isoformat(), "event": "RESOLVED", "resolved_in_pi": origin_pi})
                backorders_resolved_by_extra += 1
                if sp_item:
                    sp_item.qty_received = sp_item.qty_ordered
                    sp_item.qty_pending = 0
                    sp_item.status = "RECEIVED"
                    sp_item.updated_at = now
            else:
                if sp_item:
                    sp_item.qty_received = (sp_item.qty_received or 0) + apply_qty
                    sp_item.qty_pending = max(0, sp_item.qty_ordered - sp_item.qty_received)
                    sp_item.updated_at = now

            bo.history = history
            bo.updated_at = now
            applied_any = True

        if applied_any:
            rr.result = "EXTRA_APPLIED"

    lot.packing_list_received = True
    await db.commit()

    return {
        "confirmed": updated,
        "backorders_created": backorders_created,
        "backorders_resolved_by_extra": backorders_resolved_by_extra,
        "lot_id": str(lot.id),
    }


async def _upsert_backorder(
    db: AsyncSession,
    item: SparePartItem,
    origin_pi: str,
    qty_pending: int,
    source: str = 'reconciliation',
    already_charged: bool = False,
) -> Optional["Backorder"]:
    """Crea o actualiza un backorder para un item. Evita duplicados por (item, origin_pi, source)."""
    existing = (await db.execute(
        select(Backorder).where(
            Backorder.spare_part_item_id == item.id,
            Backorder.origin_pi == origin_pi,
            Backorder.source == source,
            Backorder.resolved == False,
        )
    )).scalar_one_or_none()

    if existing:
        existing.qty_pending = qty_pending
        existing.updated_at = datetime.utcnow()
        history = list(existing.history or [])
        history.append({"date": datetime.utcnow().isoformat(), "event": "UPDATE", "qty_pending": qty_pending})
        existing.history = history
        return None  # No cuenta como nuevo

    bo = Backorder(
        spare_part_item_id=item.id,
        part_number=item.part_number,
        origin_pi=origin_pi,
        qty_pending=qty_pending,
        source=source,
        already_charged=already_charged,
        resolved=False,
        history=[{"date": datetime.utcnow().isoformat(), "event": "CREATED", "qty_pending": qty_pending, "source": source}],
    )
    db.add(bo)
    await db.flush()
    return bo


async def apply_physical_inspection(
    db: AsyncSession,
    item: SparePartItem,
    qty_physical: int,
) -> None:
    """
    Aplica el resultado de la inspección física sobre un SparePartItem ya reconciliado.
    - physical_shortage = max(0, qty_received - qty_physical)  (cobrado y no llegó)
    - Recalcula status comparando qty_physical vs qty_ordered
    - Upsertea/resuelve el backorder de tipo physical_inspection
    """
    lot       = await db.get(SparePartLot, item.lot_id)
    origin_pi = lot.lot_identifier if lot else ''

    qty_received = item.qty_received or 0
    qty_ordered  = item.qty_ordered  or 0

    # Faltante no cobrado: estaba en la orden pero el proveedor nunca lo puso en el PL
    not_charged_qty   = max(0, qty_ordered - qty_received)
    # Faltante cobrado: estaba en el PL (cobrado) pero no llegó físicamente
    physical_shortage = max(0, qty_received - qty_physical)

    item.qty_physical = qty_physical
    item.qty_pending  = max(0, qty_ordered - qty_physical)
    item.status       = 'RECEIVED' if qty_physical >= qty_ordered else 'BACKORDER'
    item.updated_at   = datetime.utcnow()

    # Garantizar que el BO de reconciliación siempre refleje el faltante no cobrado.
    # Si fue resuelto o nunca existió, lo recrea con el valor correcto.
    if not_charged_qty > 0:
        await _upsert_backorder(
            db, item, origin_pi, not_charged_qty,
            source='reconciliation', already_charged=False,
        )

    # BO de inspección física: cobrado pero no llegó
    if physical_shortage > 0:
        await _upsert_backorder(
            db, item, origin_pi, physical_shortage,
            source='physical_inspection', already_charged=True,
        )
    else:
        existing = (await db.execute(
            select(Backorder).where(
                Backorder.spare_part_item_id == item.id,
                Backorder.source == 'physical_inspection',
                Backorder.resolved == False,
            )
        )).scalar_one_or_none()
        if existing:
            now = datetime.utcnow()
            existing.resolved    = True
            existing.resolved_at = now
            existing.updated_at  = now
            history = list(existing.history or [])
            history.append({"date": now.isoformat(), "event": "RESOLVED_BY_CORRECTION"})
            existing.history = history


async def _resolve_backorders_for_item(
    db: AsyncSession,
    item: SparePartItem,
    resolved_in_pi: str,
) -> int:
    """Marca como resueltos todos los backorders abiertos de un item."""
    open_bos = (await db.execute(
        select(Backorder).where(
            Backorder.spare_part_item_id == item.id,
            Backorder.resolved == False,
        )
    )).scalars().all()

    now = datetime.utcnow()
    for bo in open_bos:
        bo.resolved = True
        bo.resolved_at = now
        history = list(bo.history or [])
        history.append({"date": now.isoformat(), "event": "RESOLVED", "resolved_in_pi": resolved_in_pi})
        bo.history = history
        bo.updated_at = now

    return len(open_bos)


async def list_backorders(
    db: AsyncSession,
    resolved: Optional[bool] = None,
    part_number: Optional[str] = None,
    origin_pi: Optional[str] = None,
) -> list:
    """Listado de backorders con filtros opcionales, enriquecidos con datos del SparePartItem."""
    stmt = select(Backorder)
    if resolved is not None:
        stmt = stmt.where(Backorder.resolved == resolved)
    if part_number:
        stmt = stmt.where(Backorder.part_number.ilike(f"%{part_number}%"))
    if origin_pi:
        stmt = stmt.where(Backorder.origin_pi.ilike(f"%{origin_pi}%"))
    stmt = stmt.order_by(Backorder.resolved.asc(), Backorder.created_at.desc())
    backorders = (await db.execute(stmt)).scalars().all()

    # Enriquecer con description_es y model_applicable del SparePartItem
    item_ids = [bo.spare_part_item_id for bo in backorders if bo.spare_part_item_id]
    items_map: dict = {}
    if item_ids:
        items_result = await db.execute(
            select(SparePartItem).where(SparePartItem.id.in_(item_ids))
        )
        items_map = {i.id: i for i in items_result.scalars().all()}

    result = []
    for bo in backorders:
        sp = items_map.get(bo.spare_part_item_id)
        d = {
            "id": bo.id,
            "spare_part_item_id": bo.spare_part_item_id,
            "part_number": bo.part_number,
            "description_es": sp.description_es if sp else None,
            "model_applicable": sp.model_applicable if sp else None,
            "origin_pi": bo.origin_pi,
            "expected_in_pi": bo.expected_in_pi,
            "qty_pending": bo.qty_pending,
            "source": bo.source,
            "already_charged": bo.already_charged,
            "resolved": bo.resolved,
            "resolved_at": bo.resolved_at,
            "history": bo.history,
            "created_at": bo.created_at,
            "updated_at": bo.updated_at,
        }
        result.append(d)

    return result


# ---------------------------------------------------------------------------
# Utilidad interna: escribir audit log
# ---------------------------------------------------------------------------

def _log_audit(
    db: AsyncSession,
    actor: CurrentUser,
    action: str,
    entity_type: str,
    entity_id: str,
    payload: dict,
    shipment_order_id=None,
) -> None:
    log = ImportAuditLog(
        shipment_order_id=shipment_order_id,
        actor_id=uuid.UUID(actor.user_id) if actor.user_id else None,
        actor_role=actor.role,
        action=action,
        entity_type=entity_type,
        entity_id=entity_id,
        payload=payload,
        created_at=datetime.utcnow(),
    )
    db.add(log)
