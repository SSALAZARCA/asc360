"""
PR4 (`sdd/parts-description-source-of-truth`, Phase 4, tasks 4.1-4.8):
live-read conversion of the 6 imports-domain read paths (R1-R6, design
D11-D14).

Before this PR, `SparePartItem.description_es` / `ReconciliationResult
.description_es` were frozen copies written at import time -- editing a
part's name via Maestro de Partes, Ajuste de Pedidos, Repuestos tab or
Reconciliación (PR1-PR3's shared write path,
`parts_description_service.set_description_es`) never propagated to these
6 read paths. This PR makes every one of them resolve the displayed name
live against `parts_references.description_es_manual`, falling back to the
existing stored value when uncatalogued or blank (D12): plain Python `or`,
exactly `resolved.get(code) or stored_value`, NOT SQL COALESCE -- so a
blank/None manual name correctly degrades to the previous stored value.

Covers:
- R1 `GET /imports/spare-part-lots/{lot_id}/items`
- R2/R3 `_fetch_enriched_reconciliation` (list + export)
- R4 `GET /imports/spare-parts/export`
- R5 `GET /imports/spare-parts/search` (item branch only -- EXTRA rows stay raw)
- R6 `imports_service.list_backorders` (list + export)
- 4.7: R1/R4/R6 issue the SAME `db.execute` call count as before (no new query)
- 4.8: rename on one surface -> all read paths reflect it with no sync job;
  regression -- re-importing the OLD text does not un-correct the display

PR4 post-review fix pass (finding #1): R2/R3 and R5 are the only 2 read
paths that call `parts_description_service.resolve_names` as a BRAND-NEW
query (R1/R4/R6 just extend a query they were already running, so they have
no new failure surface). If `resolve_names` raises for either of these 2
call sites, the request must degrade to showing stored values instead of
500ing, and the failure must be logged -- see
`TestR5ResolveNamesFailureDegradesGracefully` and
`TestR2R3ResolveNamesFailureDegradesGracefully` below.
"""
import logging
import uuid
from datetime import datetime
from types import SimpleNamespace
from unittest.mock import AsyncMock, patch

from app.models.imports import SparePartLot, SparePartItem
from app.models.parts_manual import PartsReference

from tests.conftest import make_test_client
from tests.imports.conftest import (
    FakeAsyncSession,
    make_backorder,
    make_imports_editor,
    make_lot,
    make_reconciliation_result,
    make_spare_part_item,
)


def _ref(fpn: str, description_es_manual=None, rotation_class=None, prev_codes=None):
    return SimpleNamespace(
        factory_part_number=fpn,
        rotation_class=rotation_class,
        description_es_manual=description_es_manual,
    )


# ---------------------------------------------------------------------------
# R1 -- GET /imports/spare-part-lots/{lot_id}/items
# ---------------------------------------------------------------------------

class TestR1ListSparePartItems:
    def test_manual_name_wins_over_stale_stored_value(self):
        lot = make_lot()
        item = make_spare_part_item(
            lot_id=lot.id, part_number="FPN-1", qty_ordered=1,
            description_es="Old stale name",
        )

        fake_db = FakeAsyncSession(
            execute_queue=[
                [item],  # SparePartItem select
                [_ref("FPN-1", description_es_manual="New confirmed name")],  # PartsReference select
            ],
            get_objects=[lot],
        )

        with make_test_client(current_user=make_imports_editor(), fake_db_session=fake_db) as client:
            resp = client.get(f"/api/v1/imports/spare-part-lots/{lot.id}/items")

        assert resp.status_code == 200
        assert resp.json()[0]["description_es"] == "New confirmed name"

    def test_uncatalogued_code_falls_back_to_stored_value(self):
        lot = make_lot()
        item = make_spare_part_item(
            lot_id=lot.id, part_number="FPN-UNCAT", qty_ordered=1,
            description_es="Stored from import",
        )

        fake_db = FakeAsyncSession(
            execute_queue=[
                [item],
                [],  # no matching PartsReference
            ],
            get_objects=[lot],
        )

        with make_test_client(current_user=make_imports_editor(), fake_db_session=fake_db) as client:
            resp = client.get(f"/api/v1/imports/spare-part-lots/{lot.id}/items")

        assert resp.status_code == 200
        assert resp.json()[0]["description_es"] == "Stored from import"

    def test_blank_manual_name_falls_back_to_stored_value(self):
        lot = make_lot()
        item = make_spare_part_item(
            lot_id=lot.id, part_number="FPN-2", qty_ordered=1,
            description_es="Stored value survives",
        )

        fake_db = FakeAsyncSession(
            execute_queue=[
                [item],
                [_ref("FPN-2", description_es_manual=None)],
            ],
            get_objects=[lot],
        )

        with make_test_client(current_user=make_imports_editor(), fake_db_session=fake_db) as client:
            resp = client.get(f"/api/v1/imports/spare-part-lots/{lot.id}/items")

        assert resp.status_code == 200
        assert resp.json()[0]["description_es"] == "Stored value survives"

    def test_no_items_skips_the_partsreference_query_entirely(self):
        """4.7 -- same call count as before when there is nothing to resolve."""
        lot = make_lot()
        fake_db = FakeAsyncSession(execute_queue=[[]], get_objects=[lot])

        with make_test_client(current_user=make_imports_editor(), fake_db_session=fake_db) as client:
            resp = client.get(f"/api/v1/imports/spare-part-lots/{lot.id}/items")

        assert resp.status_code == 200
        assert resp.json() == []
        assert len(fake_db.executed_statements) == 1

    def test_one_item_issues_exactly_two_execute_calls_no_new_query(self):
        """4.7 -- R1 reuses the existing PartsReference select (rotation_class
        query), it does not add a THIRD query for the name resolution."""
        lot = make_lot()
        item = make_spare_part_item(lot_id=lot.id, part_number="FPN-3", qty_ordered=1)
        fake_db = FakeAsyncSession(
            execute_queue=[[item], [_ref("FPN-3", description_es_manual="X")]],
            get_objects=[lot],
        )

        with make_test_client(current_user=make_imports_editor(), fake_db_session=fake_db) as client:
            resp = client.get(f"/api/v1/imports/spare-part-lots/{lot.id}/items")

        assert resp.status_code == 200
        assert len(fake_db.executed_statements) == 2


# ---------------------------------------------------------------------------
# R2/R3 -- _fetch_enriched_reconciliation (list + export share the helper)
# ---------------------------------------------------------------------------

class TestR2R3EnrichedReconciliation:
    def test_manual_name_wins_over_stale_stored_value_on_list(self):
        lot = make_lot()
        item = make_spare_part_item(
            lot_id=lot.id, part_number="FPN-R2", qty_ordered=1, description_es="Old stale name",
        )
        rr = make_reconciliation_result(
            lot_id=lot.id, part_number="FPN-R2", result="COMPLETE",
            spare_part_item_id=item.id, description_es="Even older",
        )
        rr.created_at = datetime.utcnow()

        fake_db = FakeAsyncSession(
            execute_queue=[
                [rr],       # ReconciliationResult select
                [item],     # SparePartItem select
                [("FPN-R2", "New confirmed name")],  # resolve_names query
            ],
            get_objects=[lot],
        )

        with make_test_client(current_user=make_imports_editor(), fake_db_session=fake_db) as client:
            resp = client.get(f"/api/v1/imports/spare-part-lots/{lot.id}/reconciliation")

        assert resp.status_code == 200
        assert resp.json()[0]["description_es"] == "New confirmed name"

    def test_uncatalogued_code_falls_back_to_sp_stored_value(self):
        lot = make_lot()
        item = make_spare_part_item(
            lot_id=lot.id, part_number="FPN-R2B", qty_ordered=1, description_es="Stored from import",
        )
        rr = make_reconciliation_result(
            lot_id=lot.id, part_number="FPN-R2B", result="COMPLETE", spare_part_item_id=item.id,
        )
        rr.created_at = datetime.utcnow()

        fake_db = FakeAsyncSession(
            execute_queue=[[rr], [item], []],
            get_objects=[lot],
        )

        with make_test_client(current_user=make_imports_editor(), fake_db_session=fake_db) as client:
            resp = client.get(f"/api/v1/imports/spare-part-lots/{lot.id}/reconciliation")

        assert resp.status_code == 200
        assert resp.json()[0]["description_es"] == "Stored from import"

    def test_blank_manual_falls_back_to_sp_stored_value(self):
        lot = make_lot()
        item = make_spare_part_item(
            lot_id=lot.id, part_number="FPN-R2C", qty_ordered=1, description_es="Stored value survives",
        )
        rr = make_reconciliation_result(
            lot_id=lot.id, part_number="FPN-R2C", result="COMPLETE", spare_part_item_id=item.id,
        )
        rr.created_at = datetime.utcnow()

        fake_db = FakeAsyncSession(
            execute_queue=[[rr], [item], [("FPN-R2C", None)]],
            get_objects=[lot],
        )

        with make_test_client(current_user=make_imports_editor(), fake_db_session=fake_db) as client:
            resp = client.get(f"/api/v1/imports/spare-part-lots/{lot.id}/reconciliation")

        assert resp.status_code == 200
        assert resp.json()[0]["description_es"] == "Stored value survives"

    def test_pure_extra_row_without_linked_item_keeps_local_value_no_resolution(self):
        """Pure EXTRA rows (spare_part_item_id IS NULL) have no catalog
        identity -- design explicitly excludes them (D1/D22)."""
        lot = make_lot()
        rr = make_reconciliation_result(
            lot_id=lot.id, part_number="FPN-EXTRA", result="EXTRA",
            spare_part_item_id=None, description_es="Extra local value",
        )
        rr.created_at = datetime.utcnow()

        fake_db = FakeAsyncSession(
            execute_queue=[[rr], []],  # no item_ids -> no SparePartItem query;
            get_objects=[lot],          # resolve_names([]) short-circuits -> no 3rd query
        )

        with make_test_client(current_user=make_imports_editor(), fake_db_session=fake_db) as client:
            resp = client.get(f"/api/v1/imports/spare-part-lots/{lot.id}/reconciliation")

        assert resp.status_code == 200
        assert resp.json()[0]["description_es"] == "Extra local value"
        # Only the ReconciliationResult select ran -- no SparePartItem query
        # (no item_ids) and resolve_names([]) never issued its own execute.
        assert len(fake_db.executed_statements) == 1

    def test_export_shares_the_same_live_resolution_as_the_list(self):
        lot = make_lot(lot_identifier="E0000700-SP")
        item = make_spare_part_item(
            lot_id=lot.id, part_number="FPN-EXP", qty_ordered=1, unit_price=1.0, description_es="Old",
        )
        rr = make_reconciliation_result(
            lot_id=lot.id, part_number="FPN-EXP", result="COMPLETE",
            spare_part_item_id=item.id, qty_ordered=1, qty_in_packing=1,
        )

        fake_db = FakeAsyncSession(
            execute_queue=[[rr], [item], [("FPN-EXP", "New via export")]],
            get_objects=[lot],
        )

        with make_test_client(current_user=make_imports_editor(), fake_db_session=fake_db) as client:
            resp = client.get(f"/api/v1/imports/spare-part-lots/{lot.id}/reconciliation/export")

        assert resp.status_code == 200
        import io
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert rows[1][1] == "New via export"


class TestR2R3ResolveNamesFailureDegradesGracefully:
    """Fix pass finding #1: `resolve_names` is a BRAND-NEW query with no
    fallback of its own (unlike R1/R4/R6, which just extend a query the
    endpoint was already running). If it raises, the reconciliation list
    must still succeed and fall back to each linked item's stored
    `description_es` instead of 500ing the whole request -- and the failure
    must be logged."""

    def test_reconciliation_list_falls_back_to_stored_values_when_resolve_names_raises(self, caplog):
        lot = make_lot()
        item = make_spare_part_item(
            lot_id=lot.id, part_number="FPN-R2-FAIL", qty_ordered=1,
            description_es="Stored despite resolver failure",
        )
        rr = make_reconciliation_result(
            lot_id=lot.id, part_number="FPN-R2-FAIL", result="COMPLETE", spare_part_item_id=item.id,
        )
        rr.created_at = datetime.utcnow()

        # No slot queued for resolve_names -- it's mocked to raise before it
        # would ever reach `db.execute()`.
        fake_db = FakeAsyncSession(
            execute_queue=[[rr], [item]],
            get_objects=[lot],
        )

        with patch(
            "app.services.parts_description_service.resolve_names",
            new=AsyncMock(side_effect=RuntimeError("db unavailable")),
        ):
            with caplog.at_level(logging.WARNING, logger="app.api.v1.imports"):
                with make_test_client(current_user=make_imports_editor(), fake_db_session=fake_db) as client:
                    resp = client.get(f"/api/v1/imports/spare-part-lots/{lot.id}/reconciliation")

        assert resp.status_code == 200
        assert resp.json()[0]["description_es"] == "Stored despite resolver failure"
        assert "resolve_names" in caplog.text


# ---------------------------------------------------------------------------
# R4 -- GET /imports/spare-parts/export (existing rotation test file
# extended separately; see test_export_spare_parts_rotation.py). Additional
# manual-name-wins case lives here for locality with the other read paths.
# ---------------------------------------------------------------------------

class TestR4SparePartsExport:
    def test_manual_name_wins_over_stale_stored_value(self):
        lot = SparePartLot(
            id=uuid.uuid4(), shipment_order_id=uuid.uuid4(),
            lot_identifier="E0000800-SP", detail_loaded=True,
        )
        item = SparePartItem(
            id=uuid.uuid4(), lot_id=lot.id, part_number="FPN-R4",
            description="ENGINE PART", description_es="Old stale name",
            model_applicable="MODEL X", qty_ordered=1, qty_received=1, qty_pending=0,
            status="RECEIVED",
        )
        lot.items = [item]

        fake_db = FakeAsyncSession(execute_queue=[
            [lot],
            [("FPN-R4", None, "New confirmed name")],
        ])

        with make_test_client(current_user=make_imports_editor(role="superadmin"), fake_db_session=fake_db) as client:
            resp = client.get("/api/v1/imports/spare-parts/export")

        assert resp.status_code == 200
        import io
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        headers = rows[0]
        desc_idx = headers.index("DESCRIPCIÓN ES")
        assert rows[1][desc_idx] == "New confirmed name"


# ---------------------------------------------------------------------------
# R5 -- GET /imports/spare-parts/search (item branch only)
# ---------------------------------------------------------------------------

class TestR5SearchSpareParts:
    def test_manual_name_wins_for_ordered_item_match(self):
        lot = make_lot(lot_identifier="S05099")
        item = make_spare_part_item(
            lot_id=lot.id, part_number="FPN-R5", qty_ordered=5, qty_received=5, status="RECEIVED",
            description_es="Old stale name",
        )

        fake_db = FakeAsyncSession(execute_queue=[
            [(item, lot)],                        # SparePartItem join
            [("FPN-R5", "New confirmed name")],    # resolve_names (issued right after the item join)
            [],                                     # ReconciliationResult EXTRA join
        ])

        with make_test_client(current_user=make_imports_editor(), fake_db_session=fake_db) as client:
            resp = client.get("/api/v1/imports/spare-parts/search", params={"q": "FPN-R5"})

        assert resp.status_code == 200
        assert resp.json()[0]["items"][0]["description_es"] == "New confirmed name"

    def test_extra_row_without_linked_item_keeps_local_value(self):
        lot = make_lot(lot_identifier="S05100")
        rr = make_reconciliation_result(
            lot_id=lot.id, part_number="FPN-EXTRA5", result="EXTRA",
            spare_part_item_id=None, description_es="Extra local value", qty_in_packing=2,
        )

        fake_db = FakeAsyncSession(execute_queue=[
            [],           # no ordered item match -> resolve_names([]) short-circuits, no 3rd query
            [(rr, lot)],
        ])

        with make_test_client(current_user=make_imports_editor(), fake_db_session=fake_db) as client:
            resp = client.get("/api/v1/imports/spare-parts/search", params={"q": "FPN-EXTRA5"})

        assert resp.status_code == 200
        assert resp.json()[0]["items"][0]["description_es"] == "Extra local value"
        assert len(fake_db.executed_statements) == 2


class TestR5ResolveNamesFailureDegradesGracefully:
    """Fix pass finding #1: `resolve_names` is a BRAND-NEW query with no
    fallback of its own (unlike R1/R4/R6, which just extend a query the
    endpoint was already running). If it raises, the search must still
    succeed and fall back to each item's stored `description_es` instead of
    500ing the whole request -- and the failure must be logged."""

    def test_search_falls_back_to_stored_values_when_resolve_names_raises(self, caplog):
        lot = make_lot(lot_identifier="S05200")
        item = make_spare_part_item(
            lot_id=lot.id, part_number="FPN-R5-FAIL", qty_ordered=1,
            description_es="Stored despite resolver failure",
        )

        # No slot queued for resolve_names -- it's mocked to raise before it
        # would ever reach `db.execute()`.
        fake_db = FakeAsyncSession(execute_queue=[
            [(item, lot)],  # SparePartItem join
            [],              # ReconciliationResult EXTRA join
        ])

        with patch(
            "app.services.parts_description_service.resolve_names",
            new=AsyncMock(side_effect=RuntimeError("db unavailable")),
        ):
            with caplog.at_level(logging.WARNING, logger="app.api.v1.imports"):
                with make_test_client(current_user=make_imports_editor(), fake_db_session=fake_db) as client:
                    resp = client.get("/api/v1/imports/spare-parts/search", params={"q": "FPN-R5-FAIL"})

        assert resp.status_code == 200
        assert resp.json()[0]["items"][0]["description_es"] == "Stored despite resolver failure"
        assert "resolve_names" in caplog.text


# ---------------------------------------------------------------------------
# R6 -- imports_service.list_backorders (GET /imports/backorders + export)
# ---------------------------------------------------------------------------

class TestR6Backorders:
    def test_manual_name_wins_over_stale_stored_value(self):
        lot = make_lot()
        item = make_spare_part_item(
            lot_id=lot.id, part_number="FPN-R6", qty_ordered=1, description_es="Old stale name",
        )
        bo = make_backorder(
            spare_part_item_id=item.id, part_number="FPN-R6", origin_pi="E0000900", qty_pending=1,
        )

        fake_db = FakeAsyncSession(execute_queue=[
            [bo],
            [item],
            [_ref("FPN-R6", description_es_manual="New confirmed name")],
        ])

        with make_test_client(current_user=make_imports_editor(), fake_db_session=fake_db) as client:
            resp = client.get("/api/v1/imports/backorders")

        assert resp.status_code == 200
        assert resp.json()[0]["description_es"] == "New confirmed name"

    def test_uncatalogued_code_falls_back_to_sp_stored_value(self):
        lot = make_lot()
        item = make_spare_part_item(
            lot_id=lot.id, part_number="FPN-R6B", qty_ordered=1, description_es="Stored from import",
        )
        bo = make_backorder(
            spare_part_item_id=item.id, part_number="FPN-R6B", origin_pi="E0000901", qty_pending=1,
        )

        fake_db = FakeAsyncSession(execute_queue=[[bo], [item], []])

        with make_test_client(current_user=make_imports_editor(), fake_db_session=fake_db) as client:
            resp = client.get("/api/v1/imports/backorders")

        assert resp.status_code == 200
        assert resp.json()[0]["description_es"] == "Stored from import"

    def test_blank_manual_falls_back_to_sp_stored_value(self):
        lot = make_lot()
        item = make_spare_part_item(
            lot_id=lot.id, part_number="FPN-R6C", qty_ordered=1, description_es="Stored value survives",
        )
        bo = make_backorder(
            spare_part_item_id=item.id, part_number="FPN-R6C", origin_pi="E0000902", qty_pending=1,
        )

        fake_db = FakeAsyncSession(execute_queue=[
            [bo], [item], [_ref("FPN-R6C", description_es_manual=None)],
        ])

        with make_test_client(current_user=make_imports_editor(), fake_db_session=fake_db) as client:
            resp = client.get("/api/v1/imports/backorders")

        assert resp.status_code == 200
        assert resp.json()[0]["description_es"] == "Stored value survives"

    def test_no_part_numbers_skips_the_partsreference_query_entirely(self):
        """4.7 -- same call count as before when there are no backorders."""
        fake_db = FakeAsyncSession(execute_queue=[[]])

        with make_test_client(current_user=make_imports_editor(), fake_db_session=fake_db) as client:
            resp = client.get("/api/v1/imports/backorders")

        assert resp.status_code == 200
        assert resp.json() == []
        assert len(fake_db.executed_statements) == 1

    def test_one_backorder_issues_the_same_execute_call_count_as_before(self):
        """4.7 -- R6 reuses the existing PartsReference select (rotation_class
        query, extended with `description_es_manual`); it does not add a
        FOURTH query for the name resolution. `spare_part_item_id` is
        NOT NULL at the model level, so a real backorder always has a
        linked item -- the pre-PR4 call shape was already 3 queries
        (Backorder, SparePartItem, PartsReference)."""
        lot = make_lot()
        item = make_spare_part_item(lot_id=lot.id, part_number="FPN-R6D", qty_ordered=1)
        bo = make_backorder(
            spare_part_item_id=item.id, part_number="FPN-R6D", origin_pi="E0000903", qty_pending=1,
        )
        fake_db = FakeAsyncSession(execute_queue=[
            [bo],
            [item],
            [_ref("FPN-R6D", description_es_manual="X")],
        ])

        with make_test_client(current_user=make_imports_editor(), fake_db_session=fake_db) as client:
            resp = client.get("/api/v1/imports/backorders")

        assert resp.status_code == 200
        assert len(fake_db.executed_statements) == 3


# ---------------------------------------------------------------------------
# 4.8 -- Integration: rename on one surface -> all read paths reflect it
# with no sync job. Regression: re-importing the OLD text does not
# un-correct the display.
# ---------------------------------------------------------------------------

class TestRenamePropagatesToEveryReadPathNoSyncJob:
    """One `PartsReference.description_es_manual` value ("Nombre corregido")
    is the single fact every read path is asked to resolve against, while
    each surface's OWN stored/local column still holds the pre-correction
    text ("Nombre viejo (del import)") -- proving there is no background job
    keeping them in sync, just a live read at request time (D11)."""

    FPN = "FPN-INTEGRATION"
    CORRECTED = "Nombre corregido"
    STALE = "Nombre viejo (del import)"

    def test_r1_repuestos_list_reads_the_corrected_name(self):
        lot = make_lot()
        item = make_spare_part_item(
            lot_id=lot.id, part_number=self.FPN, qty_ordered=1, description_es=self.STALE,
        )
        fake_db = FakeAsyncSession(
            execute_queue=[[item], [_ref(self.FPN, description_es_manual=self.CORRECTED)]],
            get_objects=[lot],
        )
        with make_test_client(current_user=make_imports_editor(), fake_db_session=fake_db) as client:
            resp = client.get(f"/api/v1/imports/spare-part-lots/{lot.id}/items")
        assert resp.json()[0]["description_es"] == self.CORRECTED

    def test_r6_backorders_reads_the_same_corrected_name(self):
        lot = make_lot()
        item = make_spare_part_item(
            lot_id=lot.id, part_number=self.FPN, qty_ordered=1,
            description_es=self.STALE,  # same stale text re-imported would still write here
        )
        bo = make_backorder(spare_part_item_id=item.id, part_number=self.FPN, origin_pi="E1", qty_pending=1)
        fake_db = FakeAsyncSession(execute_queue=[
            [bo], [item], [_ref(self.FPN, description_es_manual=self.CORRECTED)],
        ])
        with make_test_client(current_user=make_imports_editor(), fake_db_session=fake_db) as client:
            resp = client.get("/api/v1/imports/backorders")
        assert resp.json()[0]["description_es"] == self.CORRECTED

    def test_reimporting_the_old_text_does_not_un_correct_the_display(self):
        """Regression: a re-imported packing list overwrites
        `SparePartItem.description_es` back to the OLD text (D14 -- ingest
        writers are NOT changed by this PR), but the live read still shows
        the corrected name because it never trusts the stored column when a
        manual name exists."""
        lot = make_lot()
        item = make_spare_part_item(
            lot_id=lot.id, part_number=self.FPN, qty_ordered=1,
            description_es=self.STALE,  # simulates a re-import overwriting the stored column
        )
        fake_db = FakeAsyncSession(
            execute_queue=[[item], [_ref(self.FPN, description_es_manual=self.CORRECTED)]],
            get_objects=[lot],
        )
        with make_test_client(current_user=make_imports_editor(), fake_db_session=fake_db) as client:
            resp = client.get(f"/api/v1/imports/spare-part-lots/{lot.id}/items")
        assert resp.json()[0]["description_es"] == self.CORRECTED
