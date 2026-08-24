"""
Approval/regression tests for `export_moto_units`
(`GET /api/v1/imports/moto-units/export`) capturing CURRENT (pre-refactor)
behavior — Phase 1 of `sdd/imports-oversized-functions-refactor-tier2`.

These exercise the endpoint via the HTTP harness (`make_test_client`) +
`FakeAsyncSession`, parsing the returned .xlsx bytes with `openpyxl`, and
MUST keep passing unchanged after the extraction (endpoint delegates to the
existing `_build_moto_unit_filters` helper + a new
`_build_moto_unit_export_row(unit)` helper) — same file, same assertions,
only the endpoint's internals change.

Locked-in behaviors (read directly off the pre-refactor source,
`backend/app/api/v1/imports.py` lines 2416-2504, before writing any test):
  - `_require_superadmin` gate: role must be `superadmin`, error string
    "Solo superadmin puede realizar esta acción" — NOTE this endpoint uses
    the SUPERADMIN gate, not `_require_imports_editor` like most other
    imports endpoints.
  - The endpoint currently builds its OWN `base_filters`/`filters` inline,
    duplicating `_build_moto_unit_filters` — the query uses `filters` (the
    full set including all 5 optional status params); `base_filters` is
    built but never consumed by anything else in this endpoint (dead code
    to dedupe by reusing the shared helper).
  - Single `select(ShipmentMotoUnit)` query, joined to `ShipmentOrder`,
    `selectinload`s `shipment_order`/`location`/`observation`, ordered by
    `created_at.desc()` — exactly ONE `db.execute()` call, no pagination.
  - Excel sheet title "Motocicletas"; headers (exact order): PI NUMBER,
    MODELO, AÑO MODELO, VIN, No. MOTOR, COLOR, COLOR RUNT, No. LEVANTE,
    UBICACIÓN, SEP. NACIONALIZACIÓN, OBSERVACIÓN, EMPADRONADO, FECHA
    EMPADRONAMIENTO, EMPADR. FÍSICO ENVIADO, FECHA EMPADR. FÍSICO,
    DISTRIBUIDOR, FACTURADO, CARGADO RUNT, DIM CARGADO.
  - Row mapping per unit: `pi_number` from the related order (fallback
    `u.source_pi`), `model` from `u.model` (fallback `o.model`)
    upper-cased/whitespace-collapsed (empty string -> `None`), `model_year`
    from `u.model_year` (fallback `o.model_year`), `vin_number`,
    `engine_number`, `color`, `color_runt`, `no_lev`, `location.name` (or
    `None`), boolean-ish fields rendered as literal "Sí"/"No" strings
    (`separada_nacionalizacion`, `certificado_generado`,
    `empadronamiento_fisico_enviado`, `facturado`, `cargado_runt`, and
    whether `dim_pdf_object_name` is set), `observation.name` (or `None`),
    dates formatted `%Y-%m-%d` (or `None` when unset).
  - Response filename: `motocicletas_{utcnow:%Y%m%d_%H%M%S}.xlsx`; media
    type `application/vnd.openxmlformats-officedocument.spreadsheetml.sheet`.
"""
import io
import uuid
from datetime import datetime

import openpyxl

from tests.conftest import make_test_client
from tests.imports.conftest import (
    FakeAsyncSession,
    make_imports_editor,
    make_moto_location,
    make_moto_observation,
    make_moto_unit,
    make_shipment_order,
)

EXPECTED_HEADERS = [
    "PI NUMBER", "MODELO", "AÑO MODELO", "VIN", "No. MOTOR",
    "COLOR", "COLOR RUNT", "No. LEVANTE",
    "UBICACIÓN", "SEP. NACIONALIZACIÓN", "OBSERVACIÓN",
    "EMPADRONADO", "FECHA EMPADRONAMIENTO",
    "EMPADR. FÍSICO ENVIADO", "FECHA EMPADR. FÍSICO", "DISTRIBUIDOR",
    "FACTURADO", "CARGADO RUNT",
    "DIM CARGADO",
]


def _read_rows(content: bytes) -> list:
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    return [list(row) for row in ws.iter_rows(values_only=True)]


def test_non_imports_editor_gets_403():
    """2026-08-24 business decision: this endpoint was widened from
    `_require_superadmin` to `_require_imports_editor` (superadmin |
    proveedor | administrativo) -- the Motocicletas tab itself is already
    visible to all three via `ALL_TABS`' `roles: null`. `technician` is
    none of those three, so it still gets 403, no DB call. (See
    `test_administrativo_now_gets_200` below for the role that changed.)"""
    fake_db = FakeAsyncSession(execute_queue=[])

    with make_test_client(current_user=make_imports_editor(role="technician"), fake_db_session=fake_db) as client:
        response = client.get("/api/v1/imports/moto-units/export")

    assert response.status_code == 403
    assert response.json()["detail"] == "Sin permisos para el módulo de importaciones"
    assert fake_db.executed_statements == []


def test_administrativo_now_gets_200():
    """2026-08-24 business decision: administrativo now matches superadmin
    on this Excel export (previously blocked, see the docstring above)."""
    fake_db = FakeAsyncSession(execute_queue=[[]])  # empty result -- headers-only workbook is enough to prove 200

    with make_test_client(current_user=make_imports_editor(role="administrativo"), fake_db_session=fake_db) as client:
        response = client.get("/api/v1/imports/moto-units/export")

    assert response.status_code == 200


def test_proveedor_still_gets_200_unchanged():
    """Regression guard: `proveedor` already had access via
    `_require_imports_editor` at other imports endpoints (e.g. the
    reconciliation export) -- this widening must not have narrowed that."""
    fake_db = FakeAsyncSession(execute_queue=[[]])

    with make_test_client(current_user=make_imports_editor(role="proveedor"), fake_db_session=fake_db) as client:
        response = client.get("/api/v1/imports/moto-units/export")

    assert response.status_code == 200


def test_default_request_returns_expected_headers_and_full_row_values():
    order = make_shipment_order(pi_number="E0000900", model="MODEL FALLBACK", model_year=2024)
    location = make_moto_location(name="BODEGA NORTE")
    observation = make_moto_observation(name="REVISAR VIN")
    unit = make_moto_unit(
        shipment_order=order,
        location=location,
        observation=observation,
        vin_number="VIN999",
        engine_number="ENG999",
        model="  cb 500   f  ",
        model_year=2025,
        color="ROJO",
        color_runt="RED",
        no_lev="LEV-1",
        certificado_generado=True,
        certificado_fecha=datetime(2026, 1, 15),
        facturado=True,
        cargado_runt=True,
        separada_nacionalizacion=True,
        empadronamiento_fisico_enviado=True,
        empadronamiento_fisico_fecha=datetime(2026, 2, 20),
        empadronamiento_fisico_distribuidor_nombre="Distribuidor X",
        dim_pdf_object_name="dim/vin999.pdf",
    )

    fake_db = FakeAsyncSession(execute_queue=[[unit]])

    with make_test_client(current_user=make_imports_editor(role="superadmin"), fake_db_session=fake_db) as client:
        response = client.get("/api/v1/imports/moto-units/export")

    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response.headers["content-disposition"].startswith(
        'attachment; filename="motocicletas_'
    )
    assert response.headers["content-disposition"].endswith('.xlsx"')

    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    assert wb.active.title == "Motocicletas"

    rows = _read_rows(response.content)
    assert rows[0] == EXPECTED_HEADERS
    assert rows[1] == [
        "E0000900", "CB 500 F", 2025, "VIN999", "ENG999",
        "ROJO", "RED", "LEV-1",
        "BODEGA NORTE", "Sí", "REVISAR VIN",
        "Sí", "2026-01-15",
        "Sí", "2026-02-20", "Distribuidor X",
        "Sí", "Sí",
        "Sí",
    ]


def test_row_fallbacks_and_boolean_false_values():
    """Unit with no location/observation and unset optional fields — locks
    in the `model`/`model_year`/`pi_number` fallback-to-order behavior and
    the "No"/`None` defaults for the boolean/date columns."""
    order = make_shipment_order(pi_number="E0000901", model="order model", model_year=2019)
    unit = make_moto_unit(
        shipment_order=order,
        location=None,
        observation=None,
        model=None,
        model_year=None,
    )

    fake_db = FakeAsyncSession(execute_queue=[[unit]])

    with make_test_client(current_user=make_imports_editor(role="superadmin"), fake_db_session=fake_db) as client:
        response = client.get("/api/v1/imports/moto-units/export")

    assert response.status_code == 200
    rows = _read_rows(response.content)
    row = rows[1]
    assert row[0] == "E0000901"  # pi_number from order
    assert row[1] == "ORDER MODEL"  # fallback to order.model, upper-cased
    assert row[2] == 2019  # fallback to order.model_year
    assert row[8] is None  # UBICACIÓN
    assert row[9] == "No"  # SEP. NACIONALIZACIÓN
    assert row[10] is None  # OBSERVACIÓN
    assert row[11] == "No"  # EMPADRONADO
    assert row[12] is None  # FECHA EMPADRONAMIENTO
    assert row[13] == "No"  # EMPADR. FÍSICO ENVIADO
    assert row[14] is None  # FECHA EMPADR. FÍSICO
    assert row[15] is None  # DISTRIBUIDOR
    assert row[16] == "No"  # FACTURADO
    assert row[17] == "No"  # CARGADO RUNT
    assert row[18] == "No"  # DIM CARGADO


def test_source_pi_used_when_no_related_order():
    """Defensive branch (`o = u.shipment_order`, guarded by `if o else ...`
    throughout) — locked in even though the endpoint always joins a real
    order today."""
    unit = make_moto_unit(model="raw model", source_pi="E-STANDALONE")
    unit.shipment_order = None

    fake_db = FakeAsyncSession(execute_queue=[[unit]])

    with make_test_client(current_user=make_imports_editor(role="superadmin"), fake_db_session=fake_db) as client:
        response = client.get("/api/v1/imports/moto-units/export")

    row = _read_rows(response.content)[1]
    assert row[0] == "E-STANDALONE"  # source_pi fallback, no order
    assert row[1] == "RAW MODEL"  # unit.model wins, no order to fall back to
    assert row[2] == unit.model_year


def test_no_units_returns_headers_only():
    fake_db = FakeAsyncSession(execute_queue=[[]])

    with make_test_client(current_user=make_imports_editor(role="superadmin"), fake_db_session=fake_db) as client:
        response = client.get("/api/v1/imports/moto-units/export")

    assert response.status_code == 200
    rows = _read_rows(response.content)
    assert rows == [EXPECTED_HEADERS]


def test_no_filters_only_is_spare_part_clause_applied():
    fake_db = FakeAsyncSession(execute_queue=[[]])

    with make_test_client(current_user=make_imports_editor(role="superadmin"), fake_db_session=fake_db) as client:
        response = client.get("/api/v1/imports/moto-units/export")

    assert response.status_code == 200
    assert len(fake_db.executed_statements) == 1
    stmt = fake_db.executed_statements[0]
    sql = str(stmt.compile())
    # Booleans compile to literal `= false`/`= true` (no bind param) with the
    # default (non-dialect) compiler — a single WHERE clause has no `AND`.
    assert "is_spare_part = false" in sql
    assert " AND " not in sql.split("WHERE", 1)[1]


def test_query_param_filters_applied_to_the_single_query():
    """All 9 optional query params (`pi_number`/`model`/`vin`/`engine` ILIKE
    plus the 5 status params) end up as bound params on the single
    `select(ShipmentMotoUnit)` statement — locks in that the export endpoint
    uses the FULL `filters` set (not the narrower `base_filters`)."""
    obs_id = uuid.uuid4()
    fake_db = FakeAsyncSession(execute_queue=[[]])

    with make_test_client(current_user=make_imports_editor(role="superadmin"), fake_db_session=fake_db) as client:
        response = client.get(
            "/api/v1/imports/moto-units/export",
            params={
                "pi_number": "E0001",
                "model": "cb500",
                "vin": "VIN1",
                "engine": "ENG1",
                "certificado_generado": "true",
                "observation_id": str(obs_id),
                "empadronamiento_fisico_enviado": "true",
                "facturado": "true",
                "cargado_runt": "true",
            },
        )

    assert response.status_code == 200
    assert len(fake_db.executed_statements) == 1
    stmt = fake_db.executed_statements[0]
    sql = str(stmt.compile())
    params = stmt.compile().params.values()

    # `.ilike()` compiles (without a postgres dialect) to `lower(x) LIKE lower(y)`,
    # not a literal `ILIKE` keyword — assert on the portable `LIKE` form.
    assert "like" in sql.lower()
    assert "coalesce" in sql.lower()
    assert "%E0001%" in params
    assert "%cb500%" in params
    assert "%VIN1%" in params
    assert "%ENG1%" in params
    assert obs_id in params
    # Booleans (`certificado_generado`/`empadronamiento_fisico_enviado`/
    # `facturado`/`cargado_runt`) compile to literal `= true` (no bind param)
    # — locks in all 10 filters (base `is_spare_part` + 4 text ILIKE + 5
    # optional status params) landing on the SAME statement, joined by AND.
    where_sql = sql.split("WHERE", 1)[1]
    assert where_sql.count(" AND ") == 9
    assert where_sql.count("= true") == 4


def test_distribuidor_id_filter_applied_to_the_single_query():
    """`distribuidor_id` query param narrows the export's single query the
    same way it narrows the list endpoint — Excel export must respect the
    "filter by distributor" dropdown, not just the paginated list."""
    distribuidor_id = uuid.uuid4()
    fake_db = FakeAsyncSession(execute_queue=[[]])

    with make_test_client(current_user=make_imports_editor(role="superadmin"), fake_db_session=fake_db) as client:
        response = client.get(
            "/api/v1/imports/moto-units/export",
            params={"distribuidor_id": str(distribuidor_id)},
        )

    assert response.status_code == 200
    stmt = fake_db.executed_statements[0]
    sql = str(stmt.compile())
    params = stmt.compile().params.values()

    assert "empadronamiento_fisico_distribuidor_id" in sql
    assert distribuidor_id in params
