"""
Regression tests for `create_sp_order_from_excel` (Tier 4 of the
imports-oversized-functions-refactor change).

Purpose: pin TODAY's observable behavior of `create_sp_order_from_excel`
(`backend/app/services/imports_service.py`, lines 826-1047) — multi-sheet
header detection, model validation, order/lot upsert with its flush
sequencing, the G6 confirmed-lot guard, the delete-vs-preserve N+1
backorder-existence check, the two-pass aggregate-then-upsert flow, and the
pricing fan-out — BEFORE it gets decomposed into 7 helpers + a thin
orchestrator (per `sdd/imports-oversized-functions-refactor-tier4/design`),
so the exact same assertions can be re-run AFTER that refactor to prove it
is byte-for-byte behavior-preserving.

This is RED-only approval testing (per `strict-tdd.md`'s Approval Testing
section): these tests currently PASS against the UNrefactored function —
they describe current reality, not new behavior — and are re-run unedited
after each extraction PR to prove nothing moved.

Every assertion below pins CURRENT behavior exactly as specified in
`sdd/imports-oversized-functions-refactor-tier4/spec`, including 3 known
quirks that are locked verbatim, not fixed (see Observed Behavior Notes in
the spec and the design's Q1/Q2/Q3):
  - Q1: the old-item delete-vs-preserve loop issues one `select(Backorder)`
    PER old item (N+1), never batched.
  - Q2: `lot.updated_at = datetime.utcnow() if hasattr(lot, 'updated_at')
    else None` — the `else None` branch is dead code today (SparePartLot
    always has the attribute), preserved as-is.
  - Q3: aggregation's `unit_price` keeps the FIRST non-null value seen
    across duplicate (part_number, modelo) rows; a later differing price is
    silently ignored.

These tests exercise the REAL `create_sp_order_from_excel` function against
a `FakeAsyncSession` (see tests/imports/conftest.py) and real .xlsx bytes
built with openpyxl — no live database, no HTTP server, matching this
repo's established test pattern (see test_reconcile_lot_packing_list_regression.py
and test_confirm_reconciliation_regression.py, this module's Tier 3
siblings, and test_new_order_sp_guard_regression.py, which already pins the
G6 guard's HTTP-boundary behavior for this same function).
"""
import io
import uuid
from unittest.mock import AsyncMock, patch

import openpyxl
import pytest
from fastapi import HTTPException

from app.services.imports_service import create_sp_order_from_excel
from app.models.imports import ShipmentOrder, SparePartLot, SparePartItem
from tests.imports.conftest import (
    build_sp_order_xlsx,
    build_malformed_xlsx,
    make_lot,
    make_spare_part_item,
    make_shipment_order,
    make_actor,
    FakeAsyncSession,
)


# ---------------------------------------------------------------------------
# Phase 1 — Structural scenarios (order/lot upsert, validation, header detect)
# ---------------------------------------------------------------------------


class TestNewOrderAndNewLot:

    async def test_new_order_and_lot_created_with_flushes_and_response_shape(self):
        """Scenario 1 [U, R]: brand-new reference creates both order and lot,
        each `db.add`ed and flushed immediately (2 upsert flushes), plus the
        old-items-loop flush (empty loop, still flushes) and the final
        post-insert flush — 4 total. Response dict has exactly the 7
        documented keys."""
        fake_db = FakeAsyncSession(execute_queue=[
            [],  # _load_models_map
            [],  # existing_order lookup -> not found
            [],  # G6 check on the freshly created lot -> never confirmed
            [],  # old_items load on the freshly created lot -> empty
        ])
        file_bytes = build_sp_order_xlsx([
            {"part_number": "ABC-001", "nombre": "Brake pad", "qty": 5},
        ])

        result = await create_sp_order_from_excel(
            fake_db, "e0001234-sp", file_bytes, make_actor()
        )

        order_added = fake_db.added_of_type(ShipmentOrder)[0]
        lot_added = fake_db.added_of_type(SparePartLot)[0]
        item_added = fake_db.added_of_type(SparePartItem)[0]

        assert result == {
            "inserted": 1,
            "updated": 0,
            "skipped": 0,
            "errors": [],
            "order_id": str(order_added.id),
            "lot_id": str(lot_added.id),
            "reference": "E0001234-SP",
        }
        assert fake_db.flush_count == 4
        assert order_added.pi_number == "E0001234-SP"
        assert order_added.model == "REPUESTOS"
        assert order_added.is_spare_part is True
        assert order_added.computed_status == "en_preparacion"
        assert lot_added.shipment_order_id == order_added.id
        assert item_added.qty_ordered == 5
        assert item_added.status == "PENDING"
        assert item_added.qty_received == 0
        assert item_added.qty_pending == 5

    async def test_first_matching_sheet_wins_later_sheets_never_inspected(self):
        """Scenario (H): iterates `wb.sheetnames` in order and uses the
        FIRST sheet where `_find_header_row` matches — a later, header-less
        sheet is never inspected (if it were, `_find_header_row` would
        return None for it, but that never gets a chance to matter here
        since sheet #2 already won)."""
        wb = openpyxl.Workbook()
        sheet1 = wb.active
        sheet1.title = "Notas"
        sheet1.append(["Foo", "Bar"])
        sheet2 = wb.create_sheet("Orden")
        sheet2.append(["Codigo Parte", "Nombre", "Cantidad", "Moto Aplica"])
        sheet2.append(["XYZ-100", "Filtro", 2, None])
        wb.create_sheet("Otra")  # header-less; must never be reached
        buf = io.BytesIO()
        wb.save(buf)
        file_bytes = buf.getvalue()

        fake_db = FakeAsyncSession(execute_queue=[
            [],  # _load_models_map
            [],  # existing_order lookup
            [],  # G6 check
            [],  # old_items
        ])

        result = await create_sp_order_from_excel(
            fake_db, "E0002000-SP", file_bytes, make_actor()
        )

        assert result["inserted"] == 1
        item = fake_db.added_of_type(SparePartItem)[0]
        assert item.part_number == "XYZ-100"


class TestExistingOrderAndExistingLot:

    async def test_existing_order_and_lot_reused_no_duplicate_creation(self):
        """Scenario 2 [U]: both order and lot already exist for the
        reference -> reused as-is, zero upsert flushes (only the old-items
        loop flush + the final post-insert flush = 2)."""
        order = make_shipment_order(pi_number="E0000573-SP", is_spare_part=True)
        lot = make_lot(lot_identifier="E0000573-SP", shipment_order_id=order.id)
        fake_db = FakeAsyncSession(execute_queue=[
            [],       # _load_models_map
            [order],  # existing_order lookup -> found
            [lot],    # lot lookup -> found
            [],       # G6 check -> not confirmed
            [],       # old_items load -> empty
        ])
        file_bytes = build_sp_order_xlsx([
            {"part_number": "ABC-001", "nombre": "x", "qty": 3},
        ])

        result = await create_sp_order_from_excel(
            fake_db, "e0000573-sp", file_bytes, make_actor()
        )

        assert result["order_id"] == str(order.id)
        assert result["lot_id"] == str(lot.id)
        assert fake_db.added_of_type(ShipmentOrder) == []
        assert fake_db.added_of_type(SparePartLot) == []
        assert fake_db.flush_count == 2


class TestExistingOrderNoLot:

    async def test_existing_order_without_lot_creates_and_flushes_lot(self):
        """Scenario [U]: an order exists for the reference but has no
        SparePartLot yet -> a new lot is created and flushed; the order is
        NOT re-created."""
        order = make_shipment_order(pi_number="E0000900-SP", is_spare_part=True)
        fake_db = FakeAsyncSession(execute_queue=[
            [],       # _load_models_map
            [order],  # existing_order lookup -> found
            [],       # lot lookup -> not found
            [],       # G6 check -> not confirmed (fresh lot)
            [],       # old_items load -> empty
        ])
        file_bytes = build_sp_order_xlsx([])  # zero data rows

        result = await create_sp_order_from_excel(
            fake_db, "E0000900-SP", file_bytes, make_actor()
        )

        lot_added = fake_db.added_of_type(SparePartLot)[0]
        assert lot_added.shipment_order_id == order.id
        assert fake_db.added_of_type(ShipmentOrder) == []
        # 1 flush for the new lot + old-items-loop flush + final flush = 3.
        assert fake_db.flush_count == 3
        assert result["order_id"] == str(order.id)
        assert result["lot_id"] == str(lot_added.id)


class TestModelValidationFailure:

    async def test_unknown_model_raises_before_any_order_lookup(self):
        """Scenario 3 [M]: an unrecognized "Moto Aplica" value raises
        ValueError from `_collect_unknown_models` BEFORE the order/lot
        lookup -> only ONE execute() call ever happens (`_load_models_map`);
        anything more would trip `FakeAsyncSession`'s over-call guard."""
        fake_db = FakeAsyncSession(execute_queue=[
            [],  # _load_models_map — the only call that must happen
        ])
        file_bytes = build_sp_order_xlsx([
            {"part_number": "ABC-001", "nombre": "x", "qty": 3, "modelo": "GHOST9000"},
        ])

        with pytest.raises(ValueError, match="Modelos no reconocidos"):
            await create_sp_order_from_excel(
                fake_db, "E0001000-SP", file_bytes, make_actor()
            )

        assert fake_db.added == []
        assert fake_db.deleted == []
        assert fake_db.flush_count == 0


class TestSheetNotFound:

    async def test_no_matching_sheet_raises_before_any_db_access(self):
        """Scenario 4 [H]: no sheet has >= 3 of `SP_ORDER_COLS` -> ValueError
        (missing-columns message) raised with ZERO execute() calls."""
        fake_db = FakeAsyncSession(execute_queue=[])  # any execute() is a bug
        file_bytes = build_malformed_xlsx()

        with pytest.raises(ValueError, match="columnas requeridas"):
            await create_sp_order_from_excel(
                fake_db, "E0001000-SP", file_bytes, make_actor()
            )

        assert fake_db.added == []
        assert fake_db.flush_count == 0


# ---------------------------------------------------------------------------
# Phase 2 — Mutation scenarios (G6 guard, delete-vs-preserve N+1, two-pass
# aggregation, skipped/errors, pricing fan-out)
# ---------------------------------------------------------------------------


class TestG6ConfirmedLotGuard:

    async def test_confirmed_lot_rejects_reimport_before_item_mutation(self):
        """Scenario 5 [G]: `lot_has_confirmed_reconciliation` True ->
        HTTPException 409/LOT_ALREADY_CONFIRMED raised strictly BEFORE any
        old-item read/delete/update — no SparePartItem is touched."""
        order = make_shipment_order(pi_number="E0000999-SP", is_spare_part=True)
        lot = make_lot(lot_identifier="E0000999-SP", shipment_order_id=order.id)
        fake_db = FakeAsyncSession(execute_queue=[
            [],              # _load_models_map
            [order],         # existing_order lookup -> found
            [lot],           # lot lookup -> found
            [uuid.uuid4()],  # G6 check -> confirmed (truthy row)
        ])
        file_bytes = build_sp_order_xlsx([
            {"part_number": "ABC-001", "nombre": "x", "qty": 3},
        ])

        with pytest.raises(HTTPException) as exc_info:
            await create_sp_order_from_excel(
                fake_db, "e0000999-sp", file_bytes, make_actor()
            )

        assert exc_info.value.status_code == 409
        assert exc_info.value.detail["code"] == "LOT_ALREADY_CONFIRMED"
        assert fake_db.deleted == []
        assert fake_db.added_of_type(SparePartItem) == []


class TestDeleteVsPreserveWithN1Lock:

    async def test_item_with_backorder_preserved_item_without_deleted(self):
        """Scenario 6 [D]: old item WITH a Backorder is preserved and
        updated in-place (FK intact, never `db.delete`d); old item WITHOUT
        one is deleted. Locks the N+1 invariant: exactly one
        `select(Backorder.id)` PER old item, in iteration order — the fake
        session's queue is drained to exactly zero, proving no batching."""
        order = make_shipment_order(pi_number="E0000700-SP", is_spare_part=True)
        lot = make_lot(lot_identifier="E0000700-SP", shipment_order_id=order.id)
        item_with_bo = make_spare_part_item(lot.id, "ABC-001", qty_ordered=5)
        item_without_bo = make_spare_part_item(lot.id, "XYZ-002", qty_ordered=2)
        fake_db = FakeAsyncSession(execute_queue=[
            [],                                # _load_models_map
            [order],                            # existing_order lookup
            [lot],                              # lot lookup
            [],                                 # G6 check -> not confirmed
            [item_with_bo, item_without_bo],    # old_items load
            [uuid.uuid4()],                      # Backorder check: item_with_bo -> has one
            [],                                  # Backorder check: item_without_bo -> none
        ])
        file_bytes = build_sp_order_xlsx([
            {"part_number": "ABC-001", "nombre": "updated name", "qty": 9},
        ])

        result = await create_sp_order_from_excel(
            fake_db, "e0000700-sp", file_bytes, make_actor()
        )

        assert item_with_bo.qty_ordered == 9  # updated in-place
        assert item_with_bo not in fake_db.deleted
        assert item_without_bo in fake_db.deleted
        assert fake_db._execute_queue == []  # exactly 2 backorder checks, no more/less
        assert result["updated"] == 1
        assert result["inserted"] == 0


class TestTwoPassAggregation:

    async def test_sums_qty_backfills_unit_price_leaves_stale_item_untouched(self):
        """Scenario 7 [A], bundling the 3 aggregation scenarios:
          - same part_number+modelo across 2 rows sums qty (3 + 5 = 8)
          - unit_price keeps the FIRST non-null seen (10, not the second
            row's 20) — Q3, pinned not fixed
          - an old item preserved via backorder but whose key has no
            matching row in the new file is left completely untouched:
            neither updated, deleted, nor reported in errors/counters."""
        order = make_shipment_order(pi_number="E0000800-SP", is_spare_part=True)
        lot = make_lot(lot_identifier="E0000800-SP", shipment_order_id=order.id)
        stale_item = make_spare_part_item(lot.id, "STALE-1", qty_ordered=1)
        fake_db = FakeAsyncSession(execute_queue=[
            [],                # _load_models_map
            [order],           # existing_order lookup
            [lot],             # lot lookup
            [],                # G6 check
            [stale_item],      # old_items load
            [uuid.uuid4()],    # Backorder check: stale_item -> has one, preserved
        ])
        file_bytes = build_sp_order_xlsx([
            {"part_number": "ABC-001", "nombre": "x", "qty": 3, "unit_price": 10},
            {"part_number": "ABC-001", "nombre": "x2", "qty": 5, "unit_price": 20},
        ])

        with patch(
            "app.services.pricing_service.recalculate_part_cost",
            new=AsyncMock(),
        ):
            result = await create_sp_order_from_excel(
                fake_db, "e0000800-sp", file_bytes, make_actor()
            )

        item = fake_db.added_of_type(SparePartItem)[0]
        assert item.qty_ordered == 8  # 3 + 5 summed
        assert item.unit_price == 10  # first non-null kept; 20 ignored (Q3)
        assert stale_item.qty_ordered == 1  # untouched — no key match in new file
        assert stale_item not in fake_db.deleted
        assert result["errors"] == []
        assert result["skipped"] == 0
        assert result["inserted"] == 1


class TestSkippedErrorsAndDetailLoadedUnchanged:

    async def test_zero_valid_rows_leaves_detail_loaded_unchanged_and_shape_holds(self):
        """Scenario 8 [F, R]: every row skipped (qty<=0 or missing) ->
        skipped counter increments, `lot.detail_loaded` is NOT modified
        (stays False, never reset), and the response shape holds even for a
        zero-effect run — exactly the 7 documented keys."""
        order = make_shipment_order(pi_number="E0000600-SP", is_spare_part=True)
        lot = make_lot(lot_identifier="E0000600-SP", shipment_order_id=order.id, detail_loaded=False)
        fake_db = FakeAsyncSession(execute_queue=[
            [],       # _load_models_map
            [order],  # existing_order lookup
            [lot],    # lot lookup
            [],       # G6 check
            [],       # old_items load
        ])
        file_bytes = build_sp_order_xlsx([
            {"part_number": "ABC-001", "nombre": "x", "qty": 0},     # qty<=0 -> skipped
            {"part_number": "ABC-002", "nombre": "x", "qty": None},  # missing qty -> skipped
        ])

        result = await create_sp_order_from_excel(
            fake_db, "e0000600-sp", file_bytes, make_actor()
        )

        assert lot.detail_loaded is False
        assert result["skipped"] == 2
        assert result["inserted"] == 0
        assert result["updated"] == 0
        assert result["errors"] == []
        assert set(result.keys()) == {
            "inserted", "updated", "skipped", "errors", "order_id", "lot_id", "reference",
        }

    async def test_row_parse_exception_lands_in_errors_not_double_counted_as_skipped(self):
        """Per-row exceptions are caught and appended to `errors` (row NOT
        also counted as skipped); other rows in the same file still
        process normally."""
        order = make_shipment_order(pi_number="E0000400-SP", is_spare_part=True)
        lot = make_lot(lot_identifier="E0000400-SP", shipment_order_id=order.id)
        fake_db = FakeAsyncSession(execute_queue=[
            [],       # _load_models_map
            [order],  # existing_order lookup
            [lot],    # lot lookup
            [],       # G6 check
            [],       # old_items load
        ])
        file_bytes = build_sp_order_xlsx([
            {"part_number": "BOOM", "nombre": "x", "qty": 3},
            {"part_number": "OK-001", "nombre": "y", "qty": 2},
        ])

        from app.services import imports_service as svc
        original_normalize = svc.normalize_part_number

        def _boom(pn):
            if str(pn) == "BOOM":
                raise ValueError("simulated parse failure")
            return original_normalize(pn)

        with patch("app.services.imports_service.normalize_part_number", side_effect=_boom):
            result = await create_sp_order_from_excel(
                fake_db, "e0000400-sp", file_bytes, make_actor()
            )

        assert result["errors"] == [{"row": 2, "reason": "simulated parse failure"}]
        assert result["skipped"] == 0
        assert result["inserted"] == 1  # OK-001 still processed normally
        item = fake_db.added_of_type(SparePartItem)[0]
        assert item.part_number == "OK-001"


class TestPricingFanOutOrdering:

    async def test_recalc_fires_once_per_part_after_flush_even_when_price_unchanged(self):
        """Scenario 9 [F]: `recalculate_part_cost` runs once per distinct
        part_number in `parts_with_price` — a part lands there on ANY
        `unit_price is not None` write, regardless of whether the value
        actually changed. Also locks that the orchestrator's final
        `await db.flush()` (both flushes already done: old-items-loop +
        post-upsert) happens strictly BEFORE the recalc call."""
        order = make_shipment_order(pi_number="E0000500-SP", is_spare_part=True)
        lot = make_lot(lot_identifier="E0000500-SP", shipment_order_id=order.id)
        existing_item = make_spare_part_item(
            lot.id, "ABC-001", qty_ordered=5, unit_price=10.0
        )
        fake_db = FakeAsyncSession(execute_queue=[
            [],                # _load_models_map
            [order],           # existing_order lookup
            [lot],             # lot lookup
            [],                # G6 check
            [existing_item],   # old_items load
            [uuid.uuid4()],    # Backorder check -> has one, preserved
        ])
        file_bytes = build_sp_order_xlsx([
            {"part_number": "ABC-001", "nombre": "x", "qty": 5, "unit_price": 10.0},
        ])

        recorded = {}

        async def _fake_recalc(db, pn, lot_identifier=None):
            recorded["flush_count_at_call"] = db.flush_count
            recorded["part_number"] = pn
            recorded["lot_identifier"] = lot_identifier

        with patch(
            "app.services.pricing_service.recalculate_part_cost",
            new=AsyncMock(side_effect=_fake_recalc),
        ) as mock_recalc:
            result = await create_sp_order_from_excel(
                fake_db, "e0000500-sp", file_bytes, make_actor()
            )

        mock_recalc.assert_awaited_once()
        assert existing_item.unit_price == 10.0  # unchanged value, still written
        assert result["updated"] == 1
        assert recorded["part_number"] == "ABC-001"
        assert recorded["lot_identifier"] == "E0000500-SP"
        # Both flushes (old-items loop + final post-upsert) must have
        # already happened by the time recalc runs.
        assert recorded["flush_count_at_call"] == 2
