"""
Shared fixtures/helpers for `imports_service` packing-list reconciliation tests.

Follows the project's established pattern (see tests/remisiones/conftest.py,
tests/test_parts_manual_catalog.py): import the REAL production code and
exercise it against lightweight fakes for the DB session — no live database,
no HTTP server. Excel fixtures are built with real `openpyxl` (already a hard
dependency of the app) so the parser is exercised against real, valid .xlsx
bytes instead of hand-rolled approximations.
"""
import io
import uuid
from datetime import datetime, timezone
from typing import Optional

import openpyxl


# ---------------------------------------------------------------------------
# Excel fixture builders
# ---------------------------------------------------------------------------

def build_packing_list_xlsx(
    rows: list[dict],
    sheet_name: str = "PACKING LIST",
) -> bytes:
    """
    Builds a minimal, valid Packing List (non-invoice) .xlsx: headers
    `Part #`, `Complete Description`, `Qty(PCS)`, `N.W`, `G.W`, `CBM` on row 1,
    one data row per dict in `rows` (keys: part_number, description, qty,
    nw, gw, cbm — all optional except part_number/qty).

    NOTE: `CBM` is included so `_detect_sheet_type` classifies the sheet as
    `sp_packing_list` (it requires has_part AND has_weight AND has_meas AND
    NOT has_price) — without it, multi-sheet lot-identifier preference in
    `_locate_packing_list_sheet` never engages and silently falls back to
    `wb.active` (see test_prefers_sheet_matching_lot_identifier).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(["Part #", "Complete Description", "Qty(PCS)", "N.W", "G.W", "CBM"])
    for r in rows:
        ws.append([
            r.get("part_number"),
            r.get("description"),
            r.get("qty"),
            r.get("nw", 1.0),
            r.get("gw", 1.2),
            r.get("cbm", 0.01),
        ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_invoice_xlsx(
    rows: list[dict],
    sheet_name: str = "INVOICE",
) -> bytes:
    """
    Builds a minimal, valid Invoice .xlsx: headers `Part #`, `Model`,
    `Complete Description`, `Spanish Description`, `Qty(PCS)`, `Unit Price`,
    `Amount` on row 1, one data row per dict in `rows` (keys: part_number,
    model, description, description_es, qty, unit_price, amount).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append([
        "Part #", "Model", "Complete Description", "Spanish Description",
        "Qty(PCS)", "Unit Price", "Amount",
    ])
    for r in rows:
        ws.append([
            r.get("part_number"),
            r.get("model"),
            r.get("description"),
            r.get("description_es"),
            r.get("qty"),
            r.get("unit_price"),
            r.get("amount"),
        ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_sp_order_xlsx(
    rows: list[dict],
    sheet_name: str = "ORDEN",
) -> bytes:
    """
    Builds a minimal, valid SP-order-request .xlsx (see `SP_ORDER_COLS` in
    `imports_service.py`, consumed by `create_sp_order_from_excel`): headers
    `Codigo Parte`, `Nombre`, `Cantidad`, `Moto Aplica`, `Unit Price` on row
    1, one data row per dict in `rows` (keys: part_number, nombre, qty,
    modelo, unit_price — all optional).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(["Codigo Parte", "Nombre", "Cantidad", "Moto Aplica", "Unit Price"])
    for r in rows:
        ws.append([
            r.get("part_number"),
            r.get("nombre"),
            r.get("qty"),
            r.get("modelo"),
            r.get("unit_price"),
        ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_malformed_xlsx() -> bytes:
    """An .xlsx with no recognizable Packing List/Invoice headers at all."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Foo", "Bar", "Baz"])
    ws.append(["a", "b", "c"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Lightweight ORM-instance factories (real model classes, not mocks — these
# are plain SQLAlchemy declarative instances, safe to construct without a DB)
# ---------------------------------------------------------------------------

def make_lot(
    lot_identifier: str = "E0000573-SP",
    detail_loaded: bool = False,
    packing_list_received: bool = False,
    created_at: Optional[datetime] = None,
    total_declared_value: Optional[float] = None,
    currency: Optional[str] = "USD",
    shipment_order_id=None,
) -> "SparePartLot":
    from app.models.imports import SparePartLot
    lot = SparePartLot(
        id=uuid.uuid4(),
        shipment_order_id=shipment_order_id or uuid.uuid4(),
        lot_identifier=lot_identifier,
        detail_loaded=detail_loaded,
        packing_list_received=packing_list_received,
        created_at=created_at or datetime.utcnow(),
        total_declared_value=total_declared_value,
        currency=currency,
    )
    return lot


def make_spare_part_item(
    lot_id,
    part_number: str,
    qty_ordered: int,
    model_applicable: Optional[str] = None,
    qty_received: int = 0,
    status: str = "PENDING",
    unit_price: Optional[float] = None,
    fob_pi: Optional[float] = None,
    description_es: Optional[str] = None,
    created_at: Optional[datetime] = None,
) -> "SparePartItem":
    # `created_at` has a Python-level `default=datetime.utcnow` on the model
    # column, but that default only applies at INSERT/flush time, not at
    # plain construction -- these tests never flush against a real DB, so
    # `SparePartItemRead.created_at` (required, non-optional) would be
    # `None` unless we set it explicitly here.
    from app.models.imports import SparePartItem
    item = SparePartItem(
        id=uuid.uuid4(),
        lot_id=lot_id,
        part_number=part_number,
        qty_ordered=qty_ordered,
        model_applicable=model_applicable,
        qty_received=qty_received,
        status=status,
        unit_price=unit_price,
        fob_pi=fob_pi,
        description_es=description_es,
        created_at=created_at or datetime.utcnow(),
    )
    return item


def make_reconciliation_result(
    lot_id,
    part_number: str,
    result: str = "EXTRA",
    spare_part_item_id=None,
    packing_list_id=None,
    description_es: Optional[str] = None,
    model_applicable: Optional[str] = None,
    qty_ordered: Optional[int] = None,
    qty_in_packing: Optional[int] = None,
) -> "ReconciliationResult":
    from app.models.imports import ReconciliationResult
    return ReconciliationResult(
        id=uuid.uuid4(),
        lot_id=lot_id,
        packing_list_id=packing_list_id or uuid.uuid4(),
        spare_part_item_id=spare_part_item_id,
        part_number=part_number,
        description_es=description_es,
        model_applicable=model_applicable,
        qty_ordered=qty_ordered,
        qty_in_packing=qty_in_packing,
        result=result,
    )


def make_backorder(
    spare_part_item_id,
    part_number: str,
    origin_pi: str,
    qty_pending: int,
    resolved: bool = False,
    created_at: Optional[datetime] = None,
    source: str = "reconciliation",
    already_charged: bool = False,
) -> "Backorder":
    # `created_at`/`source`/`already_charged` all have Python-level defaults
    # on the model column or the `BackorderRead` schema, but neither applies
    # at plain construction time (no flush against a real DB in these
    # tests) -- set explicitly so `BackorderRead` (which requires
    # `created_at: datetime`) validates.
    from app.models.imports import Backorder
    return Backorder(
        id=uuid.uuid4(),
        spare_part_item_id=spare_part_item_id,
        part_number=part_number,
        origin_pi=origin_pi,
        qty_pending=qty_pending,
        resolved=resolved,
        history=[],
        created_at=created_at or datetime.utcnow(),
        source=source,
        already_charged=already_charged,
    )


def make_backorder_reconciliation(
    lot_id,
    file_name: str = "remainder.xlsx",
    content_hash: str = "hash",
    status: str = "PENDING",
    is_invoice: bool = False,
) -> "BackorderReconciliation":
    from app.models.imports import BackorderReconciliation
    return BackorderReconciliation(
        id=uuid.uuid4(),
        lot_id=lot_id,
        file_name=file_name,
        content_hash=content_hash,
        minio_object_name=f"minio/{file_name}",
        status=status,
        is_invoice=is_invoice,
    )


def make_backorder_reconciliation_result(
    reconciliation_id,
    part_number: str,
    result: str,
    backorder_id=None,
    spare_part_item_id=None,
    qty_pending_snapshot: Optional[int] = None,
    qty_in_packing: Optional[int] = None,
    qty_applied: Optional[int] = None,
    unit_price: Optional[float] = None,
    model_applicable: Optional[str] = None,
) -> "BackorderReconciliationResult":
    from app.models.imports import BackorderReconciliationResult
    return BackorderReconciliationResult(
        id=uuid.uuid4(),
        reconciliation_id=reconciliation_id,
        backorder_id=backorder_id,
        spare_part_item_id=spare_part_item_id,
        part_number=part_number,
        model_applicable=model_applicable,
        qty_pending_snapshot=qty_pending_snapshot,
        qty_in_packing=qty_in_packing,
        qty_applied=qty_applied,
        unit_price=unit_price,
        result=result,
        created_at=datetime.utcnow(),
    )


def make_shipment_order(
    pi_number: str = "E0000573",
    model: Optional[str] = None,
    model_year: Optional[int] = None,
    is_spare_part: bool = False,
) -> "ShipmentOrder":
    from app.models.imports import ShipmentOrder
    return ShipmentOrder(
        id=uuid.uuid4(),
        pi_number=pi_number,
        model=model,
        model_year=model_year,
        is_spare_part=is_spare_part,
    )


def make_moto_location(name: str = "BODEGA CENTRAL") -> "MotoLocation":
    from app.models.imports import MotoLocation
    return MotoLocation(id=uuid.uuid4(), name=name)


def make_moto_observation(name: str = "PENDIENTE REVISION") -> "MotoObservation":
    from app.models.imports import MotoObservation
    return MotoObservation(id=uuid.uuid4(), name=name)


def make_moto_unit(
    shipment_order: Optional["ShipmentOrder"] = None,
    location: Optional["MotoLocation"] = None,
    observation: Optional["MotoObservation"] = None,
    vin_number: str = "VIN0001",
    engine_number: str = "ENG0001",
    model: Optional[str] = None,
    model_year: Optional[int] = None,
    created_at: Optional[datetime] = None,
    **kwargs,
) -> "ShipmentMotoUnit":
    """
    Real `ShipmentMotoUnit` ORM instance, unattached to any DB session.
    `shipment_order`/`location`/`observation` relationships are set directly
    (safe: `FakeAsyncSession` never triggers lazy-load, it hands back queued
    rows verbatim) so `_serialize_moto_unit`-style code can read
    `u.shipment_order`/`u.location`/`u.observation` without a real DB.
    Boolean columns default to `False` here (mirrors the DB column defaults,
    which do NOT apply to a plain unflushed Python instance).
    """
    from app.models.imports import ShipmentMotoUnit
    order = shipment_order if shipment_order is not None else make_shipment_order()
    defaults = dict(
        item_no=1,
        color=None,
        color_runt=None,
        container_no=None,
        seal_no=None,
        source_pi=None,
        no_acep=None,
        f_acep=None,
        no_lev=None,
        f_lev=None,
        certificado_generado=False,
        certificado_fecha=None,
        empadronamiento_fisico_enviado=False,
        empadronamiento_fisico_fecha=None,
        empadronamiento_fisico_distribuidor_id=None,
        empadronamiento_fisico_distribuidor_nombre=None,
        dim_pdf_object_name=None,
        separada_nacionalizacion=False,
        facturado=False,
        cargado_runt=False,
    )
    defaults.update(kwargs)
    unit = ShipmentMotoUnit(
        id=uuid.uuid4(),
        shipment_order_id=order.id,
        vin_number=vin_number,
        engine_number=engine_number,
        model=model,
        model_year=model_year,
        location_id=(location.id if location else None),
        observation_id=(observation.id if observation else None),
        created_at=created_at or datetime.utcnow(),
        **defaults,
    )
    unit.shipment_order = order
    unit.location = location
    unit.observation = observation
    return unit


def make_actor(user_id: Optional[str] = None) -> "CurrentUser":
    from app.api.deps import CurrentUser
    return CurrentUser(
        user_id=user_id or str(uuid.uuid4()),
        role="jefe_taller",
        tenant_id=None,
        name="Test User",
    )


def make_imports_editor(role: str = "administrativo", user_id: Optional[str] = None) -> "CurrentUser":
    """
    GOTCHA: every imports endpoint calls `_require_imports_editor`, which
    requires `role in ("superadmin", "proveedor", "administrativo")`.
    `make_actor()` defaults to role=`jefe_taller`, which is NOT an editor
    and 403s here (see `test_imports_editor_role_regression.py`). Use this
    factory instead of `make_actor()` for any imports-endpoint HTTP test.
    """
    from app.api.deps import CurrentUser
    return CurrentUser(
        user_id=user_id or str(uuid.uuid4()),
        role=role,
        tenant_id=None,
        name="Test Imports Editor",
    )


# ---------------------------------------------------------------------------
# Fake AsyncSession — simulates just enough of AsyncSession's surface for
# `reconcile_lot_packing_list` to run end-to-end without a live database.
# ---------------------------------------------------------------------------

class _ScalarsResult:
    def __init__(self, items: list):
        self._items = items

    def all(self):
        return list(self._items)

    def first(self):
        return self._items[0] if self._items else None


class _ExecuteResult:
    """
    Fakes the subset of SQLAlchemy's `CursorResult`/`ChunkedIteratorResult`
    surface this project's endpoints call on `await db.execute(stmt)`:
    `.scalars().all()` (ORM entities), `.all()` (multi-column `Row` tuples,
    e.g. `select(Model.col_a, Model.col_b)`), `.scalar_one()`/
    `.scalar_one_or_none()` (single-column, single-row results, e.g.
    `select(func.count())`), and `.first()` (raw `Row`-or-`None` existence
    checks called directly on the execute result without `.scalars()`
    first, e.g. `create_sp_order_from_excel`'s per-item Backorder-existence
    check), and `.rowcount` (targeted `UPDATE ... WHERE ...` statements,
    e.g. `import_rotation`/`import_description_es`'s "only fill if still
    NULL" pattern — queue an empty list to simulate 0 rows matched/updated,
    a 1-item list to simulate 1). All six read the SAME queued `items`
    payload — the caller queues the right shape (list of ORM objects/Rows,
    or a 1-item list for scalar_one/rowcount) for whichever call the
    endpoint makes.
    """

    def __init__(self, items: list):
        self._items = items

    @property
    def rowcount(self) -> int:
        return len(self._items)

    def scalars(self):
        return _ScalarsResult(self._items)

    def all(self):
        return list(self._items)

    def first(self):
        return self._items[0] if self._items else None

    def scalar_one(self):
        if len(self._items) != 1:
            raise AssertionError(
                f"scalar_one() expects exactly 1 queued row, got {len(self._items)} "
                "— check the execute_queue entry for this call."
            )
        return self._items[0]

    def scalar_one_or_none(self):
        if len(self._items) > 1:
            raise AssertionError(
                f"scalar_one_or_none() expects at most 1 queued row, got {len(self._items)}."
            )
        return self._items[0] if self._items else None


class FakeAsyncSession:
    """
    Minimal fake standing in for `AsyncSession`.

    `execute_queue` MUST be provided in the exact order `reconcile_lot_packing_list`
    issues its `select(...)` calls today:
      1. `select(VehicleModel.model_name)`       (via `_load_models_map`)
      2. `select(ReconciliationResult.id)...`    (G3 defense-in-depth
         confirmed-existence check, immediately before `_replace_lot_packing_list`)
      3. `select(ReconciliationResult)...`       (old_results for the lot,
         inside `_replace_lot_packing_list`)
      4. `select(PackingList)...`                (old_pls for the lot)
      5. `select(SparePartItem)...`              (lot_items_list)

    This positional coupling is intentional: it is exactly what task 2.1 asks
    for — a regression test that captures TODAY's behavior byte-for-byte, so
    the Phase 2 refactor (moving parsing into a pure helper) can be verified
    to not reorder or alter any DB interaction. G3 (`sdd/packing-list-
    reupload-requires-rollback`) added call #2; the 8 pre-existing pinned
    scenarios queue an empty list there (none involve an already-confirmed
    lot).
    """

    def __init__(self, execute_queue: list[list], get_objects: Optional[list] = None):
        self._execute_queue = list(execute_queue)
        self.added: list = []
        self.deleted: list = []
        self.flush_count = 0
        # Objects reachable via `db.get(Model, id)` — used by tests exercising
        # functions that fetch rows by primary key instead of `select(...)`
        # (e.g. `confirm_backorder_reconciliation`).
        self._get_objects: list = list(get_objects or [])
        # Every `stmt` passed to `execute()`, in call order — lets tests
        # inspect the compiled SQL/params of statements built by filter
        # helpers (e.g. `_apply_lot_filters`) without a live DB.
        self.executed_statements: list = []

    async def execute(self, stmt):
        if not self._execute_queue:
            raise AssertionError(
                "FakeAsyncSession.execute() called more times than expected — "
                "the query order/count changed. Update the test's execute_queue."
            )
        self.executed_statements.append(stmt)
        return _ExecuteResult(self._execute_queue.pop(0))

    def add(self, obj):
        if getattr(obj, "id", None) is None:
            obj.id = uuid.uuid4()
        self.added.append(obj)

    async def delete(self, obj):
        self.deleted.append(obj)

    async def flush(self):
        self.flush_count += 1

    async def commit(self):
        """No-op: write endpoints call `await db.commit()` after mutating
        ORM instances in place; since those instances are plain unflushed
        Python objects (no real session), there is nothing to persist."""
        pass

    async def refresh(self, obj, attribute_names=None):
        """No-op: write endpoints call `await db.refresh(obj)` or
        `await db.refresh(obj, [...])` (e.g. `update_moto_unit` refreshes
        `['location', 'observation']`) to reload relationships/defaults
        from the DB after commit. Since `obj` was never detached from a
        real session, its current in-memory state (set directly by the
        test factories in this module) already reflects what a refresh
        would reload — no-op is correct."""
        pass

    async def get(self, model_cls, obj_id):
        for obj in self._get_objects + self.added:
            if isinstance(obj, model_cls) and obj.id == obj_id:
                return obj
        return None

    def added_of_type(self, cls) -> list:
        return [o for o in self.added if isinstance(o, cls)]
