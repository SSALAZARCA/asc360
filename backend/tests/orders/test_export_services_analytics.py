"""
tests/orders/test_export_services_analytics.py -- RED-first test for
`POST /orders/analytics/services/export`, the "Gestión de Órdenes"
(`frontend/app/services/page.js`, `/services`) Excel-export button.

Owner requirements this endpoint must satisfy:
  1. The export respects whatever filters are active on screen (search
     text, tipo, estado, centro) -- but ALL of that filtering already
     happens client-side today (`services/page.js`'s `filtered` useMemo).
     Rather than re-implement it on the backend (risk of drift between two
     implementations), the frontend posts the exact order IDs it currently
     has on screen and the backend just re-fetches those rows -- so this
     endpoint accepts ONLY an explicit `order_ids` list, no search/tipo/
     estado/centro params of its own.
  2. Superadmin-only.

Exercised via `make_test_client` (real FastAPI DI/serialization, no live
DB) + `openpyxl` parsing of the returned bytes, mirroring
`tests/imports/test_export_moto_units_endpoint_regression.py`'s style for
this codebase's other export endpoint.

Row-shaping is delegated to `_fetch_services_data` (the helper extracted
from `get_services_analytics` so both endpoints share one query -- see
`app/api/v1/orders.py`), which issues up to two `db.execute()` calls:
  1. The joined `select(ServiceOrder, Vehicle.plate, Tenant.name,
     Tenant.ciudad, ServiceOrderReception.mileage_km, Vehicle.model)` --
     positional-tuple rows.
  2. ONLY when (1) returned at least one plate: the per-plate visit/
     warranty aggregate `select(...)` -- attribute-style rows (`.plate`,
     `.total_visits`, `.recent_visits`, `.warranty_count`).
An empty `order_ids` list makes (1) return nothing, so the `if plates:`
guard skips (2) entirely -- the empty-list test below queues only one
result.
"""
import io
import uuid
from datetime import datetime, timedelta

import openpyxl

from app.api.deps import CurrentUser
from app.models.order import ServiceOrder, ServiceStatus, ServiceType
from tests.conftest import make_test_client

EXPECTED_HEADERS = [
    "N.° Orden", "Placa", "Modelo", "Tipo", "Estado", "Días en Taller", "KM",
    "Centro", "Ciudad", "Visitas Totales", "Visitas 2 Meses", "Garantías Totales",
]


def make_user(role: str) -> CurrentUser:
    return CurrentUser(user_id=str(uuid.uuid4()), role=role, tenant_id=None, name="T")


def make_order(
    status: ServiceStatus = ServiceStatus.in_progress,
    service_type: ServiceType = ServiceType.regular,
    created_at: datetime = None,
    delivered_at: datetime = None,
) -> ServiceOrder:
    order = ServiceOrder(
        id=uuid.uuid4(),
        tenant_id=uuid.uuid4(),
        vehicle_id=uuid.uuid4(),
        status=status,
        service_type=service_type,
        created_at=created_at or datetime(2026, 1, 1),
        delivered_at=delivered_at,
    )
    order.reception = None
    return order


class _QueueResult:
    """Fakes the `.all()` surface `_fetch_services_data` reads off both of
    its `select(...)` calls -- a plain list of already-shaped rows."""

    def __init__(self, items):
        self._items = items

    def all(self):
        return self._items


class FakeExportSession:
    def __init__(self, queue):
        self._queue = list(queue)
        self.executed_statements: list = []

    async def execute(self, stmt):
        self.executed_statements.append(stmt)
        return _QueueResult(self._queue.pop(0))


class _AggRow:
    """Fakes one row of the per-plate visit/warranty aggregate query --
    attribute access only (`.plate`/`.total_visits`/`.recent_visits`/
    `.warranty_count`), matching how `_fetch_services_data` reads it."""

    def __init__(self, plate, total_visits, recent_visits, warranty_count):
        self.plate = plate
        self.total_visits = total_visits
        self.recent_visits = recent_visits
        self.warranty_count = warranty_count


def _read_rows(content: bytes) -> list:
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    return [list(row) for row in ws.iter_rows(values_only=True)]


def test_non_superadmin_gets_403():
    fake_db = FakeExportSession(queue=[])

    with make_test_client(current_user=make_user("jefe_taller"), fake_db_session=fake_db) as client:
        response = client.post("/api/v1/orders/analytics/services/export", json={"order_ids": []})

    assert response.status_code == 403
    assert fake_db.executed_statements == []


def test_distribuidor_also_gets_403():
    """Superadmin-only means EVERY other role is blocked, not just the ones
    `forbid_distribuidor` already covers elsewhere."""
    fake_db = FakeExportSession(queue=[])

    with make_test_client(current_user=make_user("parts_dealer"), fake_db_session=fake_db) as client:
        response = client.post("/api/v1/orders/analytics/services/export", json={"order_ids": []})

    assert response.status_code == 403
    assert fake_db.executed_statements == []


def test_administrativo_now_gets_200():
    """2026-08-24 business decision: administrativo now matches superadmin
    on this Excel export (previously blocked, see `test_non_superadmin_gets_403`
    above -- `jefe_taller` and `parts_dealer` stay blocked, unaffected by
    this change)."""
    fake_db = FakeExportSession(queue=[[]])

    with make_test_client(current_user=make_user("administrativo"), fake_db_session=fake_db) as client:
        response = client.post("/api/v1/orders/analytics/services/export", json={"order_ids": []})

    assert response.status_code == 200


def test_empty_id_list_returns_headers_only():
    fake_db = FakeExportSession(queue=[[]])

    with make_test_client(current_user=make_user("superadmin"), fake_db_session=fake_db) as client:
        response = client.post("/api/v1/orders/analytics/services/export", json={"order_ids": []})

    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    rows = _read_rows(response.content)
    assert rows == [EXPECTED_HEADERS]


def test_returns_only_the_requested_order_ids_rows():
    order = make_order(
        status=ServiceStatus.completed,
        service_type=ServiceType.warranty,
        created_at=datetime(2026, 1, 1),
        delivered_at=datetime(2026, 1, 6),
    )
    main_row = (order, "ABC12D", "Taller Centro", "Bogotá", 12000, "XTREET 200")
    agg_row = _AggRow(plate="ABC12D", total_visits=3, recent_visits=1, warranty_count=2)

    fake_db = FakeExportSession(queue=[[main_row], [agg_row]])

    with make_test_client(current_user=make_user("superadmin"), fake_db_session=fake_db) as client:
        response = client.post(
            "/api/v1/orders/analytics/services/export",
            json={"order_ids": [str(order.id)]},
        )

    assert response.status_code == 200
    assert response.headers["content-disposition"].startswith(
        'attachment; filename="ordenes_servicio_'
    )
    assert response.headers["content-disposition"].endswith('.xlsx"')

    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    assert wb.active.title == "Ordenes"

    rows = _read_rows(response.content)
    assert rows[0] == EXPECTED_HEADERS
    assert rows[1] == [
        str(order.id), "ABC12D", "XTREET 200", "Garantía", "Finalizado", 5, 12000,
        "Taller Centro", "Bogotá", 3, 1, 2,
    ]

    # Locks in that the query was narrowed by the posted IDs (not by any
    # search/tipo/estado/centro param -- this endpoint has none).
    stmt = fake_db.executed_statements[0]
    params = list(stmt.compile().params.values())
    assert any(order.id == p or (isinstance(p, list) and order.id in p) for p in params)


def test_administrativo_with_tenant_id_gets_network_wide_export():
    """2026-08-24 regression: an `administrativo` actor CAN carry a non-null
    `tenant_id` (see `users.py:110-112` -- non-superadmin-created users
    inherit the creating actor's tenant). The export must stay
    network-wide for this role, matching what `services/page.js`'s
    `(isSuperadmin || isAdministrativo)` gate promises in the UI -- NOT
    silently narrow to the actor's own tenant like a `jefe_taller` would.

    Unlike `make_user()` above (which hardcodes `tenant_id=None` for every
    role and is exactly why this bug slipped through review), this actor
    is built with an explicit non-null `tenant_id`.
    """
    actor_tenant_id = str(uuid.uuid4())

    other_tenant_order = make_order(
        status=ServiceStatus.in_progress,
        service_type=ServiceType.regular,
        created_at=datetime(2026, 1, 1),
    )
    other_tenant_order.tenant_id = uuid.uuid4()
    assert str(other_tenant_order.tenant_id) != actor_tenant_id

    main_row = (other_tenant_order, "ZZZ999", "Otro Centro", "Cali", 3000, "XPEED 125")
    agg_row = _AggRow(plate="ZZZ999", total_visits=1, recent_visits=1, warranty_count=0)
    fake_db = FakeExportSession(queue=[[main_row], [agg_row]])

    actor = CurrentUser(
        user_id=str(uuid.uuid4()), role="administrativo", tenant_id=actor_tenant_id, name="T"
    )

    with make_test_client(current_user=actor, fake_db_session=fake_db) as client:
        response = client.post(
            "/api/v1/orders/analytics/services/export",
            json={"order_ids": [str(other_tenant_order.id)]},
        )

    assert response.status_code == 200

    # The row belonging to a DIFFERENT tenant must appear -- proving the
    # export stayed network-wide instead of narrowing to the actor's own
    # tenant.
    rows = _read_rows(response.content)
    assert rows[1][0] == str(other_tenant_order.id)

    # Direct proof of the fix: the query must NOT have been narrowed by the
    # actor's own tenant_id. If it had been, `tenant_id` would show up as a
    # bound param on the base query, silently scoping every administrativo
    # export to their own tenant regardless of what order_ids were posted.
    stmt = fake_db.executed_statements[0]
    params = list(stmt.compile().params.values())
    assert actor.tenant_id not in params


def test_multiple_ids_narrow_the_query_and_two_orders_appear():
    order_a = make_order(
        status=ServiceStatus.received,
        service_type=ServiceType.regular,
        created_at=datetime(2026, 1, 1),
    )
    order_b = make_order(
        status=ServiceStatus.delivered,
        service_type=ServiceType.km_review,
        created_at=datetime(2026, 1, 10),
        delivered_at=datetime(2026, 1, 12),
    )
    main_rows = [
        (order_a, "AAA111", "Centro Norte", "Medellín", 5000, "XPEED 125"),
        (order_b, "BBB222", "Centro Sur", "Cali", 8000, "ROCKVILLE 200"),
    ]
    agg_rows = [
        _AggRow(plate="AAA111", total_visits=1, recent_visits=1, warranty_count=0),
        _AggRow(plate="BBB222", total_visits=2, recent_visits=0, warranty_count=1),
    ]

    fake_db = FakeExportSession(queue=[main_rows, agg_rows])

    with make_test_client(current_user=make_user("superadmin"), fake_db_session=fake_db) as client:
        response = client.post(
            "/api/v1/orders/analytics/services/export",
            json={"order_ids": [str(order_a.id), str(order_b.id)]},
        )

    assert response.status_code == 200
    rows = _read_rows(response.content)
    assert len(rows) == 3  # header + 2 data rows
    order_ids_in_export = {rows[1][0], rows[2][0]}
    assert order_ids_in_export == {str(order_a.id), str(order_b.id)}
