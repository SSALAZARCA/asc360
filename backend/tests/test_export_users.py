"""
Tests for splitting "Personal y Acceso" into two Excel exports
(`GET /users/export?scope=clients|staff`), scoped identically to the
existing `GET /users` list endpoint.

Two things matter here:
1. `_scoped_users_query` must produce the EXACT same tenant-visibility
   filter `list_users` already uses -- a non-superadmin's export must
   never leak another tenant's users. Tested by inspecting the returned
   `Select` statement's `whereclause` directly (a fake execute-mock would
   silently ignore any WHERE clause and defeat the point of this test, per
   the same lesson learned earlier today in
   `test_client_geo_normalization.py`/`FakeDeliverySession`).
2. `_build_users_workbook` (the pure Excel-building function, split out of
   `export_users` specifically so it's testable without a DB/HTTP layer)
   produces the right headers/rows per scope.
"""
import uuid
from datetime import date
from unittest.mock import MagicMock

import openpyxl

from app.api.v1.endpoints.users import _scoped_users_query, _build_users_workbook
from app.models.user import Role, UserStatus


def _superadmin() -> MagicMock:
    u = MagicMock()
    u.is_superadmin = True
    u.tenant_id = None
    return u


def _tenant_admin(tenant_id=None) -> MagicMock:
    u = MagicMock()
    u.is_superadmin = False
    u.tenant_id = tenant_id or uuid.uuid4()
    return u


def _client_user(**overrides) -> MagicMock:
    u = MagicMock()
    u.name = "Cliente Uno"
    u.identification = "111222333"
    u.phone = "3001234567"
    u.email = "cliente@example.com"
    u.birth_date = date(1990, 5, 1)
    u.city = "Medellín"
    u.department = "Antioquia"
    u.address = "Cra 1 # 2-3"
    u.status = UserStatus.active
    u.telegram_id = "tg-1"
    for k, v in overrides.items():
        setattr(u, k, v)
    return u


def _staff_user(**overrides) -> MagicMock:
    u = MagicMock()
    u.name = "Técnico Uno"
    u.role = Role.technician
    u.email = "tecnico@example.com"
    u.phone = "3009876543"
    u.service_center_name = None
    u.tenant = MagicMock(name="Taller Centro")
    u.tenant.name = "Taller Centro"
    u.status = UserStatus.pending
    u.telegram_id = None
    for k, v in overrides.items():
        setattr(u, k, v)
    return u


# ---------------------------------------------------------------------------
# Security-critical: tenant-visibility scoping (must match `list_users`)
# ---------------------------------------------------------------------------

def test_scoped_users_query_superadmin_has_no_tenant_filter():
    stmt = _scoped_users_query(_superadmin())
    assert stmt.whereclause is None


def test_scoped_users_query_non_superadmin_filters_by_own_tenant():
    tid = uuid.uuid4()
    stmt = _scoped_users_query(_tenant_admin(tid))
    assert stmt.whereclause is not None
    compiled = str(stmt.whereclause.compile(compile_kwargs={"literal_binds": True}))
    assert "tenant_id" in compiled
    assert tid.hex in compiled.replace("-", "")


# ---------------------------------------------------------------------------
# Workbook content
# ---------------------------------------------------------------------------

def test_build_users_workbook_clients_scope_headers_and_row():
    buf = _build_users_workbook([_client_user()], "clients")
    wb = openpyxl.load_workbook(buf)
    ws = wb.active

    assert ws.title == "Clientes"
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert headers == [
        "Nombre", "Cédula", "Teléfono", "Email", "Fecha de nacimiento",
        "Ciudad", "Departamento", "Dirección", "Estado", "Telegram Vinculado",
    ]
    row = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
    assert row == [
        "Cliente Uno", "111222333", "3001234567", "cliente@example.com",
        "1990-05-01", "Medellín", "Antioquia", "Cra 1 # 2-3", "Activo", "Sí",
    ]


def test_build_users_workbook_staff_scope_headers_and_row():
    buf = _build_users_workbook([_staff_user()], "staff")
    wb = openpyxl.load_workbook(buf)
    ws = wb.active

    assert ws.title == "Usuarios del Sistema"
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert headers == ["Nombre", "Rol", "Email", "Teléfono", "Taller Asignado", "Estado", "Telegram Vinculado"]
    row = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
    assert row == [
        "Técnico Uno", "Técnico", "tecnico@example.com", "3009876543",
        "Taller Centro", "Pendiente", "No",
    ]


def test_build_users_workbook_staff_scope_falls_back_to_acceso_global():
    u = _staff_user(service_center_name=None, tenant=None, role=Role.superadmin)
    buf = _build_users_workbook([u], "staff")
    wb = openpyxl.load_workbook(buf)
    ws = wb.active
    row = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
    assert row[1] == "Super Admin"
    assert row[4] == "Acceso Global"


def test_build_users_workbook_empty_list_still_has_headers():
    buf = _build_users_workbook([], "clients")
    wb = openpyxl.load_workbook(buf)
    ws = wb.active
    assert ws.max_row == 1
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert headers[0] == "Nombre"
