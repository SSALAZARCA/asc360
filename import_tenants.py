"""
Script de importación de Centros de Servicio / Distribuidoras
Lee un archivo Excel y carga los registros en la tabla `tenants` de PostgreSQL.

Uso:
    python import_tenants.py centros_de_servicio.xlsx

Dependencias:
    pip install openpyxl asyncpg sqlalchemy[asyncio]
"""

import asyncio
import sys
import uuid
from pathlib import Path

import openpyxl
import asyncpg

# ─── CONFIGURACIÓN ────────────────────────────────────────────────────────────
# Ajusta el connection string si tu entorno es diferente
DATABASE_URL = "postgresql://umadmin:umadmin123@localhost:5432/um_service_db"

TENANT_TYPE_MAP = {
    "Centro de Servicio": "service_center",
    "Distribuidor / Repuestero": "parts_dealer",
    "service_center": "service_center",
    "parts_dealer": "parts_dealer",
}
# ──────────────────────────────────────────────────────────────────────────────


async def import_tenants(excel_path: str):
    conn = await asyncpg.connect(DATABASE_URL)
    
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    print(f"📋 Columnas detectadas: {headers}")
    print(f"📦 Filas a procesar: {ws.max_row - 1}\n")
    
    inserted = 0
    skipped = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        
        name = str(row_dict.get("nombre", "") or "").strip()
        nit = str(row_dict.get("nit", "") or "").strip() or None
        phone = str(row_dict.get("telefono", "") or "").strip() or None
        tipo_raw = str(row_dict.get("tipo", "Centro de Servicio") or "Centro de Servicio").strip()
        subdomain_raw = str(row_dict.get("subdominio", "") or "").strip()
        
        if not name:
            print(f"   ⚠️  Fila ignorada (nombre vacío)")
            skipped += 1
            continue
        
        tenant_type = TENANT_TYPE_MAP.get(tipo_raw, "service_center")
        
        # Generar subdominio automático si no viene
        if not subdomain_raw:
            subdomain_raw = name.lower().replace(" ", "-").replace(".", "")[:50]
        
        # Verificar si ya existe por NIT o nombre
        existing = await conn.fetchrow(
            "SELECT id FROM tenants WHERE nit = $1 OR name = $2", nit, name
        )
        
        if existing:
            print(f"   ⏭️  Ya existe: {name} (id: {existing['id']}) — saltando")
            skipped += 1
            continue
        
        # Asegura subdominio único
        base_sub = subdomain_raw
        counter = 1
        while await conn.fetchrow("SELECT id FROM tenants WHERE subdomain = $1", subdomain_raw):
            subdomain_raw = f"{base_sub}-{counter}"
            counter += 1
        
        new_id = uuid.uuid4()
        await conn.execute(
            """
            INSERT INTO tenants (id, name, subdomain, nit, phone, tenant_type, config)
            VALUES ($1, $2, $3, $4, $5, $6::tenanttype, $7::jsonb)
            """,
            new_id, name, subdomain_raw, nit, phone, tenant_type, "{}"
        )
        print(f"   ✅ Insertado: {name} ({tenant_type}) — {nit or 'sin NIT'}")
        inserted += 1
    
    await conn.close()
    print(f"\n{'─'*50}")
    print(f"✅ Insertados: {inserted} | ⏭️  Saltados: {skipped}")
    print(f"{'─'*50}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python import_tenants.py centros_de_servicio.xlsx")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    if not Path(excel_file).exists():
        print(f"❌ Archivo no encontrado: {excel_file}")
        sys.exit(1)
    
    asyncio.run(import_tenants(excel_file))
